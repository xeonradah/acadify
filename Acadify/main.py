"""
Acadify - Academic Record Management System
A comprehensive web-based academic record management system for Norzagaray College

Environment Variables Setup:
- MYSQL_HOST: Database host (default: localhost)
- MYSQL_PORT: Database port (default: 3306)
- MYSQL_USERNAME: Database username (default: root)
- MYSQL_PASSWORD: Database password (default: 102503 - CHANGE IN PRODUCTION!)
- MYSQL_DATABASE: Database name (default: acadify_main)
- SESSION_SECRET: Flask secret key (default: acadify-secret-key-2025)

For production, set these environment variables or create a .env file
"""

from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, send_file, after_this_request
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from sqlalchemy import text
import pymysql
import os
from datetime import datetime

# Initialize Flask application
app = Flask(__name__)

# Configuration
app.config['SECRET_KEY'] = os.environ.get('SESSION_SECRET', 'acadify-secret-key-2025')

# Database Configuration - Use environment variables for security
DB_HOST = os.environ.get('MYSQL_HOST', 'mathtry-db.c9aqi8mg6z1y.ap-southeast-2.rds.amazonaws.com')
DB_PORT = os.environ.get('MYSQL_PORT', '3306')
DB_USERNAME = os.environ.get('MYSQL_USERNAME', 'admin')
DB_PASSWORD = os.environ.get('MYSQL_PASSWORD', 'mathtry123')  # Default for development only
DB_NAME = os.environ.get('MYSQL_DATABASE', 'acadify')

app.config['SQLALCHEMY_DATABASE_URI'] = f'mysql+pymysql://{DB_USERNAME}:{DB_PASSWORD}@{DB_HOST}:{DB_PORT}/{DB_NAME}'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Initialize extensions
db = SQLAlchemy(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'  # type: ignore
login_manager.login_message = 'Please log in to access this page.'

# =====================================
# DATABASE MODELS
# =====================================

class User(UserMixin, db.Model):
    """User model for authentication and basic information - CLEAN VERSION"""
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    role = db.Column(db.String(20), nullable=False)  # student, instructor, registrar, dean, mis_it
    first_name = db.Column(db.String(50), nullable=False)
    middle_name = db.Column(db.String(50), nullable=True)
    last_name = db.Column(db.String(50), nullable=False)
    suffix = db.Column(db.String(50), nullable=True)
    
    # Common Fields (Applicable to All Roles)
    department = db.Column(db.String(10), nullable=True)  # For instructors/staff
    mobile_no = db.Column(db.String(20), nullable=True)
    active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    theme_preference = db.Column(db.String(10), nullable=True, default='light')
    
    def set_password(self, password):
        self.password_hash = generate_password_hash(password)
    
    def check_password(self, password):
        return check_password_hash(self.password_hash, password)
    
    def get_id(self):
        """Return user ID with prefix for Flask-Login"""
        return f"user_{self.id}"
    
    # Note: Students are now completely separate from Users
    # Users are only for staff: Dean, Instructor, MISIT, Registrar

class Student(UserMixin, db.Model):
    """Student model - COMPLETELY INDEPENDENT from User table"""
    __tablename__ = 'students'
    
    id = db.Column(db.Integer, primary_key=True)
    
    # Authentication (Independent from User table)
    username = db.Column(db.String(80), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    
    # Student Identity
    student_id = db.Column(db.String(20), unique=True, nullable=False)
    student_lrn = db.Column(db.String(20), nullable=True)
    student_status = db.Column(db.String(10), default='Regular')
    
    # Basic Information
    first_name = db.Column(db.String(50), nullable=False)
    middle_name = db.Column(db.String(50), nullable=True)
    last_name = db.Column(db.String(50), nullable=False)
    suffix = db.Column(db.String(50), nullable=True)
    
    # Academic Information
    department = db.Column(db.String(50), nullable=False)
    course = db.Column(db.String(100), nullable=True)
    year_level = db.Column(db.Integer, nullable=False)
    semester = db.Column(db.Integer, nullable=False)
    section = db.Column(db.String(10), nullable=True)
    section_type = db.Column(db.String(50), nullable=True)
    academic_year = db.Column(db.String(10), nullable=False, default='2025-2026')
    curriculum = db.Column(db.String(50), nullable=True)
    graduating = db.Column(db.String(10), default='No')
    
    # Personal Information
    gender = db.Column(db.String(10), nullable=True)
    age = db.Column(db.Integer, nullable=True)
    date_birth = db.Column(db.String(20), nullable=True)
    place_birth = db.Column(db.String(50), nullable=True)
    nationality = db.Column(db.String(20), default='Filipino')
    religion = db.Column(db.String(50), nullable=True)
    
    # Address Information
    province = db.Column(db.String(100), nullable=True)
    city_municipality = db.Column(db.String(100), nullable=True)
    barangay = db.Column(db.String(50), nullable=True)
    house_no = db.Column(db.String(50), nullable=True)
    
    # Contact Information
    mobile_no = db.Column(db.String(20), nullable=True)
    
    # Status and Tracking
    enrollment_status = db.Column(db.String(20), default='PENDING')  # PENDING, ENROLLED, DROPPED
    total_units = db.Column(db.Integer, default=0)
    gwa = db.Column(db.Float, nullable=True)
    active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    # Authentication Methods (Same as User model)
    def set_password(self, password):
        self.password_hash = generate_password_hash(password)
    
    def check_password(self, password):
        return check_password_hash(self.password_hash, password)
    
    def get_id(self):
        """Return student ID with prefix for Flask-Login"""
        return f"student_{self.id}"
    
    # Role property for compatibility
    @property
    def role(self):
        return 'student'

# =====================================
# Account Creation
# =====================================

@app.route('/create-account')
@login_required
def create_account_page():
    """Render the create account page for STAFF only"""
    if current_user.role != 'mis_it':
        flash('Access denied. Only MIS/IT administrators can create accounts.')
        return redirect(url_for('dashboard'))
    return render_template('auth/create-account.html')

@app.route('/create_account', methods=['POST'])
@login_required
def create_account():
    """Create a new STAFF account (Dean, Instructor, MISIT, Registrar) - MIS/IT only"""
    # Verify that the current user is MIS/IT
    if current_user.role != 'mis_it':
        return jsonify({'status': 'error', 'message': 'Unauthorized access'}), 403
    
    try:
        # Get form data
        data = request.form
        username = data.get('username')
        email = data.get('email')
        password = data.get('password')
        role = data.get('role')
        first_name = data.get('first_name')
        last_name = data.get('last_name')
        
        # Validate required fields
        if not all([username, email, password, role, first_name, last_name]):
            return jsonify({'status': 'error', 'message': 'Missing required fields'}), 400
            
        # Check if username or email already exists
        if User.query.filter_by(username=username).first():
            return jsonify({'status': 'error', 'message': 'Username already exists'}), 400
        if User.query.filter_by(email=email).first():
            return jsonify({'status': 'error', 'message': 'Email already exists'}), 400
        
        # Initialize user data dictionary with common fields
        user_data = {
            'username': username,
            'email': email,
            'role': role,
            'first_name': first_name,
            'last_name': last_name,
            'active': True
        }
        
        # Optional personal details (only fields that exist in User table)
        optional_fields = [
            'middle_name', 'suffix', 'mobile_no'
        ]
        for field in optional_fields:
            value = data.get(field)
            if value is not None and value != '':
                user_data[field] = value

        # Validate role (only staff roles allowed)
        valid_staff_roles = ['instructor', 'registrar', 'dean', 'mis_it']
        if role not in valid_staff_roles:
            return jsonify({'status': 'error', 'message': f'Invalid role. Only staff roles allowed: {valid_staff_roles}'}), 400
        
        # Add role-specific fields
        if role == 'instructor':
            # Only instructors need department in user table
            department = data.get('department')
            if not department:
                return jsonify({'status': 'error', 'message': 'Department is required for instructors'}), 400
            user_data['department'] = department
        
        # Create new user with only the necessary fields
        new_user = User(**user_data)
        new_user.set_password(password)
        
        # Save to database
        db.session.add(new_user)
        db.session.commit()
        
        # Log account creation
        log_audit_action(
            user_id=current_user.id,
            action='create_account',
            resource_type='user',
            resource_id=new_user.id,
            description=f"Created new {role} account for {new_user.username} ({new_user.first_name} {new_user.last_name})",
            status='success',
            request=request
        )
        
        return jsonify({
            'status': 'success',
            'message': 'Account created successfully' + (' and student record added' if role == 'student' else ''),
            'user': {
                'id': new_user.id,
                'username': new_user.username,
                'email': new_user.email,
                'role': new_user.role
            }
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/misit/create-student-account')
@login_required
def misit_create_student_account_page():
    """Render the student account creation page (MIS/IT only)"""
    # Verify that the current user is MIS/IT
    if current_user.role != 'mis_it':
        flash('Unauthorized access', 'error')
        return redirect(url_for('login'))
    return render_template('misit/misit_create_student.html')

@app.route('/create_student_account', methods=['POST'])
def create_student_account():
    """Create a new STUDENT account (Public access - Self registration) - DEPRECATED"""
    try:
        # Get form data
        data = request.form
        username = data.get('username')
        email = data.get('email')
        password = data.get('password')
        student_id = data.get('student_id')
        first_name = data.get('first_name')
        last_name = data.get('last_name')
        
        # Validate required fields
        if not all([username, email, password, student_id, first_name, last_name]):
            return jsonify({'status': 'error', 'message': 'Missing required fields'}), 400
            
        # Check if username, email, or student_id already exists
        if Student.query.filter_by(username=username).first():
            return jsonify({'status': 'error', 'message': 'Username already exists'}), 400
        if Student.query.filter_by(email=email).first():
            return jsonify({'status': 'error', 'message': 'Email already exists'}), 400
        if Student.query.filter_by(student_id=student_id).first():
            return jsonify({'status': 'error', 'message': 'Student ID already exists'}), 400
        
        # Get additional required fields
        department = data.get('department')
        
        if not department:
            return jsonify({'status': 'error', 'message': 'Missing required department field'}), 400
        
        # Create new student record (completely independent)
        new_student = Student(
            # Authentication
            username=username,
            email=email,
            
            # Basic Information
            first_name=first_name,
            middle_name=data.get('middle_name'),
            last_name=last_name,
            suffix=data.get('suffix'),
            
            # Student Identity
            student_id=student_id,
            student_lrn=data.get('student_lrn'),
            student_status=data.get('student_status', 'Regular'),
            
            # Academic Information (Basic only - detailed enrollment handled by Registrar)
            department=department,
            course=data.get('course'),
            year_level=1,  # Default - will be updated during enrollment
            semester=1,    # Default - will be updated during enrollment
            section=None,  # Will be set during enrollment
            section_type=data.get('section_type'),
            academic_year='2025-2026',  # Default - will be updated during enrollment
            curriculum=None,  # Will be set during enrollment
            graduating='No',
            
            # Personal Information
            gender=data.get('gender'),
            age=int(data.get('age')) if data.get('age') else None,
            date_birth=data.get('date_birth'),
            place_birth=data.get('place_birth'),
            nationality=data.get('nationality', 'Filipino'),
            religion=data.get('religion'),
            
            # Address Information
            province=data.get('province'),
            city_municipality=data.get('city_municipality'),
            barangay=data.get('barangay'),
            house_no=data.get('house_no'),
            
            # Contact Information
            mobile_no=data.get('mobile_no'),
            
            # Status
            enrollment_status='PENDING',
            active=True
        )
        new_student.set_password(password)
        
        # Save to database
        db.session.add(new_student)
        db.session.commit()
        
        # Log account creation
        log_audit_action(
            user_id=None,  # No current user for self-registration
            action='create_student_account',
            description=f"Student account created: {username} ({student_id})",
            status='success',
            request=request
        )
        
        return jsonify({
            'status': 'success', 
            'message': f'Student account created successfully for {first_name} {last_name}',
            'redirect': '/login'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/misit/create_student_account', methods=['POST'])
@login_required
def misit_create_student_account():
    """Create a new STUDENT account (MIS/IT only)"""
    # Verify that the current user is MIS/IT
    if current_user.role != 'mis_it':
        return jsonify({'status': 'error', 'message': 'Unauthorized access'}), 403
    
    try:
        # Get form data
        data = request.form
        username = data.get('username')
        email = data.get('email')
        password = data.get('password')
        student_id = data.get('student_id')
        first_name = data.get('first_name')
        last_name = data.get('last_name')
        
        # Validate required fields
        if not all([username, email, password, student_id, first_name, last_name]):
            return jsonify({'status': 'error', 'message': 'Missing required fields'}), 400
            
        # Check if username, email, or student_id already exists
        if Student.query.filter_by(username=username).first():
            return jsonify({'status': 'error', 'message': 'Username already exists'}), 400
        if Student.query.filter_by(email=email).first():
            return jsonify({'status': 'error', 'message': 'Email already exists'}), 400
        if Student.query.filter_by(student_id=student_id).first():
            return jsonify({'status': 'error', 'message': 'Student ID already exists'}), 400
        
        # Get additional required fields
        department = data.get('department')
        
        if not department:
            return jsonify({'status': 'error', 'message': 'Missing required department field'}), 400
        
        # Create new student record (completely independent)
        new_student = Student(
            # Authentication
            username=username,
            email=email,
            
            # Basic Information
            first_name=first_name,
            middle_name=data.get('middle_name'),
            last_name=last_name,
            suffix=data.get('suffix'),
            
            # Student Identity
            student_id=student_id,
            student_lrn=data.get('student_lrn'),
            student_status=data.get('student_status', 'Regular'),
            
            # Academic Information (Basic only - detailed enrollment handled by Registrar)
            department=department,
            course=data.get('course'),
            year_level=1,  # Default - will be updated during enrollment
            semester=1,    # Default - will be updated during enrollment
            section=None,  # Will be set during enrollment
            section_type=data.get('section_type'),
            academic_year='2025-2026',  # Default - will be updated during enrollment
            curriculum=None,  # Will be set during enrollment
            graduating='No',
            
            # Personal Information
            gender=data.get('gender'),
            age=int(data.get('age')) if data.get('age') else None,
            date_birth=data.get('date_birth'),
            place_birth=data.get('place_birth'),
            nationality=data.get('nationality', 'Filipino'),
            religion=data.get('religion'),
            
            # Address Information
            province=data.get('province'),
            city_municipality=data.get('city_municipality'),
            barangay=data.get('barangay'),
            house_no=data.get('house_no'),
            
            # Contact Information
            mobile_no=data.get('mobile_no'),
            
            # Status
            enrollment_status='PENDING',
            active=True
        )
        new_student.set_password(password)
        
        # Save to database
        db.session.add(new_student)
        db.session.commit()
        
        # Log account creation
        log_audit_action(
            user_id=current_user.id,
            action='create_student_account',
            resource_type='Student',
            resource_id=new_student.id,
            description=f"Student account created by MIS/IT: {username} ({student_id})",
            status='success',
            request=request
        )
        
        return jsonify({
            'status': 'success', 
            'message': f'Student account created successfully for {first_name} {last_name}'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500

# =====================================
# DATABASE MODELS (continued)
# =====================================

class GradeEncodingSchedule(db.Model):
    """Grade Encoding Schedule model for managing grade submission periods"""
    id = db.Column(db.Integer, primary_key=True)
    academic_year = db.Column(db.String(10), nullable=False)
    semester = db.Column(db.Integer, nullable=False)
    start_date = db.Column(db.DateTime, nullable=False)
    end_date = db.Column(db.DateTime, nullable=False)
    start_time = db.Column(db.Time, nullable=True)  # For testing purposes
    end_time = db.Column(db.Time, nullable=True)  # For testing purposes
    department = db.Column(db.String(10), nullable=True)  # Optional, if schedule varies by department
    grading_period = db.Column(db.String(20), nullable=False, default='all')  # all, prelim, midterm, final
    status = db.Column(db.String(20), nullable=False, default='upcoming')  # upcoming, active, completed
    created_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

class Subject(db.Model):
    """Subject model for course management - Core subject information only"""
    id = db.Column(db.Integer, primary_key=True)
    subject_code = db.Column(db.String(20), unique=True, nullable=False)
    subject_name = db.Column(db.String(100), nullable=False)
    subject_type = db.Column(db.String(20), nullable=False, default='Academic')
    units = db.Column(db.Integer, nullable=False)
    department = db.Column(db.String(10), nullable=False)
    year_level = db.Column(db.Integer, nullable=False)
    semester = db.Column(db.Integer, nullable=False)
    section = db.Column(db.String(10), nullable=True)
    academic_year = db.Column(db.String(10), nullable=False, default='2025-2026')
    instructor_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    max_capacity = db.Column(db.Integer, nullable=True, default=50)
    
    # Relationships
    instructor = db.relationship('User', foreign_keys=[instructor_id], backref='subjects')
    class_assignments = db.relationship('ClassAssignment', backref='subject', lazy=True, cascade='all, delete-orphan')

class ClassAssignment(db.Model):
    """Class assignment model for section-based instructor assignments"""
    id = db.Column(db.Integer, primary_key=True)
    subject_id = db.Column(db.Integer, db.ForeignKey('subject.id'), nullable=False)
    instructor_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    school_year = db.Column(db.String(20), nullable=False, default='2025-2026')
    semester = db.Column(db.Integer, nullable=False)
    section = db.Column(db.String(10), nullable=True)
    schedule_time = db.Column(db.String(50), nullable=True)  # Format: "8:00 AM - 11:00 AM"
    schedule_day = db.Column(db.String(20), nullable=True)    # Monday, Tuesday, etc.
    room = db.Column(db.String(20), nullable=True)          # Room number/name
    subject_type = db.Column(db.String(20), nullable=True, default='Lecture')  # Lecture or Laboratory
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    # Relationships
    instructor = db.relationship('User', backref='class_assignments')
    
    # Unique constraint to prevent duplicate assignments for the same subject-section combination
    # This allows the same instructor to teach the same subject in different sections
    __table_args__ = (
        db.UniqueConstraint('subject_id', 'school_year', 'semester', 'section', name='unique_subject_section_assignment'),
    )
    
class Enrollment(db.Model):
    """Enrollment model for tracking student course enrollments"""
    __tablename__ = 'enrollment'
    
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.Integer, db.ForeignKey('students.id'), nullable=False)
    subject_id = db.Column(db.Integer, db.ForeignKey('subject.id'), nullable=False)
    academic_year = db.Column(db.String(10), nullable=False)
    semester = db.Column(db.Integer, nullable=False)
    enrollment_date = db.Column(db.DateTime, default=datetime.utcnow)
    status = db.Column(db.String(20), default='Active')  # Active, Dropped, Completed
    enrolled_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    dropped_date = db.Column(db.DateTime, nullable=True)
    completion_date = db.Column(db.DateTime, nullable=True)
    notes = db.Column(db.Text, nullable=True)
    
    # Relationships
    student = db.relationship('Student', foreign_keys=[student_id], backref='enrollments')
    subject = db.relationship('Subject', backref='enrollments')
    enrolled_by_user = db.relationship('User', foreign_keys=[enrolled_by], backref='enrolled_students')
    
    # Unique constraint
    __table_args__ = (
        db.UniqueConstraint('student_id', 'subject_id', 'academic_year', 'semester', name='unique_enrollment'),
    )

class Grade(db.Model):
    """Grade model for academic records"""
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.Integer, db.ForeignKey('students.id'), nullable=False)
    subject_id = db.Column(db.Integer, db.ForeignKey('subject.id'), nullable=False)
    prelim_grade = db.Column(db.Float, nullable=True)
    midterm_grade = db.Column(db.Float, nullable=True)
    final_grade = db.Column(db.Float, nullable=True)
    final_average = db.Column(db.Float, nullable=True)
    equivalent_grade = db.Column(db.Float, nullable=True)  # GPA equivalent (1.00-5.00)
    remarks = db.Column(db.String(20), nullable=True)  # Excellent, Outstanding, etc.
    semester = db.Column(db.Integer, nullable=False)
    academic_year = db.Column(db.String(10), nullable=False)
    is_locked = db.Column(db.Boolean, default=False)
    is_complete = db.Column(db.Boolean, default=False)  # True when all three grades are entered
    is_historical = db.Column(db.Boolean, default=False)  # For imported historical grades
    import_date = db.Column(db.DateTime, nullable=True)  # When it was imported
    import_source = db.Column(db.String(50), nullable=True)  # Excel, CSV, Manual, etc.
    submitted_at = db.Column(db.DateTime, nullable=True)
    approved_at = db.Column(db.DateTime, nullable=True)
    approved_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    
    # Relationships
    student = db.relationship('Student', foreign_keys=[student_id], backref='student_grades')
    subject = db.relationship('Subject', backref='grades')
    approver = db.relationship('User', foreign_keys=[approved_by], backref='approved_grades')

class DeansListRecord(db.Model):
    """Dean's List records for academic achievers"""
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.Integer, db.ForeignKey('students.id'), nullable=False)
    semester = db.Column(db.Integer, nullable=False)
    academic_year = db.Column(db.String(10), nullable=False)
    gwa = db.Column(db.Float, nullable=False)  # General Weighted Average
    rank = db.Column(db.Integer, nullable=True)
    total_units = db.Column(db.Integer, nullable=False)
    qualified = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

class Notification(db.Model):
    """Notification system for user communications"""
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    type = db.Column(db.String(50), default='info')
    title = db.Column(db.String(255), nullable=False)
    message = db.Column(db.Text, nullable=False)
    is_read = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    # Relationships
    user = db.relationship('User', backref='notifications')

class AuditLog(db.Model):
    """Audit log model for tracking system activities"""
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)  # Allow null for system actions
    action = db.Column(db.String(50), nullable=False)  # 'login', 'logout', 'create_account', 'update_grade', etc.
    resource_type = db.Column(db.String(50), nullable=True)  # 'user', 'grade', 'subject', 'schedule', etc.
    resource_id = db.Column(db.Integer, nullable=True)  # ID of the affected resource
    description = db.Column(db.Text, nullable=False)  # Detailed description of the action
    ip_address = db.Column(db.String(45), nullable=True)  # IPv4 or IPv6 address
    user_agent = db.Column(db.Text, nullable=True)  # Browser/client information
    status = db.Column(db.String(20), nullable=False, default='success')  # 'success', 'failed', 'error'
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    # Relationships
    user = db.relationship('User', backref='audit_logs')

class EncodingException(db.Model):
    """Grade encoding exception model for granting access after schedule closure"""
    id = db.Column(db.Integer, primary_key=True)
    instructor_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    academic_year = db.Column(db.String(10), nullable=False)
    semester = db.Column(db.Integer, nullable=False)
    grading_period = db.Column(db.String(20), nullable=False)  # 'all', 'prelim', 'midterm', 'final'
    expiration_date = db.Column(db.DateTime, nullable=False)
    reason = db.Column(db.Text, nullable=True)
    granted_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    revoked_at = db.Column(db.DateTime, nullable=True)
    
    # Relationships
    instructor = db.relationship('User', foreign_keys=[instructor_id], backref='encoding_exceptions')
    granted_by_user = db.relationship('User', foreign_keys=[granted_by], backref='granted_exceptions')

class StudentEnrollment(db.Model):
    """Model for tracking student enrollment per academic year/semester"""
    __tablename__ = 'student_enrollments'
    
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.Integer, db.ForeignKey('students.id'), nullable=False)
    academic_year = db.Column(db.String(20), nullable=False)  # Increased from 9 to 20
    semester = db.Column(db.Integer, nullable=False)
    year_level = db.Column(db.Integer, nullable=False)
    section = db.Column(db.String(10), nullable=True)
    curriculum = db.Column(db.String(20), nullable=True)  # Increased from 9 to 20
    enrollment_date = db.Column(db.DateTime, default=datetime.utcnow)
    status = db.Column(db.String(20), default='ACTIVE')  # ACTIVE, DROPPED, COMPLETED
    enrolled_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    # Relationships
    student = db.relationship('Student', backref='student_enrollments')
    enrolled_by_user = db.relationship('User', foreign_keys=[enrolled_by], backref='student_enrollment_records')
    
    # Unique constraint
    __table_args__ = (
        db.UniqueConstraint('student_id', 'academic_year', 'semester', name='unique_student_enrollment'),
    )

class Section(db.Model):
    """Model for storing available sections"""
    __tablename__ = 'sections'
    
    id = db.Column(db.Integer, primary_key=True)
    section_name = db.Column(db.String(10), unique=True, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    created_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    
    # Relationships
    creator = db.relationship('User', backref='created_sections')

class Curriculum(db.Model):
    """Model for storing available curricula"""
    __tablename__ = 'curricula'
    
    id = db.Column(db.Integer, primary_key=True)
    curriculum_name = db.Column(db.String(20), unique=True, nullable=False)
    curriculum_description = db.Column(db.String(100), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    created_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    
    # Relationships
    creator = db.relationship('User', backref='created_curricula')

class AcademicYear(db.Model):
    """Model for storing available academic years"""
    __tablename__ = 'academic_years'
    
    id = db.Column(db.Integer, primary_key=True)
    academic_year_name = db.Column(db.String(20), unique=True, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    created_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    
    # Relationships
    creator = db.relationship('User', backref='created_academic_years')

class StudentSubject(db.Model):
    """Model for many-to-many relationship between students and subjects"""
    __tablename__ = 'student_subjects'
    
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.Integer, db.ForeignKey('students.id'), nullable=False)
    subject_id = db.Column(db.Integer, db.ForeignKey('subject.id'), nullable=False)
    academic_year = db.Column(db.String(20), nullable=False)  # Increased from 9 to 20
    semester = db.Column(db.Integer, nullable=False)
    enrollment_date = db.Column(db.DateTime, default=datetime.utcnow)
    status = db.Column(db.String(20), default='ENROLLED')  # ENROLLED, DROPPED, COMPLETED
    enrolled_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    # Relationships
    student = db.relationship('Student', backref='student_subject_assignments')
    subject = db.relationship('Subject', backref='student_subject_assignments')
    enrolled_by_user = db.relationship('User', foreign_keys=[enrolled_by], backref='student_subject_assignments')
    
    # Unique constraint
    __table_args__ = (
        db.UniqueConstraint('student_id', 'subject_id', 'academic_year', 'semester', name='unique_student_subject'),
    )


# =====================================
# LOGIN MANAGER
# =====================================

@login_manager.user_loader
def load_user(user_id):
    """Load user for Flask-Login - handles both User and Student models"""
    # Format: "user_123" for User table, "student_456" for Student table
    if user_id.startswith('user_'):
        actual_id = int(user_id.replace('user_', ''))
        return User.query.get(actual_id)
    elif user_id.startswith('student_'):
        actual_id = int(user_id.replace('student_', ''))
        return Student.query.get(actual_id)
    else:
        # Backward compatibility - assume it's a User ID
        return User.query.get(int(user_id))

# =====================================
# UTILITY FUNCTIONS
# =====================================

def calculate_grade_equivalent(numerical_grade):
    """Convert numerical grade to GPA equivalent based on Norzagaray College scale"""
    if numerical_grade is None:
        return None, None
    
    if numerical_grade >= 98:
        return 1.00, "Excellent"
    elif numerical_grade >= 95:
        return 1.25, "Outstanding"
    elif numerical_grade >= 92:
        return 1.50, "Superior"
    elif numerical_grade >= 89:
        return 1.75, "Very Good"
    elif numerical_grade >= 86:
        return 2.00, "Good"
    elif numerical_grade >= 83:
        return 2.25, "Satisfactory"
    elif numerical_grade >= 80:
        return 2.50, "Fairly Satisfactory"
    elif numerical_grade >= 76:
        return 2.75, "Fair"
    elif numerical_grade >= 75:
        return 3.00, "Passed"
    else:
        return 5.00, "Failed"

def calculate_final_grade(prelim, midterm, final):
    """Calculate final average from three grading periods"""
    if prelim is None or midterm is None or final is None:
        return None
    return round((prelim + midterm + final) / 3, 2)

def create_notification(user_id, notification_type, title, message):
    """Create a new notification for a user"""
    try:
        notification = Notification(
            user_id=user_id,
            type=notification_type,
            title=title,
            message=message
        )
        db.session.add(notification)
        db.session.commit()
        return notification
    except Exception as e:
        db.session.rollback()
        print(f"Error creating notification: {e}")
        return None

def get_unread_notifications_count(user_id):
    """Get count of unread notifications for a user"""
    return Notification.query.filter_by(user_id=user_id, is_read=False).count()

def log_audit_action(user_id, action, resource_type=None, resource_id=None, description="", status="success", request=None):
    """Create an audit log entry for tracking user actions"""
    try:
        ip_address = None
        user_agent = None
        
        if request:
            ip_address = request.environ.get('HTTP_X_FORWARDED_FOR', request.environ.get('REMOTE_ADDR'))
            user_agent = request.headers.get('User-Agent')
        
        audit_log = AuditLog(
            user_id=user_id,
            action=action,
            resource_type=resource_type,
            resource_id=resource_id,
            description=description,
            ip_address=ip_address,
            user_agent=user_agent,
            status=status
        )
        
        db.session.add(audit_log)
        db.session.commit()
        return True
    except Exception as e:
        print(f"Error creating audit log: {e}")
        db.session.rollback()
        return False

def mark_notification_as_read(notification_id):
    """Mark a notification as read"""
    try:
        notification = Notification.query.get(notification_id)
        if notification:
            notification.is_read = True
            db.session.commit()
            return True
    except Exception as e:
        db.session.rollback()
        print(f"Error marking notification as read: {e}")
    return False

def check_deans_list_eligibility(student_id, semester, academic_year):
    """Check if student qualifies for Dean's List based on complete grades"""
    # Get student record to check section_type
    student = Student.query.get(student_id)
    if not student:
        return False, 0, 0, "Student not found"
    
    # Check if student has regular status (Block Section)
    if not student.section_type or student.section_type.lower() != 'block section':
        return False, 0, 0, "Not a regular student (Block Section required)"
    
    grades = Grade.query.filter_by(
        student_id=student_id,
        semester=semester,
        academic_year=academic_year
    ).all()
    
    if not grades:
        return False, 0, 0, "No grades found"
    
    # Filter for complete grades (all three components entered)
    complete_grades = [g for g in grades if g.prelim_grade is not None and g.midterm_grade is not None and g.final_grade is not None]
    
    if not complete_grades:
        return False, 0, 0, "Incomplete grades"
    
    # Calculate total units (ALL subjects - Academic + Non-Academic)
    total_all_units = 0
    for grade in complete_grades:
        subject = Subject.query.get(grade.subject_id)
        if subject:
            total_all_units += subject.units
    
    # Check minimum units requirement (18+ total units)
    if total_all_units < 18:
        return False, 0, total_all_units, f"Insufficient units ({total_all_units}/18 required)"
    
    # Calculate Academic units and GWA (Academic subjects only)
    academic_units = 0
    total_points = 0
    
    for grade in complete_grades:
        # Get subject information first
        subject = Subject.query.get(grade.subject_id)
        if not subject:
            continue
            
        # Skip Non Academic subjects - only Academic subjects count for Dean's List
        if subject.subject_type != 'Academic':
            continue
        
        # Check for failing grades (5.00)
        if not grade.equivalent_grade or grade.equivalent_grade >= 5.00:
            return False, 0, 0, "Has failing grades"
        
        # Check for grades below 2.00 (No grade below 2.00)
        if grade.equivalent_grade > 2.00:
            return False, 0, 0, "Has grades below 2.00"
        
        # Check for INC/AW/UW marks
        if grade.remarks and grade.remarks in ['INC', 'AW', 'UW']:
            return False, 0, 0, f"Has {grade.remarks} mark"
        
        # Add to calculation (only Academic subjects)
        academic_units += subject.units
        total_points += grade.equivalent_grade * subject.units
    
    gwa = total_points / academic_units if academic_units > 0 else 0
    
    # Check GWA requirement (1.75 or better)
    if gwa > 1.75:
        return False, gwa, total_all_units, f"GWA too high ({gwa:.2f} > 1.75)"
    
    return True, gwa, total_all_units, "Eligible for Dean's List"

def check_encoding_exception(instructor_id, academic_year, semester, grading_period):
    """Check if instructor has an active encoding exception for the given period"""
    now = datetime.now()
    
    # Check for specific grading period exception
    exception = EncodingException.query.filter(
        EncodingException.instructor_id == instructor_id,
        EncodingException.academic_year == academic_year,
        EncodingException.semester == semester,
        db.or_(
            EncodingException.grading_period == grading_period,
            EncodingException.grading_period == 'all'
        ),
        EncodingException.is_active == True,
        EncodingException.expiration_date >= now
    ).first()
    
    return exception is not None

def create_demo_accounts():
    """Create demo accounts for all user roles"""
    # Demo data for staff users (User table)
    demo_users = [
        {
            'username': 'instructor_demo',
            'email': 'instructor@norzagaray.edu.ph',
            'password': 'instructor123',
            'role': 'instructor',
            'first_name': 'Maria',
            'last_name': 'Santos',
            'department': 'BSCS'
        },
        {
            'username': 'registrar_demo',
            'email': 'registrar@norzagaray.edu.ph',
            'password': 'registrar123',
            'role': 'registrar',
            'first_name': 'Roberto',
            'last_name': 'Garcia',
            'department': None
        },
        {
            'username': 'dean_demo',
            'email': 'dean@norzagaray.edu.ph',
            'password': 'dean123',
            'role': 'dean',
            'first_name': 'Elizabeth',
            'last_name': 'Reyes',
            'department': None
        },
        {
            'username': 'misit_demo',
            'email': 'misit@norzagaray.edu.ph',
            'password': 'misit123',
            'role': 'mis_it',
            'first_name': 'Carlos',
            'last_name': 'Lopez',
            'department': None
        }
    ]
    
    # Demo data for students (Student table - independent)
    demo_students = [
        {
            'username': 'student_demo',
            'email': 'student@norzagaray.edu.ph',
            'password': 'student123',
            'first_name': 'Juan',
            'last_name': 'Dela Cruz',
            'student_id': '2021-00001',
            'department': 'BSCS',
            'year_level': 3,
            'semester': 1,
            'academic_year': '2025-2026',
            'student_status': 'Old',
            'enrollment_status': 'Active'
        }
    ]
    
    # Create staff users (User table)
    for user_data in demo_users:
        existing_user = User.query.filter_by(username=user_data['username']).first()
        if not existing_user:
            # Create User record (clean - no student fields)
            user = User()
            user.username = user_data['username']
            user.email = user_data['email']
            user.set_password(user_data['password'])
            user.role = user_data['role']
            user.first_name = user_data['first_name']
            user.last_name = user_data['last_name']
            user.department = user_data.get('department')
            
            db.session.add(user)
    
    # Create students (Student table - independent)
    for student_data in demo_students:
        existing_student = Student.query.filter_by(username=student_data['username']).first()
        if not existing_student:
            student = Student(
                # Authentication (independent from User table)
                username=student_data['username'],
                email=student_data['email'],
                password_hash='',  # Will be set below
                
                # Basic Information
                first_name=student_data['first_name'],
                last_name=student_data['last_name'],
                middle_name=None,
                suffix=None,
                
                # Student Identity
                student_id=student_data['student_id'],
                student_lrn=None,
                student_status=student_data['student_status'],
                
                # Academic Information
                department=student_data['department'],
                course=None,
                year_level=student_data['year_level'],
                semester=student_data['semester'],
                section=None,
                section_type=data.get('section_type'),
                academic_year=student_data['academic_year'],
                curriculum=None,
                graduating='No',
                
                # Personal Information
                gender=None,
                age=None,
                date_birth=None,
                place_birth=None,
                nationality='Filipino',
                religion=None,
                
                # Address Information
                province=None,
                city_municipality=None,
                barangay=None,
                house_no=None,
                
                # Contact Information
                mobile_no=None,
                
                # Status
                enrollment_status=student_data['enrollment_status'],
                total_units=0,
                gwa=None,
                active=True
            )
            student.set_password(student_data['password'])
            db.session.add(student)
    
    db.session.commit()

# =====================================
# ROUTES
# =====================================

@app.route('/')
def index():
    """Home page route"""
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Login page with demo accounts"""
    if request.method == 'POST':
        username = request.form.get('username') or request.form.get('id_number')
        password = request.form.get('password')

        # Try User authentication first (Staff: Dean, Instructor, MISIT, Registrar)
        user = User.query.filter(User.username == username).first()
        authenticated_user = None
        
        if user and user.check_password(password):
            authenticated_user = user
        else:
            # Try Student authentication (Students: username or student_id)
            student = Student.query.filter(
                (Student.username == username) | (Student.student_id == username)
            ).first()
            
            if student and student.check_password(password):
                authenticated_user = student

        if authenticated_user:
            login_user(authenticated_user)
            # Log successful login
            log_audit_action(
                user_id=authenticated_user.id,
                action='login',
                description=f"User {authenticated_user.username} logged in successfully",
                status='success',
                request=request
            )
            # Role-based redirect
            if authenticated_user.role == 'student':
                return redirect(url_for('student_dashboard'))
            elif authenticated_user.role == 'instructor':
                return redirect(url_for('instructor_dashboard'))
            elif authenticated_user.role == 'registrar':
                return redirect(url_for('registrar_dashboard'))
            elif authenticated_user.role == 'dean':
                return redirect(url_for('dean_dashboard'))
            elif authenticated_user.role in ['misit', 'mis_it', 'MISIT', 'MIS/IT']:
                return redirect(url_for('misit_dashboard'))
            else:
                flash('Invalid user role', 'error')
                return redirect(url_for('login'))
        else:
            # Log failed login attempt
            if username:
                log_audit_action(
                    user_id=None,
                    action='login_failed',
                    description=f"Failed login attempt for username: {username}",
                    status='failed',
                    request=request
                )
            # HTMX: return error toast if hx-request
            if request.headers.get('HX-Request'):
                return '<div id="toast-error" class="toast toast-top toast-end"><div class="alert alert-error">Invalid username or password</div></div>', 401
            flash('Invalid username or password', 'error')
    
    # Demo accounts for display
    demo_accounts = [
        {'role': 'Student', 'username': 'student_demo', 'password': 'student123'},
        {'role': 'Instructor', 'username': 'instructor_demo', 'password': 'instructor123'},
        {'role': 'Registrar', 'username': 'registrar_demo', 'password': 'registrar123'},
        {'role': 'Dean', 'username': 'dean_demo', 'password': 'dean123'},
        {'role': 'MIS/IT', 'username': 'misit_demo', 'password': 'misit123'}
    ]
    
    return render_template('auth/login.html', demo_accounts=demo_accounts)

@app.route('/logout')
@login_required
def logout():
    """Logout route"""
    # Log logout action before logging out
    log_audit_action(
        user_id=current_user.id,
        action='logout',
        description=f"User {current_user.username} logged out",
        status='success',
        request=request
    )
    logout_user()
    flash('You have been logged out successfully', 'success')
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    """Role-based dashboard routing"""
    if current_user.role == 'student':
        return redirect(url_for('student_dashboard'))
    elif current_user.role == 'instructor':
        return redirect(url_for('instructor_dashboard'))
    elif current_user.role == 'registrar':
        return redirect(url_for('registrar_dashboard'))
    elif current_user.role == 'dean':
        return redirect(url_for('dean_dashboard'))
    elif current_user.role == 'mis_it':
        return redirect(url_for('misit_dashboard'))
    else:
        flash('Invalid user role', 'error')
        return redirect(url_for('login'))

@app.route('/student/dashboard')
@login_required
def student_dashboard():
    """Student dashboard with grades and Dean's List status"""
    if current_user.role != 'student':
        flash('Access denied', 'error')
        return redirect(url_for('dashboard'))
    
    # Get student record - current_user is already a Student object for students
    if current_user.role == 'student':
        student = current_user
    else:
        # For other roles, this shouldn't happen, but handle gracefully
        flash('Access denied', 'error')
        return redirect(url_for('dashboard'))
    
    # Calculate previous semester (last completed semester)
    current_year = student.academic_year or '2025-2026'
    current_semester = student.semester or 1
    
    # Determine previous semester
    if current_semester == 1:
        # If current is 1st semester, previous is 2nd semester of previous year
        prev_year_parts = current_year.split('-')
        prev_year_start = int(prev_year_parts[0]) - 1
        prev_year_end = int(prev_year_parts[1]) - 1
        previous_academic_year = f"{prev_year_start}-{prev_year_end}"
        previous_semester = 2
    else:
        # If current is 2nd semester, previous is 1st semester of same year
        previous_academic_year = current_year
        previous_semester = 1
    
    # Check if current semester grades are complete
    current_semester_grades = Grade.query.join(Subject).filter(
        Grade.student_id == current_user.id,
        Grade.semester == current_semester,
        Grade.academic_year == current_year,
        Grade.is_complete == True
    ).all()
    
    # Check if current semester has any enrolled subjects
    current_semester_enrolled = Grade.query.join(Subject).filter(
        Grade.student_id == current_user.id,
        Grade.semester == current_semester,
        Grade.academic_year == current_year
    ).count()
    
    # Determine which semester to show based on completion
    if current_semester_enrolled > 0 and len(current_semester_grades) == current_semester_enrolled:
        # Current semester is complete, show current semester
        display_academic_year = current_year
        display_semester = current_semester
        is_current_semester = True
    else:
        # Current semester not complete, show previous semester
        display_academic_year = previous_academic_year
        display_semester = previous_semester
        is_current_semester = False
    
    # Get grades for display (either current or previous semester)
    current_grades = Grade.query.join(Subject).filter(
        Grade.student_id == current_user.id,
        Grade.semester == display_semester,
        Grade.academic_year == display_academic_year,
        Grade.is_complete == True  # Only show complete grades to students
    ).order_by(Subject.subject_code).all()
    
    # Calculate GPA and units only for complete grades
    complete_grades = [g for g in current_grades if g.prelim_grade is not None and g.midterm_grade is not None and g.final_grade is not None]
    
    # Total units (ALL subjects - Academic + Non-Academic)
    total_units = sum(g.subject.units for g in complete_grades if g.equivalent_grade)
    
    # GWA calculation (Academic subjects only)
    academic_grades = [g for g in complete_grades if g.subject.subject_type == 'Academic' and g.equivalent_grade]
    academic_units = sum(g.subject.units for g in academic_grades)
    academic_points = sum(g.subject.units * g.equivalent_grade for g in academic_grades)
    
    gwa = round(academic_points / academic_units, 2) if academic_units > 0 else None
    
    # Get all academic years for this student to find the latest one
    all_student_grades = Grade.query.join(Subject).filter(
        Grade.student_id == current_user.id,
        Grade.is_complete == True
    ).all()
    
    # Get unique academic years and sort them to get the latest
    academic_years = sorted(set(g.academic_year for g in all_student_grades), reverse=True)
    latest_academic_year = academic_years[0] if academic_years else (student.academic_year or '2025-2026')
    
    # Calculate total units for the latest academic year (all semesters)
    current_academic_year_grades = [g for g in all_student_grades if g.academic_year == latest_academic_year]
    current_year_total_units = sum(g.subject.units for g in current_academic_year_grades if g.equivalent_grade)
    
    # Check Dean's List eligibility - prioritize current semester if complete, otherwise use display semester
    deans_list_eligible = False
    eligibility_gwa = 0
    eligibility_units = 0
    eligibility_reason = "No GWA calculated"
    eligibility_semester = display_semester
    eligibility_year = display_academic_year
    
    # First, check if current semester is complete and has grades
    if is_current_semester and current_semester_grades:
        # Current semester is complete, check eligibility for current semester
        deans_list_eligible, eligibility_gwa, eligibility_units, eligibility_reason = check_deans_list_eligibility(
            current_user.id,
            current_semester,
            current_year
        )
        eligibility_semester = current_semester
        eligibility_year = current_year
    elif gwa:
        # Current semester not complete, check eligibility for display semester (previous semester)
        deans_list_eligible, eligibility_gwa, eligibility_units, eligibility_reason = check_deans_list_eligibility(
            current_user.id,
            display_semester,
            display_academic_year
        )
        eligibility_semester = display_semester
        eligibility_year = display_academic_year
    
    # Get latest Dean's List record if it exists
    deans_record = DeansListRecord.query.filter_by(
        student_id=current_user.id,
        semester=display_semester,
        academic_year=display_academic_year
    ).first()
    
    return render_template('dashboards/student_dashboard.html',
                         user=current_user,
                         student=student,
                         grades=current_grades,
                         gwa=gwa,
                         total_units=total_units,
                         current_year_total_units=current_year_total_units,
                         latest_academic_year=latest_academic_year,
                         display_academic_year=display_academic_year,
                         display_semester=display_semester,
                         is_current_semester=is_current_semester,
                         current_semester=current_semester,
                         current_year=current_year,
                         deans_list_eligible=deans_list_eligible,
                         eligibility_reason=eligibility_reason,
                         eligibility_semester=eligibility_semester,
                         eligibility_year=eligibility_year,
                         deans_record=deans_record)

@app.route('/registrar/view-student/<int:student_id>')
@login_required
def view_student_record(student_id):
    """View student record in read-only mode for registrars"""
    if current_user.role != 'registrar':
        flash('Access denied. Only registrars can view student records.', 'error')
        return redirect(url_for('dashboard'))
    
    # Get student record
    student = Student.query.get_or_404(student_id)
    
    # Calculate previous semester (last completed semester)
    current_year = student.academic_year or '2025-2026'
    current_semester = student.semester or 1
    
    # Determine previous semester
    if current_semester == 1:
        # If current is 1st semester, previous is 2nd semester of previous year
        prev_year_parts = current_year.split('-')
        prev_year_start = int(prev_year_parts[0]) - 1
        prev_year_end = int(prev_year_parts[1]) - 1
        previous_academic_year = f"{prev_year_start}-{prev_year_end}"
        previous_semester = 2
    else:
        # If current is 2nd semester, previous is 1st semester of same year
        previous_academic_year = current_year
        previous_semester = 1
    
    # Check if current semester grades are complete
    current_semester_grades = Grade.query.join(Subject).filter(
        Grade.student_id == student_id,
        Grade.semester == current_semester,
        Grade.academic_year == current_year,
        Grade.is_complete == True
    ).all()
    
    # Check if current semester has any enrolled subjects
    current_semester_enrolled = Grade.query.join(Subject).filter(
        Grade.student_id == student_id,
        Grade.semester == current_semester,
        Grade.academic_year == current_year
    ).count()
    
    # Determine which semester to show based on completion
    if current_semester_enrolled > 0 and len(current_semester_grades) == current_semester_enrolled:
        # Current semester is complete, show current semester
        display_academic_year = current_year
        display_semester = current_semester
        is_current_semester = True
    else:
        # Current semester not complete, show previous semester
        display_academic_year = previous_academic_year
        display_semester = previous_semester
        is_current_semester = False
    
    # Get grades for display (either current or previous semester)
    current_grades = Grade.query.join(Subject).filter(
        Grade.student_id == student_id,
        Grade.semester == display_semester,
        Grade.academic_year == display_academic_year,
        Grade.is_complete == True
    ).order_by(Subject.subject_code).all()
    
    # Calculate GPA and units only for complete grades
    complete_grades = [g for g in current_grades if g.prelim_grade is not None and g.midterm_grade is not None and g.final_grade is not None]
    
    # Total units (ALL subjects - Academic + Non-Academic)
    total_units = sum(g.subject.units for g in complete_grades if g.equivalent_grade)
    
    # GWA calculation (Academic subjects only)
    academic_grades = [g for g in complete_grades if g.subject.subject_type == 'Academic' and g.equivalent_grade]
    academic_units = sum(g.subject.units for g in academic_grades)
    academic_points = sum(g.subject.units * g.equivalent_grade for g in academic_grades)
    
    gwa = round(academic_points / academic_units, 2) if academic_units > 0 else None
    
    # Get all academic years for this student to find the latest one
    all_student_grades = Grade.query.join(Subject).filter(
        Grade.student_id == student_id,
        Grade.is_complete == True
    ).all()
    
    # Get unique academic years and sort them to get the latest
    academic_years = sorted(set(g.academic_year for g in all_student_grades), reverse=True)
    latest_academic_year = academic_years[0] if academic_years else current_year
    
    # Calculate total units for current academic year
    current_year_grades = Grade.query.join(Subject).filter(
        Grade.student_id == student_id,
        Grade.academic_year == current_year,
        Grade.is_complete == True
    ).all()
    current_year_total_units = sum(g.subject.units for g in current_year_grades if g.equivalent_grade)
    
    # Check Dean's List eligibility
    is_eligible, eligibility_gwa, eligibility_units, eligibility_reason = check_deans_list_eligibility(
        student_id, display_semester, display_academic_year
    )
    
    deans_list_eligible = is_eligible
    eligibility_semester = display_semester
    eligibility_year = display_academic_year
    
    # Get latest Dean's List record if it exists
    deans_record = DeansListRecord.query.filter_by(
        student_id=student_id,
        semester=display_semester,
        academic_year=display_academic_year
    ).first()
    
    return render_template('dashboards/student_dashboard.html',
                         user=current_user,
                         student=student,
                         grades=current_grades,
                         gwa=gwa,
                         total_units=total_units,
                         current_year_total_units=current_year_total_units,
                         latest_academic_year=latest_academic_year,
                         display_academic_year=display_academic_year,
                         display_semester=display_semester,
                         is_current_semester=is_current_semester,
                         current_semester=current_semester,
                         current_year=current_year,
                         deans_list_eligible=deans_list_eligible,
                         eligibility_reason=eligibility_reason,
                         eligibility_semester=eligibility_semester,
                         eligibility_year=eligibility_year,
                         deans_record=deans_record,
                         view_only=True)  # Add view_only flag

@app.route('/api/student/progress-data')
@login_required
def student_progress_data():
    """API endpoint to get student's academic progress data for charts"""
    if current_user.role != 'student':
        return jsonify({'status': 'error', 'message': 'Access denied'}), 403
    
    try:
        # Get all complete grades for the student across all academic years
        all_complete_grades = Grade.query.join(Subject).filter(
            Grade.student_id == current_user.id,
            Grade.is_complete == True
        ).order_by(Grade.academic_year, Grade.semester).all()
        
        # Group grades by academic year and semester
        progress_data = {}
        academic_years = []
        
        for grade in all_complete_grades:
            year_sem_key = f"{grade.academic_year}-S{grade.semester}"
            academic_year = grade.academic_year
            
            if academic_year not in academic_years:
                academic_years.append(academic_year)
            
            if year_sem_key not in progress_data:
                progress_data[year_sem_key] = {
                    'academic_year': academic_year,
                    'semester': grade.semester,
                    'semester_name': f"{grade.semester}{'st' if grade.semester == 1 else 'nd'} Semester",
                    'grades': [],
                    'total_units': 0,
                    'weighted_sum': 0,
                    'gwa': 0
                }
            
            progress_data[year_sem_key]['grades'].append(grade)
        
        # Calculate GWA for each semester
        semester_data = []
        for year_sem_key, data in progress_data.items():
            if data['grades']:
                # Calculate total units and weighted sum
                total_units = sum(g.subject.units for g in data['grades'] if g.equivalent_grade)
                weighted_sum = sum(g.subject.units * g.equivalent_grade 
                                 for g in data['grades'] if g.equivalent_grade)
                
                gwa = round(weighted_sum / total_units, 2) if total_units > 0 else 0
                
                semester_data.append({
                    'label': f"{data['academic_year']} - {data['semester_name']}",
                    'academic_year': data['academic_year'],
                    'semester': data['semester'],
                    'gwa': gwa,
                    'total_units': total_units,
                    'subject_count': len(data['grades'])
                })
        
        # Sort by academic year and semester
        semester_data.sort(key=lambda x: (x['academic_year'], x['semester']))
        
        # Calculate yearly averages
        yearly_data = {}
        for semester in semester_data:
            year = semester['academic_year']
            if year not in yearly_data:
                yearly_data[year] = {
                    'academic_year': year,
                    'semesters': [],
                    'yearly_gwa': 0,
                    'total_units': 0,
                    'total_subjects': 0
                }
            
            yearly_data[year]['semesters'].append(semester)
            yearly_data[year]['total_units'] += semester['total_units']
            yearly_data[year]['total_subjects'] += semester['subject_count']
        
        # Calculate yearly GWA
        yearly_averages = []
        for year, data in yearly_data.items():
            if data['semesters']:
                total_weighted_sum = sum(s['gwa'] * s['total_units'] for s in data['semesters'])
                total_units = data['total_units']
                yearly_gwa = round(total_weighted_sum / total_units, 2) if total_units > 0 else 0
                
                yearly_averages.append({
                    'label': year,
                    'academic_year': year,
                    'gwa': yearly_gwa,
                    'total_units': total_units,
                    'semester_count': len(data['semesters'])
                })
        
        yearly_averages.sort(key=lambda x: x['academic_year'])
        
        return jsonify({
            'status': 'success',
            'data': {
                'semester_data': semester_data,
                'yearly_data': yearly_averages,
                'academic_years': sorted(academic_years)
            }
        })
        
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/instructor/dashboard')
@login_required
def instructor_dashboard():
    """Instructor dashboard for grade management (also accessible by registrars)"""
    if current_user.role not in ['instructor', 'registrar']:
        flash('Access denied', 'error')
        return redirect(url_for('dashboard'))
    
    # Get subjects with enrolled students and grade status information
    if current_user.role == 'instructor':
        # Get subjects assigned to this instructor via ClassAssignment with section info
        instructor_subjects_query = db.session.query(Subject, ClassAssignment).join(
            ClassAssignment, Subject.id == ClassAssignment.subject_id
        ).filter(
            ClassAssignment.instructor_id == current_user.id
        ).all()
        
        # Create separate entries for each subject-section combination
        subjects_with_assignments = []
        for subject, assignment in instructor_subjects_query:
            # Create a unique identifier for subject-section combination
            unique_key = f"{subject.id}_{assignment.section or 'General'}_{assignment.school_year}"
            
            subjects_with_assignments.append({
                'subject': subject,
                'assignment': assignment,
                'unique_key': unique_key,
                'display_name': f"{subject.subject_code} - {subject.subject_name} ({assignment.section or 'General'})"
            })
        
        subjects = [item['subject'] for item in subjects_with_assignments]
        subject_assignments = {item['unique_key']: item['assignment'] for item in subjects_with_assignments}
    else:  # registrar
        # Registrars can see all subjects that have enrolled students
        subjects_with_enrollments = db.session.query(Subject).join(
            StudentSubject, Subject.id == StudentSubject.subject_id
        ).filter(
            StudentSubject.status == 'ENROLLED'
        ).distinct().all()
        subjects = subjects_with_enrollments
        subject_assignments = {}
    
    # Get grade status for each subject
    subject_status = {}
    
    if current_user.role == 'instructor':
        # For instructors, loop through assignments to get status for each subject-section combination
        for unique_key, assignment in subject_assignments.items():
            subject = Subject.query.get(assignment.subject_id)
            if not subject:
                continue
            
            academic_year_for_grades = assignment.school_year if assignment.school_year else subject.academic_year
            
            # Try multiple academic year combinations to find enrolled students (same approach as encode_grades)
            academic_years_to_try = []
            if assignment.school_year:
                academic_years_to_try.append(assignment.school_year)
            if subject.academic_year:
                academic_years_to_try.append(subject.academic_year)
            # Remove duplicates while preserving order
            academic_years_to_try = list(dict.fromkeys(academic_years_to_try))
            
            # Count enrolled students for this subject with multiple academic year options
            enrolled_students = StudentSubject.query.filter(
                StudentSubject.subject_id == subject.id,
                StudentSubject.semester == subject.semester,
                StudentSubject.status == 'ENROLLED'
            )
            
            if academic_years_to_try:
                enrolled_students = enrolled_students.filter(StudentSubject.academic_year.in_(academic_years_to_try))
            
            enrolled_students = enrolled_students.all()
            
            # Filter by section if this is for a specific assignment with a section
            # Count all enrolled students first (including active check)
            student_ids = [es.student_id for es in enrolled_students]
            if student_ids:
                active_students = Student.query.filter(
                    Student.id.in_(student_ids),
                    Student.active == True
                ).all()
                enrolled_count = len(active_students)
            else:
                enrolled_count = 0
            
            # Further filter by section if assignment has a section
            if assignment.section and student_ids:
                # Filter by section
                section_students = Student.query.filter(
                    Student.id.in_(student_ids),
                    Student.section == assignment.section
                ).all()
                # If students found in the specific section, use that count
                if section_students:
                    enrolled_count = len(section_students)
                # Otherwise, keep the count from all active students (already set above)
            
            # Count grades by status using the correct academic year
            total_grades = Grade.query.filter_by(
                subject_id=subject.id,
                semester=subject.semester,
                academic_year=academic_year_for_grades
            ).count()
            
            # Count complete grades (using is_complete field)
            complete_grades = Grade.query.filter_by(
                subject_id=subject.id,
                semester=subject.semester,
                academic_year=academic_year_for_grades,
                is_complete=True
            ).count()
            
            # Count draft grades (incomplete grades)
            draft_grades = Grade.query.filter_by(
                subject_id=subject.id,
                semester=subject.semester,
                academic_year=academic_year_for_grades,
                is_complete=False
            ).count()
            
            # Store status using unique_key for each subject-section combination
            subject_status[unique_key] = {
                'enrolled_count': enrolled_count,
                'total': total_grades,
                'complete': complete_grades,
                'draft': draft_grades
            }
    else:
        # For registrars, loop through subjects
        for subject in subjects:
            academic_year_for_grades = subject.academic_year
            
            # Try multiple academic year combinations to find enrolled students
            academic_years_to_try = []
            if subject.academic_year:
                academic_years_to_try.append(subject.academic_year)
            
            # Count enrolled students for this subject with multiple academic year options
            enrolled_students = StudentSubject.query.filter(
                StudentSubject.subject_id == subject.id,
                StudentSubject.semester == subject.semester,
                StudentSubject.status == 'ENROLLED'
            )
            
            if academic_years_to_try:
                enrolled_students = enrolled_students.filter(StudentSubject.academic_year.in_(academic_years_to_try))
            
            enrolled_students = enrolled_students.all()
            enrolled_count = len(enrolled_students)
            
            # Count grades by status using the correct academic year
            total_grades = Grade.query.filter_by(
                subject_id=subject.id,
                semester=subject.semester,
                academic_year=academic_year_for_grades
            ).count()
            
            # Count complete grades (using is_complete field)
            complete_grades = Grade.query.filter_by(
                subject_id=subject.id,
                semester=subject.semester,
                academic_year=academic_year_for_grades,
                is_complete=True
            ).count()
            
            # Count draft grades (incomplete grades)
            draft_grades = Grade.query.filter_by(
                subject_id=subject.id,
                semester=subject.semester,
                academic_year=academic_year_for_grades,
                is_complete=False
            ).count()
            
            subject_status[subject.id] = {
                'enrolled_count': enrolled_count,
                'total': total_grades,
                'complete': complete_grades,
                'draft': draft_grades
            }
    
    return render_template('dashboards/instructor_dashboard.html', 
                         user=current_user, 
                         subjects=subjects,
                         subject_status=subject_status,
                         subject_assignments=subject_assignments)

@app.route('/registrar/instructor-selection')
@login_required
def instructor_selection():
    """Registrar instructor selection page"""
    if current_user.role != 'registrar':
        flash('Access denied. Only registrars can access this page.', 'error')
        return redirect(url_for('dashboard'))
    
    # Get all instructors
    instructors = User.query.filter_by(role='instructor', active=True).order_by(User.last_name, User.first_name).all()
    
    # Get instructor statistics
    instructor_stats = []
    instructor_stats_json = []  # JSON-serializable version
    
    for instructor in instructors:
        # Count ClassAssignment records (assignments with sections) for this instructor
        subject_count = ClassAssignment.query.filter_by(
            instructor_id=instructor.id
        ).count()
        
        # Original format for template rendering
        instructor_stats.append({
            'instructor': instructor,
            'subject_count': subject_count
        })
        
        # JSON-serializable format for JavaScript
        instructor_stats_json.append({
            'instructor': {
                'id': instructor.id,
                'first_name': instructor.first_name,
                'last_name': instructor.last_name,
                'email': instructor.email,
                'department': instructor.department,
                'active': instructor.active
            },
            'subject_count': subject_count
        })
    
    return render_template('registrar/registrar_instructor_selection.html', 
                         user=current_user, 
                         instructor_stats=instructor_stats,
                         instructor_stats_json=instructor_stats_json)

@app.route('/registrar/instructor-dashboard/<int:instructor_id>')
@login_required
def registrar_instructor_dashboard(instructor_id):
    """Registrar view of specific instructor's dashboard"""
    if current_user.role != 'registrar':
        flash('Access denied. Only registrars can access this page.', 'error')
        return redirect(url_for('dashboard'))
    
    # Get the instructor
    instructor = User.query.get_or_404(instructor_id)
    if instructor.role != 'instructor':
        flash('Invalid instructor selected.', 'error')
        return redirect(url_for('instructor_selection'))
    
    # Get subjects assigned to the instructor that have enrolled students
    subjects_query = db.session.query(Subject, ClassAssignment).join(
        StudentSubject, Subject.id == StudentSubject.subject_id
    ).join(
        ClassAssignment, Subject.id == ClassAssignment.subject_id
    ).filter(
        ClassAssignment.instructor_id == instructor.id,
        StudentSubject.status == 'ENROLLED'
    ).all()
    
    # Group subjects by subject_id and collect their assignments
    subjects_dict = {}
    for subject, assignment in subjects_query:
        if subject.id not in subjects_dict:
            subjects_dict[subject.id] = {
                'subject': subject,
                'assignments': []
            }
        subjects_dict[subject.id]['assignments'].append(assignment)
    
    subjects = [data['subject'] for data in subjects_dict.values()]
    subject_assignments = subjects_dict
    
    # Get grade status for each subject
    subject_status = {}
    for subject in subjects:
        # Get the correct academic year from the assignment
        assignment = ClassAssignment.query.filter_by(
            subject_id=subject.id,
            instructor_id=instructor.id
        ).first()
        academic_year_for_grades = assignment.school_year if assignment and assignment.school_year else subject.academic_year
        
        # Try multiple academic year combinations to find enrolled students
        academic_years_to_try = []
        if assignment and assignment.school_year:
            academic_years_to_try.append(assignment.school_year)
        if subject.academic_year:
            academic_years_to_try.append(subject.academic_year)
        # Remove duplicates while preserving order
        academic_years_to_try = list(dict.fromkeys(academic_years_to_try))
        
        # Count enrolled students for this subject with multiple academic year options
        enrolled_students = StudentSubject.query.filter(
            StudentSubject.subject_id == subject.id,
            StudentSubject.semester == subject.semester,
            StudentSubject.status == 'ENROLLED'
        )
        
        if academic_years_to_try:
            enrolled_students = enrolled_students.filter(StudentSubject.academic_year.in_(academic_years_to_try))
        
        enrolled_students = enrolled_students.all()
        enrolled_count = len(enrolled_students)
        
        # Count grades by status using the correct academic year
        total_grades = Grade.query.filter_by(
            subject_id=subject.id,
            semester=subject.semester,
            academic_year=academic_year_for_grades
        ).count()
        
        # Count complete grades (using is_complete field)
        complete_grades = Grade.query.filter_by(
            subject_id=subject.id,
            semester=subject.semester,
            academic_year=academic_year_for_grades,
            is_complete=True
        ).count()
        
        # Count draft grades (incomplete grades)
        draft_grades = Grade.query.filter_by(
            subject_id=subject.id,
            semester=subject.semester,
            academic_year=academic_year_for_grades,
            is_complete=False
        ).count()
        
        subject_status[subject.id] = {
            'enrolled_count': enrolled_count,
            'total': total_grades,
            'complete': complete_grades,
            'draft': draft_grades
        }
    
    return render_template('dashboards/instructor_dashboard.html', 
                         user=current_user, 
                         subjects=subjects,
                         subject_status=subject_status,
                         subject_assignments=subject_assignments,
                         selected_instructor=instructor)

@app.route('/instructor/my-classes')
@login_required
def instructor_my_classes():
    """Deprecated route: removed as requested."""
    flash('This page has been removed.', 'info')
    return redirect(url_for('instructor_dashboard'))

@app.route('/registrar/grade-approval')
@login_required
def grade_approval():
    """Grade approval page for registrar"""
    if current_user.role != 'registrar':
        flash('Access denied. Only registrars can access this page.', 'error')
        return redirect(url_for('dashboard'))
    
    # Get all submitted grades that need approval
    grades = db.session.query(
        Grade, Student, Subject
    ).join(
        Student, Grade.student_id == Student.id
    ).join(
        Subject, Grade.subject_id == Subject.id
    ).filter(
        Grade.is_locked == True,
        Grade.submitted_at.isnot(None),
        Grade.approved_at.is_(None)
    ).all()
    
    return render_template('registrar_approval.html', user=current_user, grades=grades)


@app.route('/registrar/dashboard')
@login_required
def registrar_dashboard():
    """Registrar dashboard for academic oversight"""
    if current_user.role != 'registrar':
        flash('Access denied. Only registrars can access this page.', 'error')
        return redirect(url_for('dashboard'))

    # Get filter parameters from request
    filter_semester = request.args.get('semester', '')
    
    # Get available academic years from database
    academic_years_db = AcademicYear.query.order_by(AcademicYear.academic_year_name.desc()).all()
    academic_years = [academic_year.academic_year_name for academic_year in academic_years_db]
    
    # Always use the latest academic year
    latest_academic_year = academic_years[0] if academic_years else "2024-2025"
    
    # Set current academic period - always use latest academic year, semester from filter
    current_academic_year = latest_academic_year
    current_semester = int(filter_semester) if filter_semester else 1
    
    # Get all students for current academic period
    students_query = Student.query.filter(
        Student.semester == current_semester,
        Student.academic_year == current_academic_year
    ).all()
    
    # Filter for Dean's List eligible students
    deans_list_students = []
    for student in students_query:
        is_eligible, gwa, total_units, reason = check_deans_list_eligibility(
            student.id,  # Use student.id instead of user.id
            current_semester,
            current_academic_year
        )
        
        if is_eligible:
            deans_list_students.append({
                'student': student,
                'gwa': gwa,
                'total_units': total_units,
                'reason': reason
            })
    
    # Sort by GWA (best first)
    deans_list_students.sort(key=lambda x: x['gwa'])
    
    # Assign ranks with tie handling
    current_rank = 1
    for i, student_data in enumerate(deans_list_students):
        if i == 0:
            # First student gets rank 1
            student_data['rank'] = 1
        else:
            # Check if current student has same GWA as previous
            if student_data['gwa'] == deans_list_students[i-1]['gwa']:
                # Same GWA, same rank as previous student
                student_data['rank'] = deans_list_students[i-1]['rank']
            else:
                # Different GWA, rank is position in list (i+1)
                student_data['rank'] = i + 1
        current_rank = student_data['rank']
    
    # Get all instructors
    instructors = User.query.filter_by(role='instructor').all()
    
    # Get all subjects
    subjects = Subject.query.all()
    
    # Get statistics
    total_students = len(deans_list_students)
    total_instructors = len(instructors)
    total_subjects = len(subjects)
    
    return render_template('dashboards/registrar_dashboard.html', 
                         user=current_user, 
                         students=deans_list_students,
                         instructors=instructors,
                         subjects=subjects,
                         total_students=total_students,
                         total_instructors=total_instructors,
                         total_subjects=total_subjects,
                         current_semester=current_semester,
                         current_academic_year=current_academic_year,
                         academic_years=academic_years)

@app.route('/registrar/promotion-report')
@login_required
def registrar_promotion_report():
    """Promotion report page for registrars and deans"""
    if current_user.role not in ['registrar', 'dean']:
        flash('Access denied. Only registrars and deans can access this page.', 'error')
        return redirect(url_for('dashboard'))

    # Get filter parameters from request
    filter_semester = request.args.get('semester', '')
    filter_academic_year = request.args.get('academic_year', '')
    filter_department = request.args.get('department', '')
    filter_year_level = request.args.get('year_level', '')
    filter_section = request.args.get('section', '')
    applied = request.args.get('applied', '')
    
    # Get available academic years from database
    academic_years_db = AcademicYear.query.order_by(AcademicYear.academic_year_name.desc()).all()
    academic_years = [academic_year.academic_year_name for academic_year in academic_years_db]
    
    # Use filter academic year if provided, otherwise use the latest
    if filter_academic_year and filter_academic_year in academic_years:
        current_academic_year = filter_academic_year
    else:
        current_academic_year = academic_years[0] if academic_years else "2024-2025"
    
    # Set current academic period - use filtered values or defaults
    current_semester = int(filter_semester) if filter_semester else 1
    
    # Check if filters were applied via Apply button
    filters_applied = applied == 'true'
    
    # Get all students for current academic period with additional filters
    students_query = Student.query.filter(
        Student.semester == current_semester,
        Student.academic_year == current_academic_year
    )
    
    # Apply additional filters if provided
    if filter_department:
        students_query = students_query.filter(Student.department == filter_department)
    if filter_year_level:
        students_query = students_query.filter(Student.year_level == filter_year_level)
    if filter_section:
        students_query = students_query.filter(Student.section == filter_section)
    
    students_query = students_query.all()
    
    # Get all unique subjects based on all selected filters
    subjects_query = Subject.query.filter(
        Subject.semester == current_semester,
        Subject.academic_year == current_academic_year
    )
    
    # Apply additional filters to subjects if provided
    if filter_department:
        subjects_query = subjects_query.filter(Subject.department == filter_department)
    if filter_year_level:
        subjects_query = subjects_query.filter(Subject.year_level == filter_year_level)
    
    all_subjects = subjects_query.order_by(Subject.subject_code).all()
    
    # Get all unique sections from students
    unique_sections = db.session.query(Student.section).filter(
        Student.section.isnot(None),
        Student.section != ''
    ).distinct().order_by(Student.section).all()
    available_sections = [section[0] for section in unique_sections if section[0]]
    
    # Get all unique year levels from students
    unique_year_levels = db.session.query(Student.year_level).filter(
        Student.year_level.isnot(None),
        Student.year_level != ''
    ).distinct().order_by(Student.year_level).all()
    available_year_levels = [year_level[0] for year_level in unique_year_levels if year_level[0]]
    
    # Get all unique departments from students and subjects
    unique_departments_students = db.session.query(Student.department).filter(
        Student.department.isnot(None),
        Student.department != ''
    ).distinct().all()
    
    unique_departments_subjects = db.session.query(Subject.department).filter(
        Subject.department.isnot(None),
        Subject.department != ''
    ).distinct().all()
    
    # Combine and deduplicate departments
    all_departments = set()
    for dept in unique_departments_students:
        if dept[0]:
            all_departments.add(dept[0])
    for dept in unique_departments_subjects:
        if dept[0]:
            all_departments.add(dept[0])
    
    available_departments = sorted(list(all_departments))
    
    # Prepare promotion data with actual calculations
    promotion_data = []
    promoted_count = 0
    conditional_count = 0
    retained_count = 0
    
    for student in students_query:
        # Get student's grades for the selected academic year and semester
        student_grades = Grade.query.join(Subject).filter(
            Grade.student_id == student.id,
            Grade.semester == current_semester,
            Grade.academic_year == current_academic_year,
            Grade.is_complete == True
        ).all()
        
        # Calculate GWA based on academic subjects only
        complete_grades = [g for g in student_grades if g.prelim_grade is not None and g.midterm_grade is not None and g.final_grade is not None]
        academic_grades = [g for g in complete_grades if g.subject.subject_type == 'Academic' and g.equivalent_grade]
        
        # Calculate GWA
        academic_units = sum(g.subject.units for g in academic_grades)
        academic_points = sum(g.subject.units * g.equivalent_grade for g in academic_grades)
        gwa = round(academic_points / academic_units, 2) if academic_units > 0 else None
        
        # Get subjects enrolled for the selected academic year and semester
        enrolled_subjects = StudentSubject.query.join(Subject).filter(
            StudentSubject.student_id == student.id,
            StudentSubject.semester == current_semester,
            StudentSubject.academic_year == current_academic_year,
            StudentSubject.status == 'ENROLLED'
        ).all()
        
        # Get subject objects from enrolled subjects
        subjects_enrolled = [enrolled.subject for enrolled in enrolled_subjects]
        
        # Calculate total units (all enrolled subjects)
        total_units = sum(subject.units for subject in subjects_enrolled)
        
        # Get individual subject grades for each subject
        subject_grades = {}
        subject_instructors = {}
        for subject in all_subjects:
            grade = Grade.query.filter(
                Grade.student_id == student.id,
                Grade.subject_id == subject.id,
                Grade.semester == current_semester,
                Grade.academic_year == current_academic_year
            ).first()
            
            if grade and grade.equivalent_grade is not None:
                subject_grades[subject.subject_code] = round(grade.equivalent_grade, 2)
            else:
                subject_grades[subject.subject_code] = None
            
            # Get instructor for this subject
            instructor = None
            if subject.instructor_id:
                instructor = User.query.get(subject.instructor_id)
            
            # Format instructor name as "First letter. Last name"
            if instructor:
                first_initial = instructor.first_name[0].upper() if instructor.first_name else ''
                last_name = instructor.last_name if instructor.last_name else ''
                subject_instructors[subject.subject_code] = f"{first_initial}. {last_name}"
            else:
                subject_instructors[subject.subject_code] = "TBA"
        
        # Determine promotion status based on GWA
        if gwa is None:
            remarks = 'No Grades'
        elif gwa <= 2.5:
            remarks = 'Passed'
            promoted_count += 1
        else:
            remarks = 'Failed'
            retained_count += 1
        
        # Get faculty information (get the most common instructor for this student)
        faculty_grades = Grade.query.filter(
            Grade.student_id == student.id,
            Grade.semester == current_semester,
            Grade.academic_year == current_academic_year
        ).join(Subject).all()
        
        faculty_names = []
        for grade in faculty_grades:
            if grade.subject.instructor:
                faculty_names.append(grade.subject.instructor)
        
        # Get the most common faculty member
        faculty = max(set(faculty_names), key=faculty_names.count) if faculty_names else 'N/A'
        
        promotion_data.append({
            'student': student,
            'subjects_enrolled': subjects_enrolled,
            'subject_grades': subject_grades,
            'subject_instructors': subject_instructors,
            'gwa': gwa,
            'total_units': total_units,
            'remarks': remarks,
            'faculty': faculty
        })
    
    # Get instructor information for all subjects (for header display)
    subject_instructors_header = {}
    for subject in all_subjects:
        instructor = None
        if subject.instructor_id:
            instructor = User.query.get(subject.instructor_id)
        
        # Format instructor name as "First letter. Last name"
        if instructor:
            first_initial = instructor.first_name[0].upper() if instructor.first_name else ''
            last_name = instructor.last_name if instructor.last_name else ''
            subject_instructors_header[subject.subject_code] = f"{first_initial}. {last_name}"
        else:
            subject_instructors_header[subject.subject_code] = "TBA"
    
    return render_template('registrar/registrar_promotion_report.html', 
                         user=current_user, 
                         promotion_data=promotion_data,
                         promoted_count=promoted_count,
                         conditional_count=conditional_count,
                         retained_count=retained_count,
                         current_semester=current_semester,
                         current_academic_year=current_academic_year,
                         academic_years=academic_years,
                         all_subjects=all_subjects,
                         subject_instructors_header=subject_instructors_header,
                         available_sections=available_sections,
                         available_year_levels=available_year_levels,
                         available_departments=available_departments,
                         filters_applied=filters_applied,
                         current_filter_semester=filter_semester,
                         current_filter_academic_year=filter_academic_year,
                         current_filter_department=filter_department,
                         current_filter_year_level=filter_year_level,
                         current_filter_section=filter_section)

@app.route('/registrar/grade-sheet')
@login_required
def registrar_grade_sheet():
    """Comprehensive grade sheet management for registrars"""
    if current_user.role != 'registrar':
        flash('Access denied. Only registrars can access this page.', 'error')
        return redirect(url_for('dashboard'))
    
    try:
        # Get current academic period
        current_semester = 1  # Default to first semester
        current_academic_year = "2024-2025"  # Default academic year
        
        # Get all grades with student and subject information
        all_grades = db.session.query(
            Grade, Student, Subject, User
        ).join(
            Student, Grade.student_id == Student.id
        ).join(
            Subject, Grade.subject_id == Subject.id
        ).outerjoin(
            User, Grade.approved_by == User.id
        ).filter(
            Grade.semester == current_semester,
            Grade.academic_year == current_academic_year
        ).all()
        
        # Calculate statistics
        total_students = len(set(grade.student_id for grade, _, _, _ in all_grades))
        approved_grades = len([g for g, _, _, _ in all_grades if g.approved_at is not None])
        pending_grades = len([g for g, _, _, _ in all_grades if g.approved_at is None and g.is_complete])
        
        # Calculate class average
        completed_grades = [g for g, _, _, _ in all_grades if g.final_average is not None]
        class_average = sum(g.final_average for g in completed_grades) / len(completed_grades) if completed_grades else 0
        
        # Get unique values for filters
        departments = list(set(subject.department for _, _, subject, _ in all_grades if subject.department))
        year_levels = list(set(subject.year_level for _, _, subject, _ in all_grades if subject.year_level))
        sections = list(set(subject.section for _, _, subject, _ in all_grades if subject.section))
        subjects = list(set((subject.subject_code, subject.subject_name) for _, _, subject, _ in all_grades))
        
        return render_template('registrar/registrar_grade_sheet.html',
                             user=current_user,
                             grades=all_grades,
                             total_students=total_students,
                             approved_grades=approved_grades,
                             pending_grades=pending_grades,
                             class_average=class_average,
                             departments=departments,
                             year_levels=year_levels,
                             sections=sections,
                             subjects=subjects,
                             current_semester=current_semester,
                             current_academic_year=current_academic_year)
    
    except Exception as e:
        print(f"Error in registrar_grade_sheet: {e}")
        flash('Error loading grade sheet data', 'error')
        return redirect(url_for('registrar_dashboard'))

@app.route('/registrar/report-grade')
@login_required
def registrar_report_grade():
    """Comprehensive grade reporting and analytics for registrars"""
    if current_user.role != 'registrar':
        flash('Access denied. Only registrars can access this page.', 'error')
        return redirect(url_for('dashboard'))
    
    try:
        # Get current academic period
        current_semester = 1  # Default to first semester
        current_academic_year = "2024-2025"  # Default academic year
        
        # Get all grades with student and subject information
        all_grades = db.session.query(
            Grade, Student, Subject, User
        ).join(
            Student, Grade.student_id == Student.id
        ).join(
            Subject, Grade.subject_id == Subject.id
        ).outerjoin(
            User, Grade.approved_by == User.id
        ).filter(
            Grade.semester == current_semester,
            Grade.academic_year == current_academic_year
        ).all()
        
        # Calculate statistics
        total_students = len(set(grade.student_id for grade, _, _, _ in all_grades))
        
        # Calculate overall average
        completed_grades = [g for g, _, _, _ in all_grades if g.final_average is not None]
        overall_average = sum(g.final_average for g in completed_grades) / len(completed_grades) if completed_grades else 0
        
        # Calculate passing rate (grades >= 75)
        passing_grades = [g for g in completed_grades if g.final_average >= 75]
        passing_rate = (len(passing_grades) / len(completed_grades) * 100) if completed_grades else 0
        
        # Calculate Dean's List count (GWA <= 1.75)
        deans_list_count = len([g for g in completed_grades if g.equivalent_grade and g.equivalent_grade <= 1.75])
        
        # Calculate grade distribution
        grade_distribution = {
            'excellent': 0,
            'good': 0,
            'satisfactory': 0,
            'passing': 0
        }
        
        for grade in completed_grades:
            if grade.equivalent_grade:
                if grade.equivalent_grade <= 1.75:
                    grade_distribution['excellent'] += 1
                elif grade.equivalent_grade <= 2.75:
                    grade_distribution['good'] += 1
                elif grade.equivalent_grade <= 3.25:
                    grade_distribution['satisfactory'] += 1
                else:
                    grade_distribution['passing'] += 1
        
        # Convert to percentages
        total_grades = len(completed_grades)
        if total_grades > 0:
            for key in grade_distribution:
                grade_distribution[key] = (grade_distribution[key] / total_grades) * 100
        
        # Calculate department averages
        department_averages = {}
        department_grades = {}
        
        for grade, _, subject, _ in all_grades:
            if grade.final_average is not None and subject.department:
                dept = subject.department
                if dept not in department_grades:
                    department_grades[dept] = []
                department_grades[dept].append(grade.final_average)
        
        for dept, grades in department_grades.items():
            department_averages[dept] = sum(grades) / len(grades)
        
        # Get unique values for filters
        departments = list(set(subject.department for _, _, subject, _ in all_grades if subject.department))
        year_levels = list(set(subject.year_level for _, _, subject, _ in all_grades if subject.year_level))
        
        return render_template('registrar/registrar_report_grade.html',
                             user=current_user,
                             grades=all_grades,
                             total_students=total_students,
                             overall_average=overall_average,
                             passing_rate=passing_rate,
                             deans_list_count=deans_list_count,
                             grade_distribution=grade_distribution,
                             department_averages=department_averages,
                             departments=departments,
                             year_levels=year_levels,
                             current_semester=current_semester,
                             current_academic_year=current_academic_year)
    
    except Exception as e:
        print(f"Error in registrar_report_grade: {e}")
        flash('Error loading grade report data', 'error')
        return redirect(url_for('registrar_dashboard'))

@app.route('/registrar/subject-offering')
@login_required
def registrar_subject_offering():
    """Registrar subject offering page"""
    if current_user.role != 'registrar':
        flash('Access denied. Only registrars can access this page.', 'error')
        return redirect(url_for('dashboard'))
    
    try:
        # Get all subjects with instructor information and class assignments
        subjects = Subject.query.options(
            db.joinedload(Subject.instructor),
            db.joinedload(Subject.class_assignments).joinedload(ClassAssignment.instructor)
        ).all()
        
        # Calculate enrolled count for each subject from Enrollment table
        for subject in subjects:
            # Count active enrollments - use current academic year
            current_year = "2025-2026"  # Default academic year
            enrolled_count = Enrollment.query.filter_by(
                subject_id=subject.id,
                semester=subject.semester,
                academic_year=current_year,
                status='Active'
            ).count()
            subject.enrolled_count = enrolled_count
        
        return render_template('registrar/registrar_subject_offering.html', subjects=subjects)
        
    except Exception as e:
        print(f"Error in registrar_subject_offering: {str(e)}")
        flash('An error occurred while loading subject offerings.', 'error')
        return redirect(url_for('registrar_dashboard'))

@app.route('/api/subject/<int:subject_id>/enrolled-students', methods=['GET'])
@login_required
def get_enrolled_students(subject_id):
    """Get list of enrolled students for a subject"""
    if current_user.role != 'registrar':
        return jsonify({'status': 'error', 'message': 'Access denied'}), 403
    
    try:
        subject = Subject.query.get_or_404(subject_id)
        
        # Get all students enrolled in this subject via Enrollment table + Student table
        enrolled_students = db.session.query(Student, Enrollment).join(
            Enrollment, Student.id == Enrollment.student_id
        ).filter(
            Enrollment.subject_id == subject_id,
            Enrollment.semester == subject.semester,
            Enrollment.academic_year == subject.academic_year
        ).all()
        
        students_data = []
        for student, enrollment in enrolled_students:
            # Get grade if exists
            grade = Grade.query.filter_by(
                student_id=student.id,
                subject_id=subject_id,
                semester=subject.semester,
                academic_year=subject.academic_year
            ).first()
            
            students_data.append({
                'id': student.id,
                'student_id': student.student_id,
                'name': f"{student.first_name} {student.last_name}",
                'year_level': student.year_level,
                'section': student.section or 'N/A',
                'enrollment_status': enrollment.status,
                'enrollment_date': enrollment.enrollment_date.strftime('%Y-%m-%d') if enrollment.enrollment_date else 'N/A',
                'final_average': grade.final_average if grade and grade.final_average else 'N/A',
                'is_complete': grade.is_complete if grade else False,
                'is_locked': grade.is_locked if grade else False
            })
        
        return jsonify({
            'status': 'success',
            'subject': {
                'code': subject.subject_code,
                'name': subject.subject_name,
                'semester': subject.semester,
                'academic_year': subject.academic_year
            },
            'students': students_data,
            'total_count': len(students_data)
        })
        
    except Exception as e:
        print(f"Error getting enrolled students: {str(e)}")
        return jsonify({'status': 'error', 'message': str(e)}), 500

# =====================================
# ENROLLMENT MANAGEMENT ROUTES
# =====================================

@app.route('/registrar/enrollment-management')
@login_required
def enrollment_management():
    """Enrollment management page for registrars"""
    if current_user.role != 'registrar':
        flash('Access denied. Only registrars can access this page.', 'error')
        return redirect(url_for('dashboard'))
    
    try:
        # Get all subjects with enrollment counts
        subjects = Subject.query.all()
        
        for subject in subjects:
            # Count active enrollments
            subject.enrolled_count = Enrollment.query.filter_by(
                subject_id=subject.id,
                semester=subject.semester,
                academic_year=subject.academic_year,
                status='Active'
            ).count()
            
            # Check capacity - ensure max_capacity is not None
            max_capacity = subject.max_capacity or 50
            subject.has_capacity = subject.enrolled_count < max_capacity
        
        return render_template('registrar_enrollment.html', subjects=subjects)
        
    except Exception as e:
        print(f"Error in enrollment_management: {str(e)}")
        flash('An error occurred while loading enrollment management.', 'error')
        return redirect(url_for('registrar_dashboard'))

@app.route('/registrar/enroll-students/<int:subject_id>', methods=['GET', 'POST'])
@login_required
def enroll_students(subject_id):
    """Enroll students to a subject"""
    if current_user.role != 'registrar':
        flash('Access denied. Only registrars can enroll students.', 'error')
        return redirect(url_for('dashboard'))
    
    subject = Subject.query.get_or_404(subject_id)
    
    if request.method == 'POST':
        try:
            student_ids = request.form.getlist('student_ids[]')
            
            if not student_ids:
                flash('Please select at least one student to enroll.', 'error')
                return redirect(url_for('enroll_students', subject_id=subject_id))
            
            enrolled_count = 0
            already_enrolled = 0
            
            for student_id in student_ids:
                # Check if already enrolled
                existing_enrollment = Enrollment.query.filter_by(
                    student_id=student_id,
                    subject_id=subject_id,
                    academic_year=subject.academic_year,
                    semester=subject.semester
                ).first()
                
                if existing_enrollment:
                    already_enrolled += 1
                    continue
                
                # Create new enrollment
                enrollment = Enrollment(
                    student_id=student_id,
                    subject_id=subject_id,
                    academic_year=subject.academic_year,
                    semester=subject.semester,
                    status='Active',
                    enrolled_by=current_user.id
                )
                db.session.add(enrollment)
                enrolled_count += 1
            
            db.session.commit()
            
            if enrolled_count > 0:
                flash(f'Successfully enrolled {enrolled_count} student(s).', 'success')
            if already_enrolled > 0:
                flash(f'{already_enrolled} student(s) were already enrolled.', 'info')
            
            return redirect(url_for('enrollment_management'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error enrolling students: {str(e)}', 'error')
            return redirect(url_for('enroll_students', subject_id=subject_id))
    
    # GET request - show eligible students from Student table
    # Get students who match the subject criteria and are not yet enrolled
    eligible_students_query = Student.query.filter(
        Student.department == subject.department,
        Student.year_level == subject.year_level,
        Student.semester == subject.semester,
        Student.academic_year == subject.academic_year,
        Student.enrollment_status == 'Active'
    ).filter(
        ~Student.id.in_(
            db.session.query(Enrollment.student_id).filter_by(
                subject_id=subject_id,
                academic_year=subject.academic_year,
                semester=subject.semester
            )
        )
    ).all()
    
    # Prepare data for template - pass student objects directly
    eligible_students_data = eligible_students_query
    
    # Get currently enrolled students
    enrolled_students = db.session.query(Student, Enrollment).join(
        Enrollment, Student.id == Enrollment.student_id
    ).filter(
        Enrollment.subject_id == subject_id,
        Enrollment.academic_year == subject.academic_year,
        Enrollment.semester == subject.semester
    ).all()
    
    return render_template('registrar_enroll_students.html',
                         subject=subject,
                         eligible_students=eligible_students_data,
                         enrolled_students=enrolled_students)

@app.route('/api/enrollment/drop/<int:enrollment_id>', methods=['POST'])
@login_required
def drop_enrollment(enrollment_id):
    """Drop a student from a subject"""
    if current_user.role not in ['registrar', 'student']:
        return jsonify({'status': 'error', 'message': 'Access denied'}), 403
    
    try:
        enrollment = Enrollment.query.get_or_404(enrollment_id)
        
        # Students can only drop their own enrollments
        if current_user.role == 'student' and enrollment.student_id != current_user.id:
            return jsonify({'status': 'error', 'message': 'Access denied'}), 403
        
        enrollment.status = 'Dropped'
        enrollment.dropped_date = datetime.utcnow()
        db.session.commit()
        
        return jsonify({'status': 'success', 'message': 'Enrollment dropped successfully'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/student/available-subjects')
@login_required
def student_available_subjects():
    """Show subjects available for student enrollment"""
    if current_user.role != 'student':
        flash('Access denied.', 'error')
        return redirect(url_for('dashboard'))
    
    try:
        # Get student record - current_user is already a Student object for students
        if current_user.role == 'student':
            student_record = current_user
        else:
            flash('Access denied.', 'error')
            return redirect(url_for('dashboard'))
        
        # Get subjects for student's year level, department, and semester from Student table
        available_subjects = Subject.query.filter_by(
            department=student_record.department,
            year_level=student_record.year_level,
            semester=student_record.semester,
            academic_year=student_record.academic_year
        ).all()
        
        # Add enrollment status for each subject
        for subject in available_subjects:
            # Check if already enrolled
            enrollment = Enrollment.query.filter_by(
                student_id=current_user.id,
                subject_id=subject.id,
                academic_year=subject.academic_year,
                semester=subject.semester
            ).first()
            
            subject.is_enrolled = enrollment is not None
            subject.enrollment_status = enrollment.status if enrollment else None
            
            # Count current enrollments
            subject.enrolled_count = Enrollment.query.filter_by(
                subject_id=subject.id,
                academic_year=subject.academic_year,
                semester=subject.semester,
                status='Active'
            ).count()
            
            subject.has_capacity = subject.enrolled_count < (subject.max_capacity or 50)
        
        # Get student's current enrollments
        my_enrollments = db.session.query(Enrollment, Subject).join(
            Subject, Enrollment.subject_id == Subject.id
        ).filter(
            Enrollment.student_id == current_user.id,
            Enrollment.academic_year == student_record.academic_year,
            Enrollment.semester == student_record.semester
        ).all()
        
        return render_template('student_enrollment.html',
                             available_subjects=available_subjects,
                             my_enrollments=my_enrollments)
        
    except Exception as e:
        print(f"Error in student_available_subjects: {str(e)}")
        flash('An error occurred while loading available subjects.', 'error')
        return redirect(url_for('student_dashboard'))

@app.route('/api/student/enroll/<int:subject_id>', methods=['POST'])
@login_required
def student_self_enroll(subject_id):
    """Student self-enrollment in a subject"""
    if current_user.role != 'student':
        return jsonify({'status': 'error', 'message': 'Access denied'}), 403
    
    try:
        subject = Subject.query.get_or_404(subject_id)
        
        # Check if already enrolled
        existing_enrollment = Enrollment.query.filter_by(
            student_id=current_user.id,
            subject_id=subject_id,
            academic_year=subject.academic_year,
            semester=subject.semester
        ).first()
        
        if existing_enrollment:
            return jsonify({'status': 'error', 'message': 'You are already enrolled in this subject'}), 400
        
        # Check capacity
        enrolled_count = Enrollment.query.filter_by(
            subject_id=subject_id,
            academic_year=subject.academic_year,
            semester=subject.semester,
            status='Active'
        ).count()
        
        if enrolled_count >= (subject.max_capacity or 50):
            return jsonify({'status': 'error', 'message': 'This subject is at full capacity'}), 400
        
        # Create enrollment
        enrollment = Enrollment(
            student_id=current_user.id,
            subject_id=subject_id,
            academic_year=subject.academic_year,
            semester=subject.semester,
            status='Active',
            enrolled_by=current_user.id
        )
        db.session.add(enrollment)
        db.session.commit()
        
        return jsonify({'status': 'success', 'message': 'Successfully enrolled in subject'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500

# =====================================
# BULK IMPORT FUNCTIONALITY
# =====================================

@app.route('/api/import-students', methods=['POST'])
@login_required
def import_students():
    """Import students from CSV data"""
    if current_user.role != 'registrar':
        return jsonify({'status': 'error', 'message': 'Access denied'}), 403
    
    try:
        import csv
        import io
        
        if 'file' not in request.files:
            return jsonify({'status': 'error', 'message': 'No file uploaded'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'status': 'error', 'message': 'No file selected'}), 400
        
        # Read CSV file
        stream = io.StringIO(file.stream.read().decode("UTF8"), newline=None)
        csv_reader = csv.DictReader(stream)
        
        created_count = 0
        error_count = 0
        errors = []
        
        for row in csv_reader:
            try:
                # Check required fields
                required_fields = ['username', 'email', 'password', 'first_name', 'last_name', 'student_id', 'department', 'year_level', 'semester']
                if not all(field in row and row[field] for field in required_fields):
                    errors.append(f"Row {row_num}: Missing required fields")
                    error_count += 1
                    continue
                
                # Check if user already exists
                if User.query.filter_by(username=row['username']).first():
                    errors.append(f"Row {row_num}: Username '{row['username']}' already exists")
                    error_count += 1
                    continue
                
                if User.query.filter_by(email=row['email']).first():
                    errors.append(f"Row {row_num}: Email '{row['email']}' already exists")
                    error_count += 1
                    continue
                
                # Check if student_id already exists in students table
                if Student.query.filter_by(student_id=row['student_id']).first():
                    errors.append(f"Row {row_num}: Student ID '{row['student_id']}' already exists")
                    error_count += 1
                    continue
                
                # Create independent student record (no user_id needed)
                new_student_record = Student(
                    # Authentication (independent from User table)
                    username=row['username'],
                    email=row['email'],
                    password_hash='',  # Will be set below
                    
                    # Basic Information
                    first_name=row['first_name'],
                    last_name=row['last_name'],
                    middle_name=row.get('middle_name'),
                    suffix=row.get('suffix'),
                    student_id=row['student_id'],
                    student_lrn=row.get('student_lrn'),
                    student_status=row.get('student_status', 'Regular'),
                    
                    # Academic Information
                    department=row['department'],
                    course=row.get('course'),
                    year_level=int(row['year_level']),
                    semester=int(row['semester']),
                    section=row.get('section'),
                    section_type=row.get('section_type'),
                    academic_year=row.get('academic_year', '2025-2026'),
                    curriculum=row.get('curriculum'),
                    graduating=row.get('graduating', 'No'),
                    
                    # Personal Information
                    gender=row.get('gender'),
                    age=int(row.get('age')) if row.get('age') else None,
                    date_birth=row.get('date_birth'),
                    place_birth=row.get('place_birth'),
                    nationality=row.get('nationality', 'Filipino'),
                    religion=row.get('religion'),
                    
                    # Address Information
                    province=row.get('province'),
                    city_municipality=row.get('city_municipality'),
                    barangay=row.get('barangay'),
                    house_no=row.get('house_no'),
                    
                    # Contact Information
                    mobile_no=row.get('mobile_no'),
                    
                    # Status
                    enrollment_status='PENDING',
                    active=True
                )
                new_student_record.set_password(row['password'])
                db.session.add(new_student_record)
                created_count += 1
                
            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")
                error_count += 1
        
        db.session.commit()
        
        return jsonify({
            'status': 'success',
            'message': f'Import completed: {created_count} students created, {error_count} errors',
            'created': created_count,
            'errors': error_count,
            'error_details': errors
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/import-historical-grades', methods=['POST'])
@login_required
def import_historical_grades():
    """Enhanced import for historical grades from Excel/CSV"""
    if current_user.role not in ['registrar', 'dean']:
        return jsonify({'status': 'error', 'message': 'Access denied'}), 403
    
    try:
        import csv
        import io
        from datetime import datetime
        from openpyxl import load_workbook
        
        if 'file' not in request.files:
            return jsonify({'status': 'error', 'message': 'No file uploaded'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'status': 'error', 'message': 'No file selected'}), 400
        
        # Validate file type
        file_extension = file.filename.lower().split('.')[-1]
        if file_extension not in ['csv', 'xlsx']:
            return jsonify({'status': 'error', 'message': 'Only CSV and Excel (.xlsx) files are allowed'}), 400
        
        # Read data based on file type
        if file_extension == 'csv':
            stream = io.StringIO(file.stream.read().decode("UTF8"), newline=None)
            csv_reader = csv.DictReader(stream)
            rows = list(csv_reader)
        else:  # Excel file
            # Load Excel file
            workbook = load_workbook(file.stream)
            worksheet = workbook.active
            
            # Get headers from first row
            headers = [cell.value for cell in worksheet[1]]
            
            # Convert Excel data to list of dictionaries
            rows = []
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                if any(cell is not None for cell in row):  # Skip empty rows
                    row_dict = dict(zip(headers, row))
                    rows.append(row_dict)
        
        # Validate required columns
        required_columns = ['student_id', 'subject_code', 'semester', 'academic_year']
        if rows and not all(col in rows[0].keys() for col in required_columns):
            return jsonify({
                'status': 'error', 
                'message': f'Missing required columns. Required: {required_columns}'
            }), 400
        
        created_count = 0
        updated_count = 0
        error_count = 0
        errors = []
        
        for row_num, row in enumerate(rows, start=2):  # Start from 2 since row 1 is headers
            try:
                # Get student by student_id (using new Student table)
                student = Student.query.filter_by(student_id=row['student_id']).first()
                if not student:
                    errors.append(f"Row {row_num}: Student ID '{row['student_id']}' not found")
                    error_count += 1
                    continue
                
                # Get subject by subject_code
                subject = Subject.query.filter_by(subject_code=row['subject_code']).first()
                if not subject:
                    errors.append(f"Row {row_num}: Subject '{row['subject_code']}' not found")
                    error_count += 1
                    continue
                
                # Parse grades with better validation
                prelim = None
                midterm = None
                final = None
                
                if row.get('prelim_grade') and str(row['prelim_grade']).strip():
                    try:
                        prelim = float(row['prelim_grade'])
                        if prelim < 0 or prelim > 100:
                            errors.append(f"Row {row_num}: Invalid prelim grade {prelim}")
                            error_count += 1
                            continue
                    except (ValueError, TypeError):
                        errors.append(f"Row {row_num}: Invalid prelim grade format")
                        error_count += 1
                        continue
                
                if row.get('midterm_grade') and str(row['midterm_grade']).strip():
                    try:
                        midterm = float(row['midterm_grade'])
                        if midterm < 0 or midterm > 100:
                            errors.append(f"Row {row_num}: Invalid midterm grade {midterm}")
                            error_count += 1
                            continue
                    except (ValueError, TypeError):
                        errors.append(f"Row {row_num}: Invalid midterm grade format")
                        error_count += 1
                        continue
                
                if row.get('final_grade') and str(row['final_grade']).strip():
                    try:
                        final = float(row['final_grade'])
                        if final < 0 or final > 100:
                            errors.append(f"Row {row_num}: Invalid final grade {final}")
                            error_count += 1
                            continue
                    except (ValueError, TypeError):
                        errors.append(f"Row {row_num}: Invalid final grade format")
                        error_count += 1
                        continue
                
                # Calculate final average if all grades present
                final_average = None
                equivalent_grade = None
                remarks = None
                
                if prelim is not None and midterm is not None and final is not None:
                    final_average = round((prelim + midterm + final) / 3, 2)
                    equivalent_grade, remarks = calculate_grade_equivalent(final_average)
                elif row.get('final_average') and row['final_average'].strip():
                    try:
                        final_average = float(row['final_average'])
                        equivalent_grade, remarks = calculate_grade_equivalent(final_average)
                    except ValueError:
                        pass
                
                # Check if grade already exists
                existing_grade = Grade.query.filter_by(
                    student_id=student.id,
                    subject_id=subject.id,
                    semester=int(row['semester']),
                    academic_year=row['academic_year']
                ).first()
                
                if existing_grade:
                    # Update existing grade
                    existing_grade.prelim_grade = prelim
                    existing_grade.midterm_grade = midterm
                    existing_grade.final_grade = final
                    existing_grade.final_average = final_average
                    existing_grade.equivalent_grade = equivalent_grade
                    existing_grade.remarks = remarks
                    existing_grade.is_historical = True
                    existing_grade.import_date = datetime.utcnow()
                    existing_grade.import_source = 'Excel Import' if file_extension == 'xlsx' else 'CSV Import'
                    existing_grade.is_complete = True if final_average else False
                    updated_count += 1
                else:
                    # Create new grade
                    new_grade = Grade(
                        student_id=student.id,
                        subject_id=subject.id,
                        semester=int(row['semester']),
                        academic_year=row['academic_year'],
                        prelim_grade=prelim,
                        midterm_grade=midterm,
                        final_grade=final,
                        final_average=final_average,
                        equivalent_grade=equivalent_grade,
                        remarks=remarks,
                        is_historical=True,
                        import_date=datetime.utcnow(),
                        import_source='Excel Import' if file_extension == 'xlsx' else 'CSV Import',
                        is_complete=True if final_average else False
                    )
                    db.session.add(new_grade)
                    created_count += 1
                
            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")
                error_count += 1
        
        db.session.commit()
        
        return jsonify({
            'status': 'success',
            'message': f'[SUCCESS] Import completed: {created_count} created, {updated_count} updated, {error_count} errors',
            'created': created_count,
            'updated': updated_count,
            'errors': error_count,
            'error_details': errors[:10],  # Limit error details to first 10
            'total_rows': len(rows)
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/download-grade-template')
@login_required
def download_grade_template():
    """Download Excel template for historical grades"""
    if current_user.role not in ['registrar', 'dean']:
        return jsonify({'status': 'error', 'message': 'Access denied'}), 403
    
    academic_year = request.args.get('academic_year', '2023-2024')
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from flask import Response
        
        # Create Excel workbook
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Grade Template"
        
        # Define headers
        headers = [
            'student_id', 'subject_code', 'semester', 'academic_year',
            'prelim_grade', 'midterm_grade', 'final_grade'
        ]
        
        # Style for headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Sample data - get some real students and subjects for better template
        sample_students = Student.query.filter_by(academic_year=academic_year).limit(3).all()
        sample_subjects = Subject.query.filter_by(academic_year=academic_year).limit(2).all()
        
        row_num = 2
        if sample_students and sample_subjects:
            for student in sample_students:
                for subject in sample_subjects:
                    data = [
                        student.student_id,
                        subject.subject_code,
                        subject.semester,
                        academic_year,
                        85.5,  # Sample prelim grade
                        88.0,  # Sample midterm grade
                        90.0   # Sample final grade
                    ]
                    for col, value in enumerate(data, 1):
                        worksheet.cell(row=row_num, column=col, value=value)
                    row_num += 1
        else:
            # Fallback sample data
            fallback_data = [
                ['2021-00001', 'SE102', '1', academic_year, 85.5, 88.0, 90.0],
                ['2021-00001', 'OS101', '1', academic_year, 82.0, 85.5, 88.0],
                ['2022-0471', 'SE102', '1', academic_year, 90.0, 92.0, 95.0],
            ]
            for data in fallback_data:
                for col, value in enumerate(data, 1):
                    worksheet.cell(row=row_num, column=col, value=value)
                row_num += 1
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        from io import BytesIO
        excel_buffer = BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)
        
        return Response(
            excel_buffer.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename=grade_template_{academic_year}.xlsx'}
        )
        
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/registrar/manual-grade-entry')
@login_required
def manual_grade_entry():
    """Simple manual grade entry interface for historical grades"""
    if current_user.role not in ['registrar', 'dean']:
        flash('Access denied', 'error')
        return redirect(url_for('dashboard'))
    
    academic_year = request.args.get('academic_year', '')
    semester = request.args.get('semester', '1')
    
    students = []
    subjects = []
    
    if academic_year:
        students = Student.query.filter_by(academic_year=academic_year).all()
        subjects = Subject.query.filter_by(
            academic_year=academic_year,
            semester=int(semester)
        ).all()
    
    return render_template('registrar/registrar_manual_grade_entry.html',
                         students=students,
                         subjects=subjects,
                         academic_year=academic_year,
                         semester=semester)

@app.route('/registrar/grade-history')
@login_required
def registrar_grade_history():
    """Grade History management page for registrars"""
    if current_user.role not in ['registrar', 'dean']:
        flash('Access denied', 'error')
        return redirect(url_for('dashboard'))
    
    # Get statistics for the dashboard
    historical_grades_count = Grade.query.filter_by(is_historical=True).count()
    
    # Count imports by source
    csv_imports_count = Grade.query.filter_by(
        is_historical=True, 
        import_source='CSV Import'
    ).count()
    
    manual_entries_count = Grade.query.filter_by(
        is_historical=True, 
        import_source='Manual Entry'
    ).count()
    
    # Count unique academic years
    academic_years_count = db.session.query(Grade.academic_year).filter_by(
        is_historical=True
    ).distinct().count()
    
    # Get recent imports (last 10)
    recent_imports = db.session.query(
        Grade.academic_year,
        Grade.semester,
        Grade.import_source,
        Grade.import_date,
        db.func.count(Grade.id).label('records_count')
    ).filter_by(is_historical=True).group_by(
        Grade.academic_year,
        Grade.semester,
        Grade.import_source,
        Grade.import_date
    ).order_by(Grade.import_date.desc()).limit(10).all()
    
    # Count recent entries (this week)
    from datetime import datetime, timedelta
    week_ago = datetime.utcnow() - timedelta(days=7)
    recent_entries = Grade.query.filter(
        Grade.is_historical == True,
        Grade.import_date >= week_ago
    ).count()
    
    return render_template('registrar/registrar_grade_history.html',
                         historical_grades_count=historical_grades_count,
                         csv_imports_count=csv_imports_count,
                         manual_entries_count=manual_entries_count,
                         academic_years_count=academic_years_count,
                         recent_imports=recent_imports,
                         recent_entries=recent_entries)

@app.route('/api/get-students', methods=['GET'])
@login_required
def get_students():
    """Get students by academic year for modal grade entry"""
    if current_user.role not in ['registrar', 'dean']:
        return jsonify({'status': 'error', 'message': 'Access denied'}), 403
    
    try:
        academic_year = request.args.get('academic_year')
        if not academic_year:
            return jsonify({'status': 'error', 'message': 'Academic year is required'}), 400
        
        students = Student.query.filter_by(academic_year=academic_year).all()
        
        students_data = []
        for student in students:
            students_data.append({
                'id': student.id,
                'student_id': student.student_id,
                'first_name': student.first_name,
                'last_name': student.last_name,
                'middle_name': student.middle_name,
                'year_level': student.year_level,
                'section': student.section,
                'course': student.course
            })
        
        return jsonify({
            'status': 'success',
            'data': students_data,
            'count': len(students_data)
        })
        
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/get-subjects', methods=['GET'])
@login_required
def get_subjects():
    """Get subjects by academic year and semester for modal grade entry"""
    if current_user.role not in ['registrar', 'dean']:
        return jsonify({'status': 'error', 'message': 'Access denied'}), 403
    
    try:
        academic_year = request.args.get('academic_year')
        semester = request.args.get('semester')
        
        if not academic_year or not semester:
            return jsonify({'status': 'error', 'message': 'Academic year and semester are required'}), 400
        
        subjects = Subject.query.filter_by(
            academic_year=academic_year,
            semester=int(semester)
        ).all()
        
        subjects_data = []
        for subject in subjects:
            subjects_data.append({
                'id': subject.id,
                'subject_code': subject.subject_code,
                'subject_name': subject.subject_name,
                'units': subject.units,
                'department': subject.department,
                'year_level': subject.year_level,
                'section': subject.section,
                'instructor_id': subject.instructor_id
            })
        
        return jsonify({
            'status': 'success',
            'data': subjects_data,
            'count': len(subjects_data)
        })
        
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/get-grade', methods=['GET'])
@login_required
def get_grade():
    """Get existing grade for a student and subject"""
    if current_user.role not in ['registrar', 'dean']:
        return jsonify({'status': 'error', 'message': 'Access denied'}), 403
    
    try:
        student_id = request.args.get('student_id')
        subject_id = request.args.get('subject_id')
        academic_year = request.args.get('academic_year')
        semester = request.args.get('semester')
        
        if not all([student_id, subject_id, academic_year, semester]):
            return jsonify({'status': 'error', 'message': 'All parameters are required'}), 400
        
        grade = Grade.query.filter_by(
            student_id=student_id,
            subject_id=subject_id,
            academic_year=academic_year,
            semester=int(semester)
        ).first()
        
        if grade:
            return jsonify({
                'status': 'success',
                'grade': {
                    'id': grade.id,
                    'prelim_grade': grade.prelim_grade,
                    'midterm_grade': grade.midterm_grade,
                    'final_grade': grade.final_grade,
                    'final_average': grade.final_average,
                    'equivalent_grade': grade.equivalent_grade,
                    'remarks': grade.remarks,
                    'is_complete': grade.is_complete,
                    'is_historical': grade.is_historical
                }
            })
        else:
            return jsonify({
                'status': 'success',
                'grade': None
            })
        
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/save-historical-grade', methods=['POST'])
@login_required
def save_historical_grade():
    """Save a single historical grade entry"""
    if current_user.role not in ['registrar', 'dean']:
        return jsonify({'status': 'error', 'message': 'Access denied'}), 403
    
    try:
        data = request.get_json()
        
        # Validate required fields
        required_fields = ['student_id', 'subject_id', 'semester', 'academic_year']
        for field in required_fields:
            if field not in data:
                return jsonify({'status': 'error', 'message': f'Missing required field: {field}'}), 400
        
        # Get student and subject
        student = Student.query.get(data['student_id'])
        subject = Subject.query.get(data['subject_id'])
        
        if not student:
            return jsonify({'status': 'error', 'message': 'Student not found'}), 404
        
        if not subject:
            return jsonify({'status': 'error', 'message': 'Subject not found'}), 404
        
        # Parse grades
        prelim = data.get('prelim_grade')
        midterm = data.get('midterm_grade')
        final = data.get('final_grade')
        
        # Calculate final average if all grades present
        final_average = None
        equivalent_grade = None
        remarks = None
        
        if prelim is not None and midterm is not None and final is not None:
            final_average = round((prelim + midterm + final) / 3, 2)
            equivalent_grade, remarks = calculate_grade_equivalent(final_average)
        
        # Check if grade already exists
        existing_grade = Grade.query.filter_by(
            student_id=student.id,
            subject_id=subject.id,
            semester=int(data['semester']),
            academic_year=data['academic_year']
        ).first()
        
        if existing_grade:
            # Update existing grade
            existing_grade.prelim_grade = prelim
            existing_grade.midterm_grade = midterm
            existing_grade.final_grade = final
            existing_grade.final_average = final_average
            existing_grade.equivalent_grade = equivalent_grade
            existing_grade.remarks = remarks
            existing_grade.is_historical = True
            existing_grade.import_date = datetime.utcnow()
            existing_grade.import_source = 'Manual Entry'
            existing_grade.is_complete = True if final_average else False
        else:
            # Create new grade
            new_grade = Grade(
                student_id=student.id,
                subject_id=subject.id,
                semester=int(data['semester']),
                academic_year=data['academic_year'],
                prelim_grade=prelim,
                midterm_grade=midterm,
                final_grade=final,
                final_average=final_average,
                equivalent_grade=equivalent_grade,
                remarks=remarks,
                is_historical=True,
                import_date=datetime.utcnow(),
                import_source='Manual Entry',
                is_complete=True if final_average else False
            )
            db.session.add(new_grade)
        
        db.session.commit()
        
        return jsonify({
            'status': 'success',
            'message': 'Grade saved successfully',
            'final_average': final_average,
            'equivalent_grade': equivalent_grade,
            'remarks': remarks
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/dean/dashboard')
@login_required
def dean_dashboard():
    """Dean dashboard for academic excellence tracking"""
    if current_user.role != 'dean':
        flash('Access denied', 'error')
        return redirect(url_for('dashboard'))
    
    # Get filter parameters from request
    filter_semester = request.args.get('semester', '')
    filter_academic_year = request.args.get('academic_year', '')
    
    # Get available academic years from database
    academic_years_db = AcademicYear.query.order_by(AcademicYear.academic_year_name.desc()).all()
    academic_years = [academic_year.academic_year_name for academic_year in academic_years_db]
    
    # Always use the latest academic year if not specified
    latest_academic_year = academic_years[0] if academic_years else "2024-2025"
    
    # Set current academic period - use filter if provided, otherwise use latest
    current_academic_year = filter_academic_year if filter_academic_year else latest_academic_year
    current_semester = int(filter_semester) if filter_semester else 1
    
    # Get Dean's List records with filters
    query = DeansListRecord.query.filter_by(qualified=True)
    
    if filter_academic_year:
        query = query.filter_by(academic_year=filter_academic_year)
    
    if filter_semester:
        query = query.filter_by(semester=current_semester)
    
    deans_list = query.order_by(DeansListRecord.rank).all()
    
    # Calculate highest GWA from current filtered list
    highest_gwa = deans_list[0].gwa if deans_list else 0
    
    return render_template('dashboards/dean_dashboard.html', 
                         user=current_user, 
                         deans_list=deans_list,
                         current_academic_year=current_academic_year,
                         current_semester=current_semester,
                         academic_years=academic_years)

@app.route('/misit/dashboard')
@login_required
def misit_dashboard():
    """MIS/IT dashboard for system administration"""
    if current_user.role != 'mis_it':
        flash('Access denied', 'error')
        return redirect(url_for('dashboard'))
    
    # Get system statistics
    total_users = User.query.count()
    total_students = User.query.filter_by(role='student').count()
    total_instructors = User.query.filter_by(role='instructor').count()
    
    return render_template('dashboards/misit_dashboard.html', 
                         user=current_user,
                         total_users=total_users,
                         total_students=total_students,
                         total_instructors=total_instructors)

@app.route('/api/system-resources', methods=['GET'])
@login_required
def system_resources():
    """Get real-time system resource usage (MIS/IT only)"""
    if current_user.role != 'mis_it':
        return jsonify({'status': 'error', 'message': 'Unauthorized access'}), 403
    
    try:
        import psutil
        import os
        from datetime import datetime
        
        # Get system resource information
        # Memory usage
        memory = psutil.virtual_memory()
        memory_percent = round(memory.percent, 1)
        memory_used_gb = round(memory.used / (1024**3), 2)
        memory_total_gb = round(memory.total / (1024**3), 2)
        
        # CPU usage
        cpu_percent = round(psutil.cpu_percent(interval=1), 1)
        cpu_count = psutil.cpu_count()
        cpu_freq = psutil.cpu_freq()
        cpu_freq_mhz = round(cpu_freq.current, 0) if cpu_freq else 0
        
        # Disk usage
        disk = psutil.disk_usage('/')
        disk_percent = round(disk.percent, 1)
        disk_used_gb = round(disk.used / (1024**3), 2)
        disk_total_gb = round(disk.total / (1024**3), 2)
        disk_free_gb = round(disk.free / (1024**3), 2)
        
        # Network usage (optional)
        try:
            network = psutil.net_io_counters()
            network_sent_mb = round(network.bytes_sent / (1024**2), 2)
            network_recv_mb = round(network.bytes_recv / (1024**2), 2)
        except:
            network_sent_mb = 0
            network_recv_mb = 0
        
        # Process count
        process_count = len(psutil.pids())
        
        # System uptime
        boot_time = psutil.boot_time()
        uptime_seconds = datetime.now().timestamp() - boot_time
        uptime_hours = round(uptime_seconds / 3600, 1)
        
        # Database size estimation (MySQL)
        try:
            db_size_query = text("""
                SELECT 
                    ROUND(SUM(data_length + index_length) / 1024 / 1024, 2) AS 'DB Size in MB'
                FROM information_schema.tables 
                WHERE table_schema = DATABASE()
            """)
            db_size_result = db.session.execute(db_size_query).fetchone()
            db_size_mb = float(db_size_result[0]) if db_size_result and db_size_result[0] else 0
        except Exception as e:
            db_size_mb = 0
        
        # Resource status indicators
        def get_status_color(percent, thresholds=(70, 85)):
            if percent < thresholds[0]:
                return 'success'
            elif percent < thresholds[1]:
                return 'warning'
            else:
                return 'error'
        
        resource_data = {
            'timestamp': datetime.now().isoformat(),
            'memory': {
                'percent': memory_percent,
                'used_gb': memory_used_gb,
                'total_gb': memory_total_gb,
                'status': get_status_color(memory_percent),
                'status_text': 'Normal' if memory_percent < 70 else 'High' if memory_percent < 85 else 'Critical'
            },
            'cpu': {
                'percent': cpu_percent,
                'cores': cpu_count,
                'frequency_mhz': cpu_freq_mhz,
                'status': get_status_color(cpu_percent),
                'status_text': 'Normal' if cpu_percent < 70 else 'High' if cpu_percent < 85 else 'Critical'
            },
            'storage': {
                'percent': disk_percent,
                'used_gb': disk_used_gb,
                'total_gb': disk_total_gb,
                'free_gb': disk_free_gb,
                'status': get_status_color(disk_percent),
                'status_text': 'Normal' if disk_percent < 70 else 'High' if disk_percent < 85 else 'Critical'
            },
            'system': {
                'processes': process_count,
                'uptime_hours': uptime_hours,
                'database_size_mb': db_size_mb
            },
            'network': {
                'sent_mb': network_sent_mb,
                'received_mb': network_recv_mb
            }
        }
        
        return jsonify({
            'status': 'success',
            'data': resource_data
        })
        
    except ImportError:
        # Fallback if psutil is not available
        return jsonify({
            'status': 'success',
            'data': {
                'timestamp': datetime.now().isoformat(),
                'memory': {
                    'percent': 65,
                    'used_gb': 4.2,
                    'total_gb': 8.0,
                    'status': 'warning',
                    'status_text': 'Simulated'
                },
                'cpu': {
                    'percent': 45,
                    'cores': 4,
                    'frequency_mhz': 2400,
                    'status': 'success',
                    'status_text': 'Simulated'
                },
                'storage': {
                    'percent': 82,
                    'used_gb': 120.5,
                    'total_gb': 150.0,
                    'free_gb': 29.5,
                    'status': 'warning',
                    'status_text': 'Simulated'
                },
                'system': {
                    'processes': 156,
                    'uptime_hours': 72.5,
                    'database_size_mb': 25.8
                },
                'network': {
                    'sent_mb': 1024.5,
                    'received_mb': 2048.3
                }
            }
        })
        
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'Error getting system resources: {str(e)}'
        }), 500

@app.route('/api/system-report', methods=['GET'])
@login_required
def system_report():
    """Generate comprehensive system report (MIS/IT only)"""
    if current_user.role != 'mis_it':
        return jsonify({'status': 'error', 'message': 'Unauthorized access'}), 403
    
    try:
        from datetime import datetime, timedelta
        from sqlalchemy import func, desc
        
        # Get current date and calculate date ranges
        now = datetime.now()
        last_30_days = now - timedelta(days=30)
        last_7_days = now - timedelta(days=7)
        
        # User Statistics
        total_users = User.query.count()
        active_users = User.query.filter_by(active=True).count()
        inactive_users = total_users - active_users
        
        # Role Distribution
        role_stats = db.session.query(
            User.role, 
            func.count(User.id).label('count')
        ).group_by(User.role).all()
        
        role_distribution = {role: count for role, count in role_stats}
        
        # Department Distribution (for students and instructors)
        dept_stats = db.session.query(
            User.department,
            func.count(User.id).label('count')
        ).filter(User.department.isnot(None)).group_by(User.department).all()
        
        department_distribution = {dept: count for dept, count in dept_stats}
        
        # Year Level Distribution (students only)
        year_stats = db.session.query(
            Student.year_level,
            func.count(Student.id).label('count')
        ).group_by(Student.year_level).all()
        
        year_distribution = {year: count for year, count in year_stats}
        
        # Recent User Activity (last 30 days)
        recent_logins = db.session.query(
            func.date(User.created_at).label('date'),
            func.count(User.id).label('count')
        ).filter(User.created_at >= last_30_days).group_by(func.date(User.created_at)).all()
        
        # System Activity (from audit logs)
        try:
            # Check if audit_log table exists and has data
            audit_query = text("""
                SELECT 
                    DATE(created_at) as date,
                    COUNT(*) as count,
                    action
                FROM audit_log 
                WHERE created_at >= :start_date 
                GROUP BY DATE(created_at), action
                ORDER BY date DESC
            """)
            
            audit_results = db.session.execute(audit_query, {'start_date': last_30_days}).fetchall()
            
            # Process audit data
            daily_activity = {}
            for row in audit_results:
                date_str = str(row.date)
                if date_str not in daily_activity:
                    daily_activity[date_str] = {}
                daily_activity[date_str][row.action] = row.count
                
        except Exception as e:
            # If audit_log table doesn't exist or has issues, use empty data
            daily_activity = {}
        
        # Grade Statistics (if grades table exists)
        try:
            grade_query = text("""
                SELECT 
                    COUNT(*) as total_grades,
                    AVG(grade) as average_grade,
                    COUNT(CASE WHEN grade >= 90 THEN 1 END) as excellent_count,
                    COUNT(CASE WHEN grade >= 80 AND grade < 90 THEN 1 END) as good_count,
                    COUNT(CASE WHEN grade >= 70 AND grade < 80 THEN 1 END) as satisfactory_count,
                    COUNT(CASE WHEN grade < 70 THEN 1 END) as needs_improvement_count
                FROM grade
            """)
            
            grade_stats = db.session.execute(grade_query).fetchone()
            
            grade_distribution = {
                'total_grades': grade_stats.total_grades or 0,
                'average_grade': round(float(grade_stats.average_grade or 0), 2),
                'excellent': grade_stats.excellent_count or 0,
                'good': grade_stats.good_count or 0,
                'satisfactory': grade_stats.satisfactory_count or 0,
                'needs_improvement': grade_stats.needs_improvement_count or 0
            }
        except Exception as e:
            grade_distribution = {
                'total_grades': 0,
                'average_grade': 0,
                'excellent': 0,
                'good': 0,
                'satisfactory': 0,
                'needs_improvement': 0
            }
        
        # Subject Statistics
        try:
            subject_query = text("""
                SELECT 
                    COUNT(*) as total_subjects,
                    COUNT(DISTINCT department) as departments_with_subjects
                FROM subject
            """)
            
            subject_stats = db.session.execute(subject_query).fetchone()
            
            subject_info = {
                'total_subjects': subject_stats.total_subjects or 0,
                'departments_with_subjects': subject_stats.departments_with_subjects or 0
            }
        except Exception as e:
            subject_info = {
                'total_subjects': 0,
                'departments_with_subjects': 0
            }
        
        # Class Assignment Statistics
        try:
            class_query = text("""
                SELECT 
                    COUNT(*) as total_assignments,
                    COUNT(DISTINCT instructor_id) as instructors_with_classes,
                    COUNT(DISTINCT subject_id) as subjects_with_assignments
                FROM class_assignment
            """)
            
            class_stats = db.session.execute(class_query).fetchone()
            
            class_info = {
                'total_assignments': class_stats.total_assignments or 0,
                'instructors_with_classes': class_stats.instructors_with_classes or 0,
                'subjects_with_assignments': class_stats.subjects_with_assignments or 0
            }
        except Exception as e:
            class_info = {
                'total_assignments': 0,
                'instructors_with_classes': 0,
                'subjects_with_assignments': 0
            }
        
        # System Health Metrics
        system_health = {
            'database_status': 'healthy',
            'total_tables': 9,  # Based on the database schema
            'last_backup': '2025-10-13 03:45:00',  # This could be dynamic
            'uptime': '99.9%',
            'response_time': '< 100ms'
        }
        
        # Generate report data
        report_data = {
            'generated_at': now.isoformat(),
            'period': 'Last 30 Days',
            'user_statistics': {
                'total_users': total_users,
                'active_users': active_users,
                'inactive_users': inactive_users,
                'role_distribution': role_distribution,
                'department_distribution': department_distribution,
                'year_distribution': year_distribution
            },
            'academic_statistics': {
                'grade_distribution': grade_distribution,
                'subject_info': subject_info,
                'class_info': class_info
            },
            'system_activity': {
                'daily_activity': daily_activity,
                'recent_logins': [{'date': str(date), 'count': count} for date, count in recent_logins]
            },
            'system_health': system_health,
            'recommendations': [
                'Monitor inactive user accounts',
                'Review grade encoding schedules',
                'Ensure regular database backups',
                'Check system performance metrics'
            ]
        }
        
        return jsonify({
            'status': 'success',
            'data': report_data
        })
        
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'Error generating report: {str(e)}'
        }), 500

@app.route('/api/database-backup', methods=['POST'])
@login_required
def database_backup():
    """Create and download database backup (MIS/IT only)"""
    if current_user.role != 'mis_it':
        return jsonify({'status': 'error', 'message': 'Unauthorized access'}), 403
    
    try:
        import tempfile
        import os
        from datetime import datetime
        from io import StringIO
        
        # Get database configuration from app config
        db_uri = app.config['SQLALCHEMY_DATABASE_URI']
        
        # Parse database URI (format: mysql+pymysql://user:pass@host/dbname)
        if 'mysql' in db_uri:
            # Extract database credentials
            parts = db_uri.split('://')[-1]  # Remove mysql+pymysql://
            credentials, host_db = parts.split('@')
            username, password = credentials.split(':')
            host_dbname = host_db.split('/')
            host = host_dbname[0].split(':')[0]  # Remove port if exists
            port = int(host_dbname[0].split(':')[1]) if ':' in host_dbname[0] else 3306
            database = host_dbname[1].split('?')[0] if '?' in host_dbname[1] else host_dbname[1]
            
            # Create timestamp for filename
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            
            # Create temporary file for backup
            temp_file = tempfile.NamedTemporaryFile(mode='w', suffix='.sql', delete=False, encoding='utf-8')
            temp_path = temp_file.name
            
            try:
                # Connect to database using pymysql
                connection = pymysql.connect(
                    host=host,
                    port=port,
                    user=username,
                    password=password,
                    database=database,
                    charset='utf8mb4'
                )
                
                cursor = connection.cursor()
                
                # Write SQL header
                temp_file.write(f"-- Acadify Database Backup\n")
                temp_file.write(f"-- Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                temp_file.write(f"-- Database: {database}\n")
                temp_file.write(f"--\n\n")
                temp_file.write(f"SET FOREIGN_KEY_CHECKS=0;\n")
                temp_file.write(f"SET SQL_MODE='NO_AUTO_VALUE_ON_ZERO';\n\n")
                
                # Get all tables
                cursor.execute("SHOW TABLES")
                tables = cursor.fetchall()
                
                for (table_name,) in tables:
                    # Get CREATE TABLE statement
                    cursor.execute(f"SHOW CREATE TABLE `{table_name}`")
                    create_table = cursor.fetchone()[1]
                    
                    temp_file.write(f"--\n-- Table structure for table `{table_name}`\n--\n\n")
                    temp_file.write(f"DROP TABLE IF EXISTS `{table_name}`;\n")
                    temp_file.write(create_table + ";\n\n")
                    
                    # Get table data
                    cursor.execute(f"SELECT * FROM `{table_name}`")
                    rows = cursor.fetchall()
                    
                    if rows:
                        temp_file.write(f"--\n-- Dumping data for table `{table_name}`\n--\n\n")
                        
                        # Get column names
                        cursor.execute(f"SHOW COLUMNS FROM `{table_name}`")
                        columns = [col[0] for col in cursor.fetchall()]
                        
                        # Insert data in batches
                        batch_size = 100
                        for i in range(0, len(rows), batch_size):
                            batch = rows[i:i + batch_size]
                            
                            temp_file.write(f"INSERT INTO `{table_name}` (`{'`, `'.join(columns)}`) VALUES\n")
                            
                            for idx, row in enumerate(batch):
                                # Escape and format values
                                values = []
                                for val in row:
                                    if val is None:
                                        values.append('NULL')
                                    elif isinstance(val, (int, float)):
                                        values.append(str(val))
                                    elif isinstance(val, datetime):
                                        values.append(f"'{val.strftime('%Y-%m-%d %H:%M:%S')}'")
                                    else:
                                        # Escape special characters
                                        escaped = str(val).replace('\\', '\\\\').replace("'", "\\'").replace('\n', '\\n').replace('\r', '\\r')
                                        values.append(f"'{escaped}'")
                                
                                comma = ',' if idx < len(batch) - 1 else ';'
                                temp_file.write(f"({', '.join(values)}){comma}\n")
                            
                            temp_file.write('\n')
                    
                    temp_file.write('\n')
                
                # Write footer
                temp_file.write(f"SET FOREIGN_KEY_CHECKS=1;\n")
                
                temp_file.close()
                cursor.close()
                connection.close()
                
                # Log the backup action
                log_audit_action(
                    user_id=current_user.id,
                    action='create_database_backup',
                    resource_type='system',
                    resource_id=None,
                    description=f'Database backup created ({len(tables)} tables)',
                    status='success',
                    request=request
                )
                
                # Send file and cleanup
                @after_this_request
                def cleanup(response):
                    try:
                        if os.path.exists(temp_path):
                            os.unlink(temp_path)
                    except Exception as e:
                        app.logger.error(f'Error cleaning up backup file: {e}')
                    return response
                
                return send_file(
                    temp_path,
                    mimetype='application/sql',
                    as_attachment=True,
                    download_name=f'acadify_backup_{timestamp}.sql'
                )
                
            except Exception as e:
                # Cleanup temp file on error
                temp_file.close()
                if os.path.exists(temp_path):
                    os.unlink(temp_path)
                raise e
                
        else:
            return jsonify({
                'status': 'error',
                'message': 'Unsupported database type'
            }), 400
            
    except Exception as e:
        app.logger.error(f'Backup error: {str(e)}')
        return jsonify({
            'status': 'error',
            'message': f'Backup failed: {str(e)}'
        }), 500

@app.route('/misit/audit-logs')
@login_required
def audit_logs():
    """Audit logs page for MIS/IT administrators"""
    if current_user.role != 'mis_it':
        flash('Access denied', 'error')
        return redirect(url_for('dashboard'))
    
    # Get pagination parameters
    page = request.args.get('page', 1, type=int)
    per_page = 50
    
    # Get filter parameters
    action_filter = request.args.get('action', '')
    user_filter = request.args.get('user', '')
    status_filter = request.args.get('status', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    
    # Build query
    query = AuditLog.query.outerjoin(User).order_by(AuditLog.created_at.desc())
    
    # Apply filters
    if action_filter:
        query = query.filter(AuditLog.action.contains(action_filter))
    if user_filter:
        query = query.filter(
            db.or_(
                User.username.contains(user_filter),
                User.first_name.contains(user_filter),
                User.last_name.contains(user_filter)
            )
        )
    if status_filter:
        query = query.filter(AuditLog.status == status_filter)
    if date_from:
        try:
            from datetime import datetime
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
            query = query.filter(AuditLog.created_at >= date_from_obj)
        except ValueError:
            pass
    if date_to:
        try:
            from datetime import datetime
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
            query = query.filter(AuditLog.created_at <= date_to_obj)
        except ValueError:
            pass
    
    # Get paginated results
    audit_logs = query.paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    # Get unique actions for filter dropdown
    unique_actions = db.session.query(AuditLog.action).distinct().all()
    unique_actions = [action[0] for action in unique_actions]
    
    return render_template('common/audit_logs.html',
                         audit_logs=audit_logs,
                         unique_actions=unique_actions,
                         current_filters={
                             'action': action_filter,
                             'user': user_filter,
                             'status': status_filter,
                             'date_from': date_from,
                             'date_to': date_to
                         })

# =====================================
# SUBJECT MANAGEMENT ROUTES
# =====================================

@app.route('/registrar/assign-subject', methods=['GET', 'POST'])
@login_required
def assign_subject():
    """Subject assignment page and functionality"""
    if current_user.role != 'registrar':
        flash('Access denied. Only registrars can assign subjects.', 'error')
        return redirect(url_for('dashboard'))
    
    # Get all instructors
    instructors = User.query.filter_by(role='instructor').all()
    # Get all subjects
    subjects = Subject.query.all()
    
    if request.method == 'POST':
        try:
            subject_id = request.form.get('subject_id')
            instructor_id = request.form.get('instructor_id')
            section = request.form.get('section', '').strip()
            school_year = request.form.get('school_year', '').strip()
            schedule_time = request.form.get('schedule_time', '').strip()
            schedule_day = request.form.get('schedule_day', '').strip()
            room = request.form.get('room', '').strip()
            subject_type = request.form.get('subject_type', 'Lecture').strip()
            assignment_id = request.form.get('assignment_id', '').strip()  # For editing existing assignments
            
            if not subject_id or not instructor_id or not section or not school_year:
                flash('Please fill in all required fields (Subject, Instructor, Section, Academic Year)', 'error')
                return redirect(url_for('assign_subject'))
            
            # Get subject details
            subject = Subject.query.get(subject_id)
            if not subject:
                flash('Subject not found', 'error')
                return redirect(url_for('assign_subject'))
            
            # Get instructor details
            instructor = User.query.get(instructor_id)
            if not instructor:
                flash('Instructor not found', 'error')
                return redirect(url_for('assign_subject'))
            
            # Use form data for section, semester, and academic year
            semester = subject.semester  # Use subject's semester
            
            # Check if assignment already exists for the same subject and section (excluding current assignment if editing)
            # This prevents multiple instructors from being assigned to the same subject-section combination
            query = ClassAssignment.query.filter_by(
                subject_id=subject_id,
                school_year=school_year,
                semester=semester,
                section=section
            )
            
            # If editing an existing assignment, exclude it from the duplicate check
            if assignment_id:
                query = query.filter(ClassAssignment.id != assignment_id)
            
            existing_assignment = query.first()
            
            if existing_assignment:
                # Check if it's the same instructor trying to assign to the same section
                if existing_assignment.instructor_id == int(instructor_id):
                    flash(f'You have already assigned this subject to this instructor for Section {section} in {school_year}', 'error')
                else:
                    flash(f'Section {section} already has an instructor assigned for this subject in {school_year}. Please choose a different section or academic year.', 'error')
                return redirect(url_for('assign_subject'))
            
            # Handle assignment creation or update
            if assignment_id:
                # Update existing assignment
                assignment = ClassAssignment.query.get(assignment_id)
                if not assignment:
                    flash('Assignment not found', 'error')
                    return redirect(url_for('assign_subject'))
                
                # Update assignment details
                assignment.instructor_id = instructor_id
                assignment.school_year = school_year
                assignment.section = section
                assignment.schedule_time = schedule_time if schedule_time else None
                assignment.schedule_day = schedule_day if schedule_day else None
                assignment.room = room if room else None
                assignment.subject_type = subject_type
                
                # Log the assignment update
                audit_log = AuditLog(
                    user_id=current_user.id,
                    action='UPDATE_ASSIGNMENT',
                    description=f'Updated assignment for subject {subject.subject_code} to instructor {instructor.first_name} {instructor.last_name} for Section {section}',
                    ip_address=request.remote_addr,
                    user_agent=request.headers.get('User-Agent', '')
                )
                db.session.add(audit_log)
                
                flash(f'Assignment updated successfully for Section {section} ({school_year})', 'success')
            else:
                # Create new class assignment
                assignment = ClassAssignment(
                    subject_id=subject_id,
                    instructor_id=instructor_id,
                    school_year=school_year,
                    semester=semester,
                    section=section,
                    schedule_time=schedule_time if schedule_time else None,
                    schedule_day=schedule_day if schedule_day else None,
                    room=room if room else None,
                    subject_type=subject_type
                )
                
                db.session.add(assignment)
                
                # Update subject's instructor_id for backward compatibility (keep the first assignment)
                if not subject.instructor_id:
                    subject.instructor_id = instructor_id
                
                # Log the assignment creation
                audit_log = AuditLog(
                    user_id=current_user.id,
                    action='ASSIGN_SUBJECT',
                    description=f'Assigned subject {subject.subject_code} to instructor {instructor.first_name} {instructor.last_name} for Section {section}',
                    ip_address=request.remote_addr,
                    user_agent=request.headers.get('User-Agent', '')
                )
                db.session.add(audit_log)
                
                flash(f'Subject successfully assigned to instructor for Section {section} ({school_year})', 'success')
            
            db.session.commit()
            return redirect(url_for('assign_subject', success='true'))
                
        except Exception as e:
            db.session.rollback()
            flash(f'Error assigning subject: {str(e)}', 'error')
            
        return redirect(url_for('assign_subject'))
    
    # Get all instructors
    instructors = User.query.filter_by(role='instructor').all()
    # Get all subjects
    subjects = Subject.query.all()
    # Get all class assignments
    class_assignments = ClassAssignment.query.join(Subject).join(User).all()
    
    # Get unique sections from students table
    sections_query = db.session.query(Student.section).filter(
        Student.section.isnot(None),
        Student.section != ''
    ).distinct().order_by(Student.section).all()
    sections = [section[0] for section in sections_query]
    
    return render_template('registrar/registrar_assignsub.html', 
                         user=current_user,
                         instructors=instructors,
                         subjects=subjects,
                         class_assignments=class_assignments,
                         sections=sections)

@app.route('/api/unassign-subject', methods=['POST'])
@login_required
def unassign_subject():
    """Unassign subject from instructor (Registrar only)"""
    if current_user.role != 'registrar':
        return jsonify({'status': 'error', 'message': 'Unauthorized access'}), 403
    
    try:
        data = request.get_json()
        assignment_id = data.get('assignment_id')
        
        if not assignment_id:
            return jsonify({'status': 'error', 'message': 'Assignment ID is required'}), 400
        
        # Find and delete the assignment
        assignment = ClassAssignment.query.get(assignment_id)
        if not assignment:
            return jsonify({'status': 'error', 'message': 'Assignment not found'}), 404
        
        # Get assignment details for response
        subject_code = assignment.subject.subject_code
        instructor_name = f"{assignment.instructor.first_name} {assignment.instructor.last_name}"
        section = assignment.section or "General"
        
        # Delete the assignment
        db.session.delete(assignment)
        db.session.commit()
        
        return jsonify({
            'status': 'success',
            'message': f'Assignment removed: {subject_code} - {instructor_name} ({section})'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/unassign-subjects-batch', methods=['POST'])
@login_required
def unassign_subjects_batch():
    """Unassign multiple subjects from instructors (Registrar only)"""
    if current_user.role != 'registrar':
        return jsonify({'status': 'error', 'message': 'Unauthorized access'}), 403
    
    try:
        data = request.get_json()
        assignment_ids = data.get('assignment_ids', [])
        
        if not assignment_ids or not isinstance(assignment_ids, list):
            return jsonify({'status': 'error', 'message': 'Assignment IDs are required'}), 400
        
        # Find all assignments
        assignments = ClassAssignment.query.filter(ClassAssignment.id.in_(assignment_ids)).all()
        
        if not assignments:
            return jsonify({'status': 'error', 'message': 'No assignments found'}), 404
        
        # Collect details for response
        removed_details = []
        for assignment in assignments:
            subject_code = assignment.subject.subject_code
            instructor_name = f"{assignment.instructor.first_name} {assignment.instructor.last_name}"
            section = assignment.section or "General"
            removed_details.append(f"{subject_code} - {instructor_name} ({section})")
        
        # Delete all assignments
        for assignment in assignments:
            db.session.delete(assignment)
        
        db.session.commit()
        
        return jsonify({
            'status': 'success',
            'message': f'Successfully removed {len(assignments)} assignment(s)',
            'removed': removed_details
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/get-enrollment-updates', methods=['GET'])
@login_required
def get_enrollment_updates():
    """Get recent student enrollment updates for real-time dashboard updates"""
    if current_user.role not in ['instructor', 'registrar']:
        return jsonify({'status': 'error', 'message': 'Access denied'}), 403
    
    try:
        # Get recent StudentSubject enrollments (last 24 hours)
        from datetime import datetime, timedelta
        recent_time = datetime.utcnow() - timedelta(hours=24)
        
        recent_enrollments = StudentSubject.query.filter(
            StudentSubject.enrollment_date >= recent_time,
            StudentSubject.status == 'ENROLLED'
        ).all()
        
        enrollment_data = []
        for enrollment in recent_enrollments:
            subject = Subject.query.get(enrollment.subject_id)
            student = Student.query.get(enrollment.student_id)
            
            if subject and student:
                enrollment_data.append({
                    'id': enrollment.id,
                    'student_name': f"{student.first_name} {student.last_name}",
                    'student_id': student.student_id,
                    'subject_code': subject.subject_code,
                    'subject_name': subject.subject_name,
                    'enrollment_date': enrollment.enrollment_date.isoformat(),
                    'academic_year': enrollment.academic_year,
                    'semester': enrollment.semester
                })
        
        return jsonify({
            'status': 'success',
            'enrollments': enrollment_data,
            'count': len(enrollment_data)
        })
        
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/instructor-subject-counts', methods=['GET'])
@login_required
def get_instructor_subject_counts():
    """Get current student counts for each subject-section combination for the instructor"""
    if current_user.role not in ['instructor', 'registrar']:
        return jsonify({'status': 'error', 'message': 'Access denied'}), 403
    
    try:
        subject_counts = {}
        
        if current_user.role == 'instructor':
            # Get subjects assigned to this instructor - same approach as main dashboard
            instructor_subjects_query = db.session.query(Subject, ClassAssignment).join(
                ClassAssignment, Subject.id == ClassAssignment.subject_id
            ).filter(
                ClassAssignment.instructor_id == current_user.id
            ).all()
            
            # Create subject assignments dict (same structure as main dashboard)
            subjects_with_assignments = []
            for subject, assignment in instructor_subjects_query:
                unique_key = f"{subject.id}_{assignment.section or 'General'}_{assignment.school_year}"
                subjects_with_assignments.append({
                    'subject': subject,
                    'assignment': assignment,
                    'unique_key': unique_key
                })
            
            subject_assignments_dict = {item['unique_key']: item['assignment'] for item in subjects_with_assignments}
            
            # Count students exactly like the main dashboard
            for unique_key, assignment in subject_assignments_dict.items():
                subject = Subject.query.get(assignment.subject_id)
                if not subject:
                    continue
                
                # Try multiple academic year combinations (same as main dashboard)
                academic_years_to_try = []
                if assignment.school_year:
                    academic_years_to_try.append(assignment.school_year)
                if subject.academic_year:
                    academic_years_to_try.append(subject.academic_year)
                academic_years_to_try = list(dict.fromkeys(academic_years_to_try))
                
                # Count enrolled students for this subject
                enrolled_students = StudentSubject.query.filter(
                    StudentSubject.subject_id == subject.id,
                    StudentSubject.semester == subject.semester,
                    StudentSubject.status == 'ENROLLED'
                )
                
                if academic_years_to_try:
                    enrolled_students = enrolled_students.filter(StudentSubject.academic_year.in_(academic_years_to_try))
                
                enrolled_students = enrolled_students.all()
                
                # Filter by section if this is for a specific assignment with a section
                # Count all enrolled students first (including active check)
                student_ids = [es.student_id for es in enrolled_students]
                if student_ids:
                    active_students = Student.query.filter(
                        Student.id.in_(student_ids),
                        Student.active == True
                    ).all()
                    enrolled_count = len(active_students)
                else:
                    enrolled_count = 0
                
                # Further filter by section if assignment has a section
                if assignment.section and student_ids:
                    # Filter by section
                    section_students = Student.query.filter(
                        Student.id.in_(student_ids),
                        Student.section == assignment.section
                    ).all()
                    # If students found in the specific section, use that count
                    if section_students:
                        enrolled_count = len(section_students)
                    # Otherwise, keep the count from all active students (already set above)
                
                subject_counts[unique_key] = {
                    'enrolled_count': enrolled_count
                }
        else:
            # For registrars, get counts for all subjects with enrollments
            subjects_with_enrollments = db.session.query(Subject).join(
                StudentSubject, Subject.id == StudentSubject.subject_id
            ).filter(
                StudentSubject.status == 'ENROLLED'
            ).distinct().all()
            
            for subject in subjects_with_enrollments:
                # Try multiple academic year combinations
                academic_years_to_try = []
                if subject.academic_year:
                    academic_years_to_try.append(subject.academic_year)
                
                # Count enrolled students
                enrolled_students = StudentSubject.query.filter(
                    StudentSubject.subject_id == subject.id,
                    StudentSubject.semester == subject.semester,
                    StudentSubject.status == 'ENROLLED'
                )
                
                if academic_years_to_try:
                    enrolled_students = enrolled_students.filter(StudentSubject.academic_year.in_(academic_years_to_try))
                
                enrolled_students = enrolled_students.all()
                enrolled_count = len(enrolled_students)
                
                subject_counts[str(subject.id)] = {
                    'enrolled_count': enrolled_count
                }
        
        return jsonify({
            'status': 'success',
            'counts': subject_counts
        })
        
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

# =====================================
# GRADE ENCODING SCHEDULE ROUTES
# =====================================

def check_duplicate_schedule(academic_year, semester, grading_period, department, start_date, end_date, exclude_id=None):
    """Check if a schedule with the same parameters already exists"""
    query = GradeEncodingSchedule.query.filter(
        GradeEncodingSchedule.academic_year == academic_year,
        GradeEncodingSchedule.semester == semester,
        GradeEncodingSchedule.grading_period == grading_period,
        GradeEncodingSchedule.start_date == start_date,
        GradeEncodingSchedule.end_date == end_date,
        GradeEncodingSchedule.status != 'completed'
    )
    
    # Add department filter
    if department:
        query = query.filter(GradeEncodingSchedule.department == department)
    else:
        query = query.filter(GradeEncodingSchedule.department.is_(None))
    
    # Exclude current schedule when updating
    if exclude_id:
        query = query.filter(GradeEncodingSchedule.id != exclude_id)
    
    return query.first()

def update_schedule_statuses():
    """Update schedule statuses based on current date and time"""
    from datetime import datetime
    
    now = datetime.now()
    today = now.date()
    current_time = now.time()
    
    # Get all non-completed schedules
    schedules = GradeEncodingSchedule.query.filter(
        GradeEncodingSchedule.status != 'completed'
    ).all()
    
    updated = False
    for schedule in schedules:
        schedule_start_date = schedule.start_date.date() if isinstance(schedule.start_date, datetime) else schedule.start_date
        schedule_end_date = schedule.end_date.date() if isinstance(schedule.end_date, datetime) else schedule.end_date
        
        # Check if schedule has time constraints (for testing)
        if schedule.start_time and schedule.end_time:
            # Time-based checking
            if schedule_start_date == today:
                if current_time >= schedule.start_time and current_time <= schedule.end_time:
                    if schedule.status != 'active':
                        schedule.status = 'active'
                        updated = True
                elif current_time > schedule.end_time:
                    if schedule.status != 'completed':
                        schedule.status = 'completed'
                        updated = True
            elif schedule_start_date < today or (schedule_start_date == today and current_time > schedule.end_time):
                if schedule.status != 'completed':
                    schedule.status = 'completed'
                    updated = True
        else:
            # Date-based checking (no time constraints)
            if schedule_end_date < today:
                # Schedule has ended
                if schedule.status != 'completed':
                    schedule.status = 'completed'
                    updated = True
            elif schedule_start_date <= today <= schedule_end_date:
                # Schedule is currently active
                if schedule.status != 'active':
                    schedule.status = 'active'
                    updated = True
            elif schedule_start_date > today:
                # Schedule hasn't started yet
                if schedule.status != 'upcoming':
                    schedule.status = 'upcoming'
                    updated = True
    
    if updated:
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            print(f"Error updating schedule statuses: {str(e)}")
    
    return updated

def update_encoding_exceptions():
    """Automatically expire encoding exceptions that have passed their expiration date"""
    from datetime import datetime
    
    now = datetime.now()
    
    # Get all active exceptions
    active_exceptions = EncodingException.query.filter_by(is_active=True).all()
    
    expired_count = 0
    for exception in active_exceptions:
        exception_date = exception.expiration_date.date() if isinstance(exception.expiration_date, datetime) else exception.expiration_date
        
        # Check if exception has expired
        if exception.expiration_date < now:
            exception.is_active = False
            exception.revoked_at = now
            expired_count += 1
    
    if expired_count > 0:
        try:
            db.session.commit()
            print(f"[INFO] Auto-expired {expired_count} encoding exception(s)")
        except Exception as e:
            db.session.rollback()
            print(f"Error expiring encoding exceptions: {str(e)}")
    
    return expired_count

@app.route('/misit/grade-encoding-schedule', methods=['GET', 'POST'])
@login_required
def grade_encoding_schedule():
    """Manage grade encoding schedules"""
    if current_user.role != 'mis_it':
        flash('Access denied. Only MIS/IT administrators can manage grade encoding schedules.', 'error')
        return redirect(url_for('dashboard'))

    # Update schedule statuses and encoding exceptions before processing request
    update_schedule_statuses()
    update_encoding_exceptions()

    if request.method == 'POST':
        try:
            academic_year = request.form.get('academic_year')
            semester = int(request.form.get('semester'))
            department = request.form.get('department', '').strip()
            if not department:  # Convert empty string to None
                department = None
            start_date = datetime.strptime(request.form.get('start_date'), '%Y-%m-%d')
            end_date = datetime.strptime(request.form.get('end_date'), '%Y-%m-%d')
            status = request.form.get('status', 'active')
            grading_period = request.form.get('grading_period', 'all')
            
            # Handle time fields for testing
            start_time_str = request.form.get('start_time', '').strip()
            end_time_str = request.form.get('end_time', '').strip()
            start_time = None
            end_time = None
            
            if start_time_str:
                start_time = datetime.strptime(start_time_str, '%H:%M').time()
            if end_time_str:
                end_time = datetime.strptime(end_time_str, '%H:%M').time()

            # Print debug information
            print(f"Creating schedule: Year={academic_year}, Sem={semester}, Dept={department}, Status={status}, Period={grading_period}")

            # Validate dates
            if start_date >= end_date:
                flash('End date must be after start date', 'error')
                return redirect(url_for('grade_encoding_schedule'))
            
            # Validate against current date
            today = datetime.now().date()
            if start_date.date() < today:
                flash('Start date cannot be in the past', 'error')
                return redirect(url_for('grade_encoding_schedule'))
            
            # Auto-adjust status based on dates
            if start_date.date() == today and status == 'upcoming':
                status = 'active'  # If starting today, make it active
                flash('Schedule status automatically set to "Active" since it starts today', 'info')

            # Check for exact duplicate schedule
            duplicate = check_duplicate_schedule(academic_year, semester, grading_period, department, start_date.date(), end_date.date())
            if duplicate:
                dept_text = f" for {department}" if department else " for all departments"
                flash(f'A schedule with the same academic year, semester, grading period, department, and dates already exists{dept_text}.', 'error')
                return redirect(url_for('grade_encoding_schedule'))

            # Check for overlapping schedules with same grading period
            query = GradeEncodingSchedule.query.filter(
                GradeEncodingSchedule.academic_year == academic_year,
                GradeEncodingSchedule.semester == semester,
                GradeEncodingSchedule.grading_period == grading_period,
                GradeEncodingSchedule.start_date <= end_date,
                GradeEncodingSchedule.end_date >= start_date,
                GradeEncodingSchedule.status != 'completed'
            )

            # Add department filter if specified
            if department:
                query = query.filter(
                    db.or_(
                        GradeEncodingSchedule.department == department,
                        GradeEncodingSchedule.department.is_(None)
                    )
                )
            else:
                query = query.filter(GradeEncodingSchedule.department.is_(None))

            overlapping = query.first()

            if overlapping:
                flash(f'An overlapping schedule already exists for {grading_period} period.', 'error')
                return redirect(url_for('grade_encoding_schedule'))

            # If setting as active, complete other active schedules for the same grading period
            if status == 'active':
                active_query = GradeEncodingSchedule.query.filter(
                    GradeEncodingSchedule.status == 'active',
                    GradeEncodingSchedule.academic_year == academic_year,
                    GradeEncodingSchedule.semester == semester,
                    GradeEncodingSchedule.grading_period == grading_period
                )
                
                if department:
                    active_query = active_query.filter(
                        db.or_(
                            GradeEncodingSchedule.department == department,
                            GradeEncodingSchedule.department.is_(None)
                        )
                    )
                else:
                    active_query = active_query.filter(GradeEncodingSchedule.department.is_(None))
                
                active_schedules = active_query.all()
                
                for schedule in active_schedules:
                    schedule.status = 'completed'
                    db.session.add(schedule)

            new_schedule = GradeEncodingSchedule(
                academic_year=academic_year,
                semester=semester,
                department=department,
                grading_period=grading_period,
                start_date=start_date.replace(hour=0, minute=0, second=0),
                end_date=end_date.replace(hour=23, minute=59, second=59),
                start_time=start_time,
                end_time=end_time,
                status=status,
                created_by=current_user.id
            )

            db.session.add(new_schedule)
            db.session.commit()
            flash('Grade encoding schedule has been set successfully.', 'success')
            
        except Exception as e:
            db.session.rollback()
            flash('An error occurred while setting the schedule.', 'error')
            print(f"Error: {str(e)}")

        return redirect(url_for('grade_encoding_schedule'))

    # Get all schedules
    schedules = GradeEncodingSchedule.query.order_by(
        GradeEncodingSchedule.academic_year.desc(),
        GradeEncodingSchedule.semester.desc(),
        GradeEncodingSchedule.start_date.desc()
    ).all()
    
    # Get all instructors for the restriction modal
    instructors = User.query.filter_by(role='instructor', active=True).order_by(User.last_name, User.first_name).all()
    
    # Get active encoding exceptions
    exceptions = EncodingException.query.filter_by(is_active=True).order_by(EncodingException.created_at.desc()).all()
    
    # Get recently expired exceptions (within last 30 days)
    from datetime import timedelta
    thirty_days_ago = datetime.now() - timedelta(days=30)
    expired_exceptions = EncodingException.query.filter(
        EncodingException.is_active == False,
        EncodingException.revoked_at >= thirty_days_ago
    ).order_by(EncodingException.revoked_at.desc()).all()

    return render_template('misit/misit_assignsched.html', 
                         user=current_user,
                         schedules=schedules,
                         instructors=instructors,
                         exceptions=exceptions,
                         expired_exceptions=expired_exceptions)

@app.route('/api/grade-encoding-schedule/check-updates', methods=['GET'])
@login_required
def check_schedule_updates():
    """Check and update schedule statuses and encoding exceptions in real-time (for auto-refresh)"""
    if current_user.role != 'mis_it':
        return jsonify({'status': 'error', 'message': 'Unauthorized access'}), 403
    
    try:
        # Update statuses and exceptions
        schedules_updated = update_schedule_statuses()
        exceptions_expired = update_encoding_exceptions()
        
        # Return current schedule counts
        active_count = GradeEncodingSchedule.query.filter_by(status='active').count()
        upcoming_count = GradeEncodingSchedule.query.filter_by(status='upcoming').count()
        completed_count = GradeEncodingSchedule.query.filter_by(status='completed').count()
        
        # Return current exception counts
        active_exceptions = EncodingException.query.filter_by(is_active=True).count()
        expired_exceptions = EncodingException.query.filter_by(is_active=False).count()
        
        return jsonify({
            'status': 'success',
            'updated': schedules_updated or exceptions_expired > 0,
            'schedules_updated': schedules_updated,
            'exceptions_expired': exceptions_expired,
            'counts': {
                'active': active_count,
                'upcoming': upcoming_count,
                'completed': completed_count
            },
            'exceptions': {
                'active': active_exceptions,
                'expired': expired_exceptions
            },
            'timestamp': datetime.now().isoformat()
        })
        
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/grade-encoding-schedule/<int:schedule_id>/close', methods=['POST'])
@login_required
def close_grade_encoding_schedule(schedule_id):
    """Close a grade encoding schedule (MIS/IT only)"""
    if current_user.role != 'mis_it':
        return jsonify({'status': 'error', 'message': 'Unauthorized access'}), 403
    
    try:
        schedule = GradeEncodingSchedule.query.get_or_404(schedule_id)
        
        if schedule.status == 'completed':
            return jsonify({'status': 'error', 'message': 'Schedule is already closed'}), 400
        
        # Mark schedule as completed
        schedule.status = 'completed'
        schedule.updated_at = datetime.utcnow()
        
        db.session.commit()
        
        return jsonify({
            'status': 'success',
            'message': 'Grade encoding schedule closed successfully'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/debug-schedules', methods=['GET'])
@login_required
def debug_schedules():
    """Debug endpoint to show all schedules"""
    if current_user.role != 'mis_it':
        return jsonify({'status': 'error', 'message': 'Unauthorized access'}), 403
    
    try:
        schedules = GradeEncodingSchedule.query.filter(GradeEncodingSchedule.status != 'completed').all()
        schedule_list = []
        for sched in schedules:
            schedule_list.append({
                'id': sched.id,
                'academic_year': sched.academic_year,
                'semester': sched.semester,
                'grading_period': sched.grading_period,
                'department': sched.department,
                'start_date': str(sched.start_date),
                'end_date': str(sched.end_date),
                'status': sched.status
            })
        
        return jsonify({
            'status': 'success',
            'schedules': schedule_list,
            'count': len(schedule_list)
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/grade-encoding-schedule/<int:schedule_id>', methods=['PUT'])
@login_required
def update_grade_encoding_schedule(schedule_id):
    """Update a grade encoding schedule"""
    if current_user.role != 'mis_it':
        return jsonify({'status': 'error', 'message': 'Unauthorized access'}), 403
    
    try:
        schedule = GradeEncodingSchedule.query.get_or_404(schedule_id)
        
        # Get data from request
        data = request.get_json()
        
        # Validate required fields
        required_fields = ['academic_year', 'semester', 'grading_period', 'status', 'start_date', 'end_date']
        for field in required_fields:
            if field not in data or not data[field]:
                return jsonify({'status': 'error', 'message': f'{field} is required'}), 400
        
        # Parse dates for validation
        new_start_date = datetime.strptime(data['start_date'], '%Y-%m-%d').date()
        new_end_date = datetime.strptime(data['end_date'], '%Y-%m-%d').date()
        
        # Validate dates
        if new_start_date >= new_end_date:
            return jsonify({'status': 'error', 'message': 'End date must be after start date'}), 400
        
        # Validate against current date
        today = datetime.now().date()
        if new_start_date < today:
            return jsonify({'status': 'error', 'message': 'Start date cannot be in the past'}), 400
        
        # Auto-adjust status based on dates
        if new_start_date == today and data['status'] == 'upcoming':
            data['status'] = 'active'  # If starting today, make it active
        
        # Check for exact duplicate schedule (excluding current schedule) BEFORE updating
        duplicate = check_duplicate_schedule(
            data['academic_year'], 
            int(data['semester']), 
            data['grading_period'], 
            data.get('department'), 
            new_start_date,  # Use the new dates from form
            new_end_date,    # Use the new dates from form
            exclude_id=schedule_id
        )
        if duplicate:
            dept_text = f" for {data.get('department')}" if data.get('department') else " for all departments"
            return jsonify({'status': 'error', 'message': f'A schedule with the same academic year, semester, grading period, department, and dates already exists{dept_text}.'}), 400
        
        # Update schedule fields
        schedule.academic_year = data['academic_year']
        schedule.semester = int(data['semester'])
        schedule.grading_period = data['grading_period']
        schedule.status = data['status']
        schedule.department = data.get('department')
        schedule.start_date = new_start_date
        schedule.end_date = new_end_date
        
        # Handle optional time fields
        if data.get('start_time'):
            schedule.start_time = datetime.strptime(data['start_time'], '%H:%M').time()
        else:
            schedule.start_time = None
            
        if data.get('end_time'):
            schedule.end_time = datetime.strptime(data['end_time'], '%H:%M').time()
        else:
            schedule.end_time = None
        
        db.session.commit()
        return jsonify({'status': 'success', 'message': 'Schedule updated successfully'})
        
    except ValueError as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': f'Invalid date format: {str(e)}'}), 400
    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/grade-encoding-schedule/<int:schedule_id>', methods=['DELETE'])
@login_required
def delete_grade_encoding_schedule(schedule_id):
    """Delete a grade encoding schedule"""
    if current_user.role != 'mis_it':
        return jsonify({'status': 'error', 'message': 'Unauthorized access'}), 403
    
    try:
        schedule = GradeEncodingSchedule.query.get_or_404(schedule_id)
        if schedule.status == 'active':
            return jsonify({'status': 'error', 'message': 'Cannot delete an active schedule'}), 400
            
        db.session.delete(schedule)
        db.session.commit()
        return jsonify({'status': 'success', 'message': 'Schedule deleted successfully'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/encoding-exception/grant', methods=['POST'])
@login_required
def grant_encoding_exception():
    """Grant encoding access to an instructor after schedule closure (MIS/IT only)"""
    if current_user.role != 'mis_it':
        return jsonify({'status': 'error', 'message': 'Unauthorized access'}), 403
    
    try:
        data = request.get_json()
        instructor_id = int(data.get('instructor_id'))
        academic_year = data.get('academic_year')
        semester = int(data.get('semester'))
        grading_period = data.get('grading_period')
        expiration_date_str = data.get('expiration_date', '').strip()
        expiration_time_str = data.get('expiration_time', '').strip()
        reason = data.get('reason', '').strip() or None
        
        # Validate: at least date or time must be provided
        if not expiration_date_str and not expiration_time_str:
            return jsonify({'status': 'error', 'message': 'Please specify at least an expiration date or time'}), 400
        
        # Determine expiration datetime based on what was provided
        now = datetime.now()
        
        if expiration_date_str and expiration_time_str:
            # Both date and time provided
            expiration_date = datetime.strptime(expiration_date_str, '%Y-%m-%d')
            expiration_time = datetime.strptime(expiration_time_str, '%H:%M').time()
            expiration_date = expiration_date.replace(
                hour=expiration_time.hour,
                minute=expiration_time.minute,
                second=59
            )
        elif expiration_date_str:
            # Date only - default to 11:59 PM on that date
            expiration_date = datetime.strptime(expiration_date_str, '%Y-%m-%d')
            expiration_date = expiration_date.replace(hour=23, minute=59, second=59)
        else:
            # Time only - use today's date with specified time
            expiration_time = datetime.strptime(expiration_time_str, '%H:%M').time()
            expiration_date = now.replace(
                hour=expiration_time.hour,
                minute=expiration_time.minute,
                second=59,
                microsecond=0
            )
            
            # If the time has already passed today, it's invalid
            if expiration_date <= now:
                return jsonify({'status': 'error', 'message': 'Time-only expiration must be in the future (later today)'}), 400
        
        # Validate instructor exists and is an instructor
        instructor = User.query.get(instructor_id)
        if not instructor or instructor.role != 'instructor':
            return jsonify({'status': 'error', 'message': 'Invalid instructor selected'}), 400
        
        # Check if exception already exists for this instructor and period
        existing = EncodingException.query.filter_by(
            instructor_id=instructor_id,
            academic_year=academic_year,
            semester=semester,
            grading_period=grading_period,
            is_active=True
        ).first()
        
        if existing:
            return jsonify({
                'status': 'error', 
                'message': f'An active exception already exists for {instructor.first_name} {instructor.last_name} for this period'
            }), 400
        
        # Create new exception
        exception = EncodingException(
            instructor_id=instructor_id,
            academic_year=academic_year,
            semester=semester,
            grading_period=grading_period,
            expiration_date=expiration_date,
            reason=reason,
            granted_by=current_user.id
        )
        
        db.session.add(exception)
        db.session.commit()
        
        # Log the action
        log_audit_action(
            user_id=current_user.id,
            action='grant_encoding_exception',
            resource_type='encoding_exception',
            resource_id=exception.id,
            description=f"Granted encoding access to {instructor.first_name} {instructor.last_name} for {academic_year} Sem {semester} - {grading_period}",
            status='success',
            request=request
        )
        
        # Format success message based on what was provided
        if expiration_date_str and expiration_time_str:
            expiry_msg = expiration_date.strftime("%b %d, %Y at %I:%M %p")
        elif expiration_date_str:
            expiry_msg = expiration_date.strftime("%b %d, %Y at 11:59 PM")
        else:
            expiry_msg = f"today at {expiration_date.strftime('%I:%M %p')}"
        
        return jsonify({
            'status': 'success',
            'message': f'Encoding access granted to {instructor.first_name} {instructor.last_name} until {expiry_msg}'
        })
        
    except ValueError as e:
        return jsonify({'status': 'error', 'message': 'Invalid date format'}), 400
    except Exception as e:
        db.session.rollback()
        print(f"Error granting exception: {str(e)}")
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/encoding-exception/<int:exception_id>/revoke', methods=['POST'])
@login_required
def revoke_encoding_exception(exception_id):
    """Revoke an encoding exception (MIS/IT only)"""
    if current_user.role != 'mis_it':
        return jsonify({'status': 'error', 'message': 'Unauthorized access'}), 403
    
    try:
        exception = EncodingException.query.get_or_404(exception_id)
        
        if not exception.is_active:
            return jsonify({'status': 'error', 'message': 'This exception is already revoked'}), 400
        
        exception.is_active = False
        exception.revoked_at = datetime.utcnow()
        db.session.commit()
        
        # Log the action
        log_audit_action(
            user_id=current_user.id,
            action='revoke_encoding_exception',
            resource_type='encoding_exception',
            resource_id=exception.id,
            description=f"Revoked encoding access for {exception.instructor.first_name} {exception.instructor.last_name}",
            status='success',
            request=request
        )
        
        return jsonify({
            'status': 'success',
            'message': 'Encoding access revoked successfully'
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Error revoking exception: {str(e)}")
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/fix-academic-year', methods=['POST'])
@login_required
def fix_academic_year():
    """Fix academic year for all subjects (MIS/IT only)"""
    if current_user.role != 'mis_it':
        return jsonify({'status': 'error', 'message': 'Unauthorized access'}), 403
    
    try:
        # Update all subjects that have NULL or empty academic_year
        result = db.session.execute(text("UPDATE subject SET academic_year = '2025-2026' WHERE academic_year IS NULL OR academic_year = ''"))
        db.session.commit()
        
        updated_count = result.rowcount
        return jsonify({
            'status': 'success', 
            'message': f'Updated {updated_count} subjects with academic year 2025-2026'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500

# =====================================
# GRADE MANAGEMENT ROUTES
# =====================================

@app.route('/instructor/encode-grades/<int:subject_id>')
@app.route('/instructor/encode-grades/<int:subject_id>/<section>')
@app.route('/instructor/encode-grades/<int:subject_id>/<section>/<school_year>')
@login_required
def encode_grades(subject_id, section=None, school_year=None):
    """Grade encoding page for instructors and registrars"""
    if current_user.role not in ['instructor', 'registrar']:
        flash('Access denied. Only instructors and registrars can access grade encoding.', 'error')
        return redirect(url_for('dashboard'))
    
    # Get subject and verify access
    subject = Subject.query.get_or_404(subject_id)
    assignment = None
    if current_user.role == 'instructor':
        # Check if instructor is assigned to this subject via ClassAssignment
        # If section and school_year are provided, use them to find the specific assignment
        if section and school_year:
            assignment = ClassAssignment.query.filter_by(
                subject_id=subject_id,
                instructor_id=current_user.id,
                section=section,
                school_year=school_year
            ).first()
        else:
            # Fallback to first assignment if no specific section/year provided
            assignment = ClassAssignment.query.filter_by(
                subject_id=subject_id,
                instructor_id=current_user.id
            ).first()
        
        if not assignment:
            flash('Access denied. You are not assigned to this subject.', 'error')
            return redirect(url_for('instructor_dashboard'))
    elif current_user.role == 'registrar':
        # For registrars, get the assignment to show section info
        if section and school_year:
            assignment = ClassAssignment.query.filter_by(
                subject_id=subject_id,
                section=section,
                school_year=school_year
            ).first()
        else:
            assignment = ClassAssignment.query.filter_by(
                subject_id=subject_id
            ).first()
        # Registrars can access any subject's grade encoding
        pass
    
    # Get students actually enrolled in this specific subject via StudentSubject table
    # Try multiple academic year combinations to find enrolled students
    academic_years_to_try = []
    
    # Add assignment's school_year if available
    if assignment and assignment.school_year:
        academic_years_to_try.append(assignment.school_year)
    
    # Add subject's academic_year
    if subject.academic_year:
        academic_years_to_try.append(subject.academic_year)
    
    # Remove duplicates while preserving order
    academic_years_to_try = list(dict.fromkeys(academic_years_to_try))
    
    # Query for enrolled students with any of the possible academic years
    enrolled_students = StudentSubject.query.filter(
        StudentSubject.subject_id == subject_id,
        StudentSubject.academic_year.in_(academic_years_to_try),
        StudentSubject.semester == subject.semester,
        StudentSubject.status == 'ENROLLED'
    ).all()
    
    # If no students found with specific academic years, try without academic year filter
    if not enrolled_students:
        enrolled_students = StudentSubject.query.filter_by(
            subject_id=subject_id,
            semester=subject.semester,
            status='ENROLLED'
        ).all()
    
    # Debug information
    print(f"[DEBUG] Student enrollment search:")
    print(f"   Subject ID: {subject_id}")
    print(f"   Subject academic_year: {subject.academic_year}")
    print(f"   Assignment school_year: {assignment.school_year if assignment else 'No assignment'}")
    print(f"   Academic years to try: {academic_years_to_try}")
    print(f"   Found enrolled students: {len(enrolled_students)}")
    
    # Get the actual Student objects for enrolled students
    student_ids = [es.student_id for es in enrolled_students]
    
    # Filter students by section if assignment has a specific section
    students_query = Student.query.filter(
        Student.id.in_(student_ids),
        Student.active == True
    )
    
    if assignment and assignment.section:
        # Try filtering by section first
        section_students = students_query.filter(Student.section == assignment.section).all()
        print(f"   Filtering by section: {assignment.section}")
        print(f"   Students in section: {len(section_students)}")
        
        # If no students found in the specific section, include all enrolled students
        if section_students:
            students = section_students
        else:
            print(f"   No students found in section {assignment.section}, showing all enrolled students")
            students = students_query.order_by(Student.last_name, Student.first_name).all()
    else:
        students = students_query.order_by(Student.last_name, Student.first_name).all()
    
    print(f"   Final student count: {len(students)}")
    
    # Get current date
    now = datetime.now()
    
    # Determine the correct academic year to use for schedule matching and grade retrieval
    # Use assignment.school_year if available, otherwise fall back to subject.academic_year
    academic_year_for_schedule = subject.academic_year
    if assignment and assignment.school_year:
        academic_year_for_schedule = assignment.school_year
    
    # Get existing grades using the correct academic year
    grades = {grade.student_id: grade for grade in Grade.query.filter_by(
        subject_id=subject_id,
        semester=subject.semester,
        academic_year=academic_year_for_schedule
    ).all()}
    
    print(f"[DEBUG] Schedule matching - Subject ID: {subject_id}")
    print(f"   Subject academic_year: {subject.academic_year}")
    print(f"   Assignment school_year: {assignment.school_year if assignment else 'No assignment'}")
    print(f"   Using academic_year_for_schedule: {academic_year_for_schedule}")
    print(f"   Subject semester: {subject.semester}")
    print(f"   Subject department: {subject.department}")
    
    # Debug: Check what schedules exist
    all_schedules = GradeEncodingSchedule.query.filter(
        GradeEncodingSchedule.status == 'active'
    ).all()
    print(f"   All active schedules: {[(s.academic_year, s.semester, s.department, s.grading_period) for s in all_schedules]}")
    
    # Debug: Check enrolled students
    print(f"   Enrolled students count: {len(enrolled_students)}")
    print(f"   Students found: {len(students)}")
    
    # First try to find department-specific schedule for all grading periods
    schedule = GradeEncodingSchedule.query.filter(
        GradeEncodingSchedule.semester == subject.semester,
        GradeEncodingSchedule.academic_year == academic_year_for_schedule,
        GradeEncodingSchedule.department == subject.department,
        GradeEncodingSchedule.status == 'active',
        GradeEncodingSchedule.grading_period == 'all'
    ).first()
    
    # If no department-specific schedule, look for general schedule for all periods
    if not schedule:
        schedule = GradeEncodingSchedule.query.filter(
            GradeEncodingSchedule.semester == subject.semester,
            GradeEncodingSchedule.academic_year == academic_year_for_schedule,
            GradeEncodingSchedule.department == None,
            GradeEncodingSchedule.status == 'active',
            GradeEncodingSchedule.grading_period == 'all'
        ).first()
    
    # Get grading period specific schedules for display
    prelim_schedule = GradeEncodingSchedule.query.filter(
        GradeEncodingSchedule.semester == subject.semester,
        GradeEncodingSchedule.academic_year == academic_year_for_schedule,
        db.or_(
            GradeEncodingSchedule.department == subject.department,
            GradeEncodingSchedule.department == None
        ),
        GradeEncodingSchedule.status == 'active',
        GradeEncodingSchedule.grading_period == 'prelim'
    ).first()
    
    midterm_schedule = GradeEncodingSchedule.query.filter(
        GradeEncodingSchedule.semester == subject.semester,
        GradeEncodingSchedule.academic_year == academic_year_for_schedule,
        db.or_(
            GradeEncodingSchedule.department == subject.department,
            GradeEncodingSchedule.department == None
        ),
        GradeEncodingSchedule.status == 'active',
        GradeEncodingSchedule.grading_period == 'midterm'
    ).first()
    
    final_schedule = GradeEncodingSchedule.query.filter(
        GradeEncodingSchedule.semester == subject.semester,
        GradeEncodingSchedule.academic_year == academic_year_for_schedule,
        db.or_(
            GradeEncodingSchedule.department == subject.department,
            GradeEncodingSchedule.department == None
        ),
        GradeEncodingSchedule.status == 'active',
        GradeEncodingSchedule.grading_period == 'final'
    ).first()
    
    # Debug: Show which schedules were found
    print(f"   Found schedules:")
    print(f"     All periods schedule: {'Yes' if schedule else 'No'}")
    print(f"     Prelim schedule: {'Yes' if prelim_schedule else 'No'}")
    print(f"     Midterm schedule: {'Yes' if midterm_schedule else 'No'}")
    print(f"     Final schedule: {'Yes' if final_schedule else 'No'}")
    
    # Determine which grading periods are currently open
    # Check if "All Periods" schedule is active
    all_periods_open = schedule and schedule.grading_period == 'all' and schedule.start_date <= now <= schedule.end_date
    
    prelim_open = (prelim_schedule and prelim_schedule.start_date <= now <= prelim_schedule.end_date) or all_periods_open
    midterm_open = (midterm_schedule and midterm_schedule.start_date <= now <= midterm_schedule.end_date) or all_periods_open
    final_open = (final_schedule and final_schedule.start_date <= now <= final_schedule.end_date) or all_periods_open
        
    # Debug information
    if schedule:
        print("Found schedule:")
        print(f"Department: {schedule.department or 'All Departments'}")
        print(f"Period: {schedule.start_date} to {schedule.end_date}")
        print(f"Status: {schedule.status}")
        print(f"Grading Period: {schedule.grading_period}")
        print(f"All Periods Open: {all_periods_open}")
        print(f"Prelim Open: {prelim_open}")
        print(f"Midterm Open: {midterm_open}")
        print(f"Final Open: {final_open}")
    else:
        print("No general schedule found!")

    # Check if encoding is allowed
    can_encode = False
    has_exception = False
    
    # Check if instructor has an encoding exception (for instructors only)
    if current_user.role == 'instructor':
        # Check for 'all' grading periods exception first
        has_all_exception = check_encoding_exception(
            current_user.id,
            academic_year_for_schedule,
            subject.semester,
            'all'
        )
        
        if has_all_exception:
            has_exception = True
            can_encode = True
            # Set all periods as open when 'all' exception is granted
            prelim_open = True
            midterm_open = True
            final_open = True
            flash('You have been granted special access to encode grades for all periods.', 'info')
            print("Encoding ALLOWED via exception - All periods accessible")
        else:
            # Check for specific period exceptions
            has_prelim_exception = check_encoding_exception(
                current_user.id,
                academic_year_for_schedule,
                subject.semester,
                'prelim'
            )
            has_midterm_exception = check_encoding_exception(
                current_user.id,
                academic_year_for_schedule,
                subject.semester,
                'midterm'
            )
            has_final_exception = check_encoding_exception(
                current_user.id,
                academic_year_for_schedule,
                subject.semester,
                'final'
            )
            
            # Set specific periods as open based on exceptions
            if has_prelim_exception:
                prelim_open = True
                has_exception = True
                can_encode = True
                print("Encoding ALLOWED via exception - Prelim period accessible")
            
            if has_midterm_exception:
                midterm_open = True
                has_exception = True
                can_encode = True
                print("Encoding ALLOWED via exception - Midterm period accessible")
            
            if has_final_exception:
                final_open = True
                has_exception = True
                can_encode = True
                print("Encoding ALLOWED via exception - Final period accessible")
            
            if has_exception:
                periods = []
                if has_prelim_exception:
                    periods.append('Prelim')
                if has_midterm_exception:
                    periods.append('Midterm')
                if has_final_exception:
                    periods.append('Final')
                flash(f'You have been granted special access to encode grades for: {", ".join(periods)}', 'info')
                print(f"Encoding ALLOWED via exception - Specific periods: {', '.join(periods)}")
    
    # Registrars always have access
    if current_user.role == 'registrar':
        can_encode = True
        prelim_open = True
        midterm_open = True
        final_open = True
    
    if not has_exception and not can_encode and schedule:
        now = now.replace(tzinfo=None)
        schedule_start = schedule.start_date.replace(tzinfo=None)
        schedule_end = schedule.end_date.replace(tzinfo=None)
        
        print(f"\nChecking dates:")
        print(f"Current time: {now}")
        print(f"Schedule start: {schedule_start}")
        print(f"Schedule end: {schedule_end}")
        
        if schedule_start <= now <= schedule_end:
            can_encode = True
            print("Encoding is ALLOWED")
        else:
            if now < schedule_start:
                flash(f'Grade encoding will start on {schedule_start.strftime("%B %d, %Y")}', 'info')
                print("Schedule hasn't started yet")
            else:
                flash(f'Grade encoding period ended on {schedule_end.strftime("%B %d, %Y")}', 'warning')
                print("Schedule has ended")
    else:
        # Check if there are any active schedules for specific grading periods
        any_active_schedule = GradeEncodingSchedule.query.filter(
            GradeEncodingSchedule.semester == subject.semester,
            GradeEncodingSchedule.academic_year == academic_year_for_schedule,
            db.or_(
                GradeEncodingSchedule.department == subject.department,
                GradeEncodingSchedule.department == None
            ),
            GradeEncodingSchedule.status == 'active'
        ).first()
        
        if any_active_schedule:
            # There's an active schedule but it's for a specific grading period
            # Check if we're within any of the grading period schedules
            if prelim_open or midterm_open or final_open:
                can_encode = True
                print("Encoding is ALLOWED for specific grading period")
            else:
                # Check dates for the active schedule
                now = now.replace(tzinfo=None)
                schedule_start = any_active_schedule.start_date.replace(tzinfo=None)
                schedule_end = any_active_schedule.end_date.replace(tzinfo=None)
                
                if now < schedule_start:
                    flash(f'Grade encoding will start on {schedule_start.strftime("%B %d, %Y")}', 'info')
                else:
                    flash(f'Grade encoding period ended on {schedule_end.strftime("%B %d, %Y")}', 'warning')
        else:
            # Check for any schedule (active, upcoming, or completed)
            any_schedule = GradeEncodingSchedule.query.filter(
                GradeEncodingSchedule.semester == subject.semester,
                GradeEncodingSchedule.academic_year == academic_year_for_schedule,
                db.or_(
                    GradeEncodingSchedule.department == subject.department,
                    GradeEncodingSchedule.department == None
                )
            ).order_by(GradeEncodingSchedule.start_date.asc()).first()
        
            if any_schedule:
                schedule = any_schedule
                if any_schedule.status == 'upcoming':
                    flash(f'Grade encoding is scheduled to start on {any_schedule.start_date.strftime("%B %d, %Y")}', 'info')
                elif any_schedule.status == 'completed':
                    flash(f'The grade encoding period for this semester has ended. Contact MIS/IT if you need to encode grades.', 'warning')
                else:
                    # This should be 'active' status but outside date range
                    now = now.replace(tzinfo=None)
                    schedule_start = any_schedule.start_date.replace(tzinfo=None)
                    schedule_end = any_schedule.end_date.replace(tzinfo=None)
                    
                    if now < schedule_start:
                        flash(f'Grade encoding will start on {schedule_start.strftime("%B %d, %Y")}', 'info')
                    else:
                        flash(f'Grade encoding period ended on {schedule_end.strftime("%B %d, %Y")}', 'warning')
            else:
                flash(f'No grade encoding schedule found for {subject.department} department. Please contact MIS/IT to set up a schedule.', 'warning')
            
            print("\nNo active schedule found")
            print(f"Any schedule found: {'Yes' if any_schedule else 'No'}")
            if any_schedule:
                print(f"Schedule status: {any_schedule.status}")
                print(f"Schedule period: {any_schedule.start_date} to {any_schedule.end_date}")
    
    # Final debug information
    print(f"\n[FINAL] FINAL DECISION:")
    print(f"   can_encode: {can_encode}")
    print(f"   schedule found: {'Yes' if schedule else 'No'}")
    print(f"   prelim_open: {prelim_open}")
    print(f"   midterm_open: {midterm_open}")
    print(f"   final_open: {final_open}")
    print(f"   has_exception: {has_exception}")
    print(f"   Current user role: {current_user.role}")
    
    return render_template('instructor/instructor_encode.html',
                         user=current_user,
                         subject=subject,
                         students=students,
                         grades=grades,
                         can_encode=can_encode,
                         schedule=schedule,
                         prelim_open=prelim_open,
                         midterm_open=midterm_open,
                         final_open=final_open,
                         prelim_schedule=prelim_schedule,
                         midterm_schedule=midterm_schedule,
                         final_schedule=final_schedule,
                         assignment=assignment,
                         is_registrar_access=current_user.role == 'registrar')

@app.route('/api/save-grades', methods=['POST'])
@login_required
def save_grades():
    """Save or update student grades"""
    if current_user.role not in ['instructor', 'registrar']:
        return jsonify({'status': 'error', 'message': 'Unauthorized access'}), 403
    
    # Prevent registrars from saving grades
    if current_user.role == 'registrar':
        return jsonify({'status': 'error', 'message': 'Registrars cannot save grades. This is a read-only view.'}), 403
    
    try:
        data = request.get_json()
        subject_id = data.get('subject_id')
        grades_data = data.get('grades', [])
        
        # Verify access to subject
        subject = Subject.query.get_or_404(subject_id)
        if current_user.role == 'instructor':
            # Check if instructor is assigned to this subject via ClassAssignment
            assignment = ClassAssignment.query.filter_by(
                subject_id=subject_id,
                instructor_id=current_user.id
            ).first()
            if not assignment:
                return jsonify({'status': 'error', 'message': 'You are not assigned to this subject'}), 403
        elif current_user.role == 'registrar':
            # Registrars can access any subject
            pass
        
        # Check if grade encoding is allowed using the same logic as the encode_grades route
        now = datetime.now()
        
        # Determine the correct academic year to use for schedule matching and grade saving
        # Use assignment.school_year if available, otherwise fall back to subject.academic_year
        academic_year_for_grades = subject.academic_year
        if current_user.role == 'instructor' and assignment and assignment.school_year:
            academic_year_for_grades = assignment.school_year
        
        print(f"[DEBUG] Save grades - Schedule matching - Subject ID: {subject_id}")
        print(f"   Subject academic_year: {subject.academic_year}")
        print(f"   Assignment school_year: {assignment.school_year if assignment else 'No assignment'}")
        print(f"   Using academic_year_for_grades: {academic_year_for_grades}")
        
        # First try to find department-specific schedule for all grading periods
        schedule = GradeEncodingSchedule.query.filter(
            GradeEncodingSchedule.semester == subject.semester,
            GradeEncodingSchedule.academic_year == academic_year_for_grades,
            GradeEncodingSchedule.department == subject.department,
            GradeEncodingSchedule.status == 'active',
            GradeEncodingSchedule.grading_period == 'all'
        ).first()
        
        # If no department-specific schedule, look for general schedule for all periods
        if not schedule:
            schedule = GradeEncodingSchedule.query.filter(
                GradeEncodingSchedule.semester == subject.semester,
                GradeEncodingSchedule.academic_year == academic_year_for_grades,
                GradeEncodingSchedule.department == None,
                GradeEncodingSchedule.status == 'active',
                GradeEncodingSchedule.grading_period == 'all'
            ).first()
        
        # Check for specific grading period schedules
        prelim_schedule = GradeEncodingSchedule.query.filter(
            GradeEncodingSchedule.semester == subject.semester,
            GradeEncodingSchedule.academic_year == academic_year_for_grades,
            db.or_(
                GradeEncodingSchedule.department == subject.department,
                GradeEncodingSchedule.department == None
            ),
            GradeEncodingSchedule.status == 'active',
            GradeEncodingSchedule.grading_period == 'prelim'
        ).first()
        
        midterm_schedule = GradeEncodingSchedule.query.filter(
            GradeEncodingSchedule.semester == subject.semester,
            GradeEncodingSchedule.academic_year == academic_year_for_grades,
            db.or_(
                GradeEncodingSchedule.department == subject.department,
                GradeEncodingSchedule.department == None
            ),
            GradeEncodingSchedule.status == 'active',
            GradeEncodingSchedule.grading_period == 'midterm'
        ).first()
        
        final_schedule = GradeEncodingSchedule.query.filter(
            GradeEncodingSchedule.semester == subject.semester,
            GradeEncodingSchedule.academic_year == academic_year_for_grades,
            db.or_(
                GradeEncodingSchedule.department == subject.department,
                GradeEncodingSchedule.department == None
            ),
            GradeEncodingSchedule.status == 'active',
            GradeEncodingSchedule.grading_period == 'final'
        ).first()
        
        # Determine which grading periods are currently open
        # Check if "All Periods" schedule is active
        all_periods_open = schedule and schedule.grading_period == 'all' and schedule.start_date <= now <= schedule.end_date
        
        prelim_open = (prelim_schedule and prelim_schedule.start_date <= now <= prelim_schedule.end_date) or all_periods_open
        midterm_open = (midterm_schedule and midterm_schedule.start_date <= now <= midterm_schedule.end_date) or all_periods_open
        final_open = (final_schedule and final_schedule.start_date <= now <= final_schedule.end_date) or all_periods_open
        
        # Check if encoding is allowed
        can_encode = False
        if schedule:
            now = now.replace(tzinfo=None)
            schedule_start = schedule.start_date.replace(tzinfo=None)
            schedule_end = schedule.end_date.replace(tzinfo=None)
            
            if schedule_start <= now <= schedule_end:
                can_encode = True
        else:
            # Check if there are any active schedules for specific grading periods
            any_active_schedule = GradeEncodingSchedule.query.filter(
                GradeEncodingSchedule.semester == subject.semester,
                GradeEncodingSchedule.academic_year == academic_year_for_grades,
                db.or_(
                    GradeEncodingSchedule.department == subject.department,
                    GradeEncodingSchedule.department == None
                ),
                GradeEncodingSchedule.status == 'active'
            ).first()
            
            if any_active_schedule:
                # Check if we're within any of the grading period schedules
                if prelim_open or midterm_open or final_open:
                    can_encode = True
        
        if not can_encode:
            return jsonify({'status': 'error', 'message': 'Grade encoding is not currently allowed'}), 403
        
        for grade_entry in grades_data:
            student_id = grade_entry.get('student_id')
            
            # Helper to convert grade to float, returns None if invalid
            def to_float(grade_val):
                if grade_val is None or grade_val == '':
                    return None
                try:
                    # Convert to string first to handle any type, then strip whitespace
                    str_val = str(grade_val).strip()
                    if not str_val:
                        return None
                    return float(str_val)
                except (ValueError, TypeError):
                    return None

            prelim = to_float(grade_entry.get('prelim'))
            midterm = to_float(grade_entry.get('midterm'))
            final = to_float(grade_entry.get('final'))
            special_remarks = grade_entry.get('remarks')
            
            # Query for existing grade record
            grade = Grade.query.filter_by(
                student_id=student_id,
                subject_id=subject_id,
                semester=subject.semester,
                academic_year=academic_year_for_grades
            ).first()
            
            # Handle special remarks (AW, UW, INC, Passed)
            if special_remarks and special_remarks in ['AW', 'UW', 'INC', 'Passed']:
                final_average = None
                equivalent_grade = None
                remarks = special_remarks
                is_complete = True  # Special remarks count as complete
            # Calculate final average if all grades are present
            elif prelim is not None and midterm is not None and final is not None:
                final_average = round((prelim + midterm + final) / 3, 2)
                equivalent_grade, remarks = calculate_grade_equivalent(final_average)
                is_complete = True
            else:
                final_average = None
                equivalent_grade = None
                remarks = None
                is_complete = False
            
            if grade:
                grade.prelim_grade = prelim
                grade.midterm_grade = midterm
                grade.final_grade = final
                grade.final_average = final_average
                grade.equivalent_grade = equivalent_grade
                grade.remarks = remarks
                grade.is_complete = is_complete
            else:
                grade = Grade(
                    student_id=student_id,
                    subject_id=subject_id,
                    prelim_grade=prelim,
                    midterm_grade=midterm,
                    final_grade=final,
                    final_average=final_average,
                    equivalent_grade=equivalent_grade,
                    remarks=remarks,
                    semester=subject.semester,
                    academic_year=academic_year_for_grades,
                    is_complete=is_complete
                )
                db.session.add(grade)
        
        db.session.commit()
        
        # Check if all grades are complete and create notifications for students
        all_complete = data.get('all_complete', False)
        
        # Count how many students now have complete grades
        complete_grades_count = Grade.query.filter_by(
            subject_id=subject_id,
            semester=subject.semester,
            academic_year=academic_year_for_grades,
            is_complete=True
        ).count()
        
        # Get total students in the subject
        total_students = Student.query.filter_by(
            department=subject.department,
            year_level=subject.year_level,
            section=subject.section,
            active=True
        ).count()
        
        if all_complete and complete_grades_count == total_students:
            # All students now have complete grades - notify students
            students = Student.query.filter_by(
                department=subject.department,
                year_level=subject.year_level,
                section=subject.section,
                active=True
            ).all()
            
            
            return jsonify({
                'status': 'success', 
                'message': 'Grades Complete! All grades have been entered and are now visible to students.',
                'all_complete': True
            })
        else:
            return jsonify({
                'status': 'success', 
                'message': 'Grades saved successfully',
                'all_complete': False
            })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500



# =====================================
# GRADE MANAGEMENT API
# =====================================

@app.route('/student/grades')
@login_required
def student_grades():
    """Enhanced view for student grades page with complete academic history"""
    if current_user.role != 'student':
        flash('Access denied. Student access only.', 'error')
        return redirect(url_for('dashboard'))
    
    # Get student record - current_user is already a Student object for students
    if current_user.role == 'student':
        student = current_user
    else:
        flash('Access denied. Student access only.', 'error')
        return redirect(url_for('dashboard'))
    
    # Get all complete grades for the student across all academic years
    # Only show grades that are marked as complete (all three components entered)
    all_complete_grades = Grade.query.join(Subject).filter(
        Grade.student_id == current_user.id,
        Grade.is_complete == True
    ).order_by(Grade.academic_year.desc(), Grade.semester, Subject.year_level, Subject.subject_code).all()
    
    # Group grades by academic year and semester
    grades_by_year_semester = {}
    academic_years = set()
    
    for grade in all_complete_grades:
        year_sem_key = f"{grade.academic_year}-{grade.semester}"
        academic_years.add(grade.academic_year)
        
        if year_sem_key not in grades_by_year_semester:
            grades_by_year_semester[year_sem_key] = {
                'academic_year': grade.academic_year,
                'semester': grade.semester,
                'grades': [],
                'total_units': 0,
                'gwa': 0,
                'average_percentage': 0
            }
        
        grades_by_year_semester[year_sem_key]['grades'].append(grade)
    
    # Calculate statistics for each semester and convert Grade objects to dictionaries
    for year_sem_key, data in grades_by_year_semester.items():
        grades = data['grades']
        if grades:
            # Calculate total units
            total_units = sum(grade.subject.units for grade in grades)
            data['total_units'] = total_units
            
            # Calculate GWA (weighted average)
            weighted_sum = sum(grade.subject.units * grade.equivalent_grade for grade in grades if grade.equivalent_grade)
            gwa = weighted_sum / total_units if total_units > 0 else 0
            data['gwa'] = round(gwa, 2)
            
            # Calculate average percentage
            avg_percentage = sum(grade.final_average for grade in grades if grade.final_average) / len(grades) if grades else 0
            data['average_percentage'] = round(avg_percentage, 2)
            
            # Convert Grade objects to serializable dictionaries
            serializable_grades = []
            for grade in grades:
                grade_dict = {
                    'id': grade.id,
                    'student_id': grade.student_id,
                    'subject_id': grade.subject_id,
                    'prelim_grade': grade.prelim_grade,
                    'midterm_grade': grade.midterm_grade,
                    'final_grade': grade.final_grade,
                    'final_average': grade.final_average,
                    'equivalent_grade': grade.equivalent_grade,
                    'remarks': grade.remarks,
                    'semester': grade.semester,
                    'academic_year': grade.academic_year,
                    'is_locked': grade.is_locked,
                    'submitted_at': grade.submitted_at.isoformat() if grade.submitted_at else None,
                    'approved_at': grade.approved_at.isoformat() if grade.approved_at else None,
                    'approved_by': grade.approved_by,
                    'subject': {
                        'id': grade.subject.id,
                        'subject_code': grade.subject.subject_code,
                        'subject_name': grade.subject.subject_name,
                        'units': grade.subject.units,
                        'department': grade.subject.department,
                        'section': grade.subject.section,
                        'year_level': grade.subject.year_level,
                        'semester': grade.subject.semester,
                        'academic_year': grade.academic_year,
                        'instructor_id': grade.subject.instructor_id
                    }
                }
                serializable_grades.append(grade_dict)
            
            # Replace the Grade objects with serializable dictionaries
            data['grades'] = serializable_grades
    
    # Group grades by year level for easier navigation
    grades_by_year_level = {}
    for grade in all_complete_grades:
        year_level = grade.subject.year_level
        if year_level not in grades_by_year_level:
            grades_by_year_level[year_level] = []
        grades_by_year_level[year_level].append(grade)
    
    # Calculate subject counts per academic year
    subjects_per_year = {}
    for grade in all_complete_grades:
        year = grade.academic_year
        if year not in subjects_per_year:
            subjects_per_year[year] = set()
        subjects_per_year[year].add(grade.subject.id)  # Use subject.id to avoid duplicates
    
    # Convert sets to counts
    subjects_per_year = {year: len(subjects) for year, subjects in subjects_per_year.items()}
    
    # Calculate overall statistics
    total_units_all = sum(grade.subject.units for grade in all_complete_grades)
    weighted_sum_all = sum(grade.subject.units * grade.equivalent_grade for grade in all_complete_grades if grade.equivalent_grade)
    overall_gwa = weighted_sum_all / total_units_all if total_units_all > 0 else 0
    
    # Get current semester grades (for dashboard integration)
    current_year = student.academic_year or '2025-2026'
    current_semester = student.semester or 1
    
    current_sem_grades = [grade for grade in all_complete_grades 
                         if grade.academic_year == current_year and grade.semester == current_semester]
    
    return render_template('student/student_grades.html',
                         all_grades=all_complete_grades,
                         grades_by_year_semester=grades_by_year_semester,
                         grades_by_year_level=grades_by_year_level,
                         academic_years=sorted(academic_years, reverse=True),
                         subjects_per_year=subjects_per_year,
                         total_units_all=total_units_all,
                         overall_gwa=round(overall_gwa, 2),
                         current_sem_grades=current_sem_grades,
                         current_year=current_year,
                         current_semester=current_semester)

# =====================================
# PASSWORD RESET ROUTES
# =====================================

@app.route('/registrar/create-subject', methods=['GET', 'POST'])
@login_required
def registrar_subcreate():
    """Create new subject (Registrar only)"""
    if current_user.role != 'registrar':
        flash('Access denied. Only registrars can create subjects.', 'error')
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        try:
            subject_code = request.form.get('subject_code')
            subject_name = request.form.get('subject_name')
            subject_academic = request.form.get('subject_academic')
            units = int(request.form.get('units'))
            department = request.form.get('department')
            year_level = int(request.form.get('year_level'))
            semester = int(request.form.get('semester'))
            
            # Check if subject code already exists
            if Subject.query.filter_by(subject_code=subject_code).first():
                flash('Subject code already exists', 'error')
                return redirect(url_for('registrar_subcreate'))
            
            new_subject = Subject(
                subject_code=subject_code,
                subject_name=subject_name,
                subject_type=subject_academic,
                units=units,
                department=department,
                year_level=year_level,
                semester=semester
            )
            
            db.session.add(new_subject)
            db.session.commit()
            flash('Subject created successfully. Use "Assign Subject" to assign instructors to specific sections.', 'success')
            return redirect(url_for('registrar_subcreate'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error creating subject: {str(e)}', 'error')
    
    # Get all subjects for display
    subjects = Subject.query.all()
    
    # Get all class assignments to check assignment status
    class_assignments = ClassAssignment.query.all()
    
    # Create a set of assigned subject IDs for quick lookup
    assigned_subject_ids = {assignment.subject_id for assignment in class_assignments}
    
    # Get all unique sections from Student table
    sections = db.session.query(Student.section)\
        .filter(Student.section.isnot(None))\
        .distinct()\
        .order_by(Student.section)\
        .all()
    sections = [section[0] for section in sections if section[0]]  # Clean up the results
    
    return render_template('registrar/registrar_subcreate.html', 
                         user=current_user,
                         subjects=subjects,
                         assigned_subject_ids=assigned_subject_ids,
                         sections=sections)

@app.route('/misit/reset-password', methods=['GET', 'POST'])
@login_required
def reset_password_page():
    """Password reset page for MIS/IT admin"""
    if current_user.role != 'mis_it':
        flash('Access denied. Only MIS/IT administrators can reset passwords.', 'error')
        return redirect(url_for('dashboard'))

    # Get staff users (instructors, registrars, deans, MIS/IT)
    users = User.query.all()
    
    # Get students from the separate Student table
    students = Student.query.all()
    
    # Combine both for the template (students will have role='student' added in template)
    return render_template('misit/misit_resetpass.html', users=users, students=students, user=current_user)

@app.route('/api/reset-password', methods=['POST'])
@login_required
def reset_password():
    """API endpoint for password reset - handles both User and Student tables"""
    if current_user.role != 'mis_it':
        return jsonify({'status': 'error', 'message': 'Unauthorized access'}), 403
    
    try:
        data = request.get_json()
        user_id = data.get('user_id')
        new_password = data.get('new_password')
        
        if not user_id or not new_password:
            return jsonify({'status': 'error', 'message': 'Missing required fields'}), 400
        
        # Try to find user in User table first (staff)
        user = User.query.get(user_id)
        if user:
            user.set_password(new_password)
            db.session.commit()
            return jsonify({
                'status': 'success',
                'message': f'Password reset successfully for {user.username}'
            })
        
        # If not found in User table, try Student table
        student = Student.query.get(user_id)
        if student:
            student.set_password(new_password)
            db.session.commit()
            return jsonify({
                'status': 'success',
                'message': f'Student password reset successfully for {student.username}'
            })
        
        # If not found in either table
        return jsonify({'status': 'error', 'message': 'User not found'}), 404
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/delete-student', methods=['POST'])
@login_required
def delete_student():
    """API endpoint for deleting a student - only MIS/IT can delete students"""
    if current_user.role != 'mis_it':
        return jsonify({'status': 'error', 'message': 'Unauthorized access'}), 403
    
    try:
        data = request.get_json()
        student_id = data.get('student_id')
        
        if not student_id:
            return jsonify({'status': 'error', 'message': 'Missing required fields'}), 400
        
        # Find student in Student table
        student = Student.query.get(student_id)
        if not student:
            return jsonify({'status': 'error', 'message': 'Student not found'}), 404
        
        # Store student info for success message
        student_name = f"{student.first_name} {student.last_name}"
        student_username = student.username
        
        # Delete student (related records will be CASCADE deleted automatically)
        db.session.delete(student)
        db.session.commit()
        
        return jsonify({
            'status': 'success',
            'message': f'Student {student_name} ({student_username}) has been deleted successfully'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500

# =====================================
# GRADE SUBMISSION AND APPROVAL ROUTES
# =====================================

@app.route('/api/submit-grades', methods=['POST'])
@login_required
def submit_grades():
    """Submit grades for registrar approval"""
    if current_user.role != 'instructor':
        return jsonify({'status': 'error', 'message': 'Unauthorized access'}), 403

    try:
        data = request.get_json()
        subject_id = data.get('subject_id')
        
        # Verify instructor is assigned to subject
        subject = Subject.query.get_or_404(subject_id)
        if subject.instructor_id != current_user.id:
            return jsonify({'status': 'error', 'message': 'Not assigned to this subject'}), 403

        # Check if grades are already locked
        existing_submission = Grade.query.filter_by(
            subject_id=subject_id,
            is_locked=True
        ).first()
        
        if existing_submission and existing_submission.approved_at is None:
            return jsonify({
                'status': 'error',
                'message': 'Grades already submitted and pending approval'
            }), 400

        # Get and validate all grades
        grades = Grade.query.filter_by(
            subject_id=subject_id,
            semester=subject.semester,
            academic_year=subject.academic_year
        ).all()
        if not grades:
            return jsonify({'status': 'error', 'message': 'No grades found'}), 404

        for grade in grades:
            if None in [grade.prelim_grade, grade.midterm_grade, grade.final_grade]:
                return jsonify({
                    'status': 'error',
                    'message': f'Incomplete grades for {grade.student.first_name} {grade.student.last_name}'
                }), 400

            # Calculate final grades
            grade.final_average = round((grade.prelim_grade + grade.midterm_grade + grade.final_grade) / 3, 2)
            grade.equivalent_grade, grade.remarks = calculate_grade_equivalent(grade.final_average)
            
            # Lock grades
            grade.is_locked = True
            grade.submitted_at = datetime.utcnow()
            grade.approved_at = None
            grade.approved_by = None

        db.session.commit()
        
        
        return jsonify({
            'status': 'success',
            'message': 'Grades submitted successfully and pending approval'
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/approve-grades', methods=['POST'])
@login_required
def approve_grades():
    """Approve submitted grades"""
    print(f"approve_grades called by user: {current_user.username} (role: {current_user.role})")  # Debug log
    
    if current_user.role != 'registrar':
        print("Unauthorized access - user is not registrar")  # Debug log
        return jsonify({'status': 'error', 'message': 'Unauthorized access'}), 403

    try:
        data = request.get_json()
        print(f"Request data: {data}")  # Debug log
        
        grade_ids = data.get('grade_ids', [])
        print(f"Grade IDs to approve: {grade_ids}")  # Debug log
        
        if not grade_ids:
            print("No grade IDs provided")  # Debug log
            return jsonify({'status': 'error', 'message': 'No grades selected'}), 400

        # Verify grades are submitted and not yet approved
        grades = Grade.query.filter(
            Grade.id.in_(grade_ids),
            Grade.is_locked == True,
            Grade.submitted_at != None,
            Grade.approved_at == None
        ).all()
        
        print(f"Found {len(grades)} valid grades for approval")  # Debug log

        if not grades:
            print("No valid grades found for approval")  # Debug log
            return jsonify({'status': 'error', 'message': 'No valid grades for approval'}), 400

        # Approve grades and create notifications for instructors
        now = datetime.utcnow()
        approved_subjects = set()  # Track unique subjects to avoid duplicate notifications
        
        for grade in grades:
            print(f"Approving grade ID {grade.id} for student {grade.student.first_name} {grade.student.last_name}")  # Debug log
            grade.approved_at = now
            grade.approved_by = current_user.id
            approved_subjects.add(grade.subject)

        db.session.commit()
        print("Grades approved successfully in database")  # Debug log
        
        
        print(f"Successfully approved {len(grades)} grades")  # Debug log
        return jsonify({
            'status': 'success',
            'message': f'Successfully approved {len(grades)} grades'
        })

    except Exception as e:
        print(f"Error in approve_grades: {str(e)}")  # Debug log
        db.session.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/unlock-grades', methods=['POST'])
@login_required
def unlock_grades():
    """Unlock grades for re-encoding (MIS/IT only)"""
    if current_user.role != 'mis_it':
        return jsonify({'status': 'error', 'message': 'Unauthorized access'}), 403

    try:
        data = request.get_json()
        subject_id = data.get('subject_id')
        
        # Reset grade status
        grades = Grade.query.filter_by(subject_id=subject_id).all()
        for grade in grades:
            grade.is_locked = False
            grade.submitted_at = None
            grade.approved_at = None
            grade.approved_by = None

        db.session.commit()
        return jsonify({
            'status': 'success',
            'message': 'Grades unlocked for re-encoding'
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'status': 'error', 'message': str(e)}), 500

# =====================================
# NEW ENROLLMENT SYSTEM ROUTES
# =====================================

@app.route('/registrar/enrollment-dashboard')
@login_required
def enrollment_dashboard():
    """Main enrollment dashboard for registrars"""
    if current_user.role != 'registrar':
        flash('Access denied. Only registrars can access this page.', 'error')
        return redirect(url_for('dashboard'))
    
    try:
        # Get enrollment statistics
        pending_students = Student.query.filter_by(enrollment_status='PENDING').count()
        enrolled_students = Student.query.filter_by(enrollment_status='ENROLLED').count()
        
        # Get recent enrollments
        recent_enrollments = StudentEnrollment.query.order_by(StudentEnrollment.enrollment_date.desc()).limit(10).all()
        
        # Get academic years with enrollments
        academic_years = db.session.query(StudentEnrollment.academic_year).distinct().all()
        academic_years = [year[0] for year in academic_years]
        
        return render_template('registrar/enrollment_dashboard.html', 
                             pending_students=pending_students,
                             enrolled_students=enrolled_students,
                             recent_enrollments=recent_enrollments,
                             academic_years=academic_years)
        
    except Exception as e:
        print(f"Error in enrollment_dashboard: {str(e)}")
        flash('An error occurred while loading enrollment dashboard.', 'error')
        return redirect(url_for('registrar_dashboard'))

@app.route('/registrar/pending-enrollments')
@login_required
def pending_enrollments():
    """View students pending enrollment"""
    if current_user.role != 'registrar':
        flash('Access denied. Only registrars can access this page.', 'error')
        return redirect(url_for('dashboard'))
    
    try:
        # Get pending students
        pending_students = Student.query.filter_by(enrollment_status='PENDING').all()
        
        return render_template('registrar/pending_enrollments.html', 
                             pending_students=pending_students)
        
    except Exception as e:
        print(f"Error in pending_enrollments: {str(e)}")
        flash('An error occurred while loading pending enrollments.', 'error')
        return redirect(url_for('enrollment_dashboard'))

@app.route('/registrar/enroll-student/<int:student_id>', methods=['GET', 'POST'])
@login_required
def enroll_student(student_id):
    """Enroll a specific student"""
    if current_user.role != 'registrar':
        flash('Access denied. Only registrars can enroll students.', 'error')
        return redirect(url_for('dashboard'))
    
    student = Student.query.get_or_404(student_id)
    
    if request.method == 'POST':
        try:
            # Get form data with validation
            academic_year = request.form.get('academic_year', '').strip()
            semester_str = request.form.get('semester', '').strip()
            year_level_str = request.form.get('year_level', '').strip()
            section = request.form.get('section', '').strip()
            curriculum = request.form.get('curriculum', '').strip()
            
            # Handle empty section (convert empty string to None)
            if section == '':
                section = None
            
            # Validate required fields
            if not academic_year:
                flash('Academic year is required.', 'error')
                return redirect(url_for('enroll_student', student_id=student_id))
            
            if not semester_str:
                flash('Semester is required.', 'error')
                return redirect(url_for('enroll_student', student_id=student_id))
                
            if not year_level_str:
                flash('Year level is required.', 'error')
                return redirect(url_for('enroll_student', student_id=student_id))
                
            if not curriculum:
                flash('Curriculum is required.', 'error')
                return redirect(url_for('enroll_student', student_id=student_id))
            
            # Convert to integers with error handling
            try:
                semester = int(semester_str)
                year_level = int(year_level_str)
            except ValueError:
                flash('Invalid semester or year level format.', 'error')
                return redirect(url_for('enroll_student', student_id=student_id))
            
            # Validate semester and year level ranges
            if semester not in [1, 2]:
                flash('Semester must be 1 or 2.', 'error')
                return redirect(url_for('enroll_student', student_id=student_id))
                
            if year_level not in [1, 2, 3, 4]:
                flash('Year level must be between 1 and 4.', 'error')
                return redirect(url_for('enroll_student', student_id=student_id))
            
            # Check if student is already enrolled for this academic year/semester
            existing_enrollment = StudentEnrollment.query.filter_by(
                student_id=student_id,
                academic_year=academic_year,
                semester=semester
            ).first()
            
            if existing_enrollment:
                flash('Student is already enrolled for this academic year and semester.', 'error')
                return redirect(url_for('enroll_student', student_id=student_id))
            
            # Create enrollment record
            enrollment = StudentEnrollment(
                student_id=student_id,
                academic_year=academic_year,
                semester=semester,
                year_level=year_level,
                section=section,
                curriculum=curriculum,
                enrolled_by=current_user.id,
                status='ACTIVE'
            )
            
            db.session.add(enrollment)
            
            # Update student status
            student.enrollment_status = 'ENROLLED'
            student.academic_year = academic_year
            student.semester = semester
            student.year_level = year_level
            student.section = section
            student.curriculum = curriculum
            
            db.session.commit()
            
            flash(f'Successfully enrolled {student.first_name} {student.last_name}', 'success')
            return redirect(url_for('pending_enrollments'))
            
        except Exception as e:
            db.session.rollback()
            print(f"Error enrolling student: {str(e)}")
            print(f"Error type: {type(e).__name__}")
            import traceback
            traceback.print_exc()
            flash(f'An error occurred while enrolling the student: {str(e)}', 'error')
            return redirect(url_for('enroll_student', student_id=student_id))
    
    # Get available academic years and semesters
    academic_years = ['2025-2026', '2026-2027', '2027-2028']
    semesters = [1, 2]
    year_levels = [1, 2, 3, 4]
    
    # Get available sections
    sections = Section.query.order_by(Section.section_name).all()
    sections = [section.section_name for section in sections]
    
    # Get available curricula
    curricula = Curriculum.query.order_by(Curriculum.curriculum_name).all()
    curricula = [curriculum.curriculum_name for curriculum in curricula]
    
    # Get available academic years
    academic_years_db = AcademicYear.query.order_by(AcademicYear.academic_year_name).all()
    academic_years = [academic_year.academic_year_name for academic_year in academic_years_db]
    
    return render_template('registrar/enroll_student.html', 
                         student=student,
                         academic_years=academic_years,
                         semesters=semesters,
                         year_levels=year_levels,
                         sections=sections,
                         curricula=curricula)

@app.route('/registrar/assign-subjects/<int:student_id>', methods=['GET', 'POST'])
@login_required
def assign_subjects(student_id):
    """Assign subjects to a student based on curriculum"""
    if current_user.role != 'registrar':
        flash('Access denied. Only registrars can assign subjects.', 'error')
        return redirect(url_for('dashboard'))
    
    student = Student.query.get_or_404(student_id)
    
    # Get student's current enrollment
    current_enrollment = StudentEnrollment.query.filter_by(
        student_id=student_id,
        status='ACTIVE'
    ).first()
    
    if not current_enrollment:
        flash('Student must be enrolled first before assigning subjects.', 'error')
        return redirect(url_for('enroll_student', student_id=student_id))
    
    if request.method == 'POST':
        try:
            subject_ids = request.form.getlist('subject_ids[]')
            
            if not subject_ids:
                flash('Please select at least one subject to assign.', 'error')
                return redirect(url_for('assign_subjects', student_id=student_id))
            
            # Remove existing subject assignments for this student
            StudentSubject.query.filter_by(
                student_id=student_id
            ).delete()
            
            # Add new subject assignments
            for subject_id in subject_ids:
                # Get the subject to use its academic_year and semester
                subject = Subject.query.get(int(subject_id))
                if subject:
                    print(f"[DEBUG] Assigning subject {subject.subject_code} to student {student_id}")
                    print(f"   Subject academic_year: {subject.academic_year}")
                    print(f"   Subject semester: {subject.semester}")
                    
                    # Check if student is already enrolled in this subject
                    existing_assignment = StudentSubject.query.filter_by(
                        student_id=student_id,
                        subject_id=int(subject_id)
                    ).first()
                    
                    if existing_assignment:
                        # Update existing assignment with correct semester and academic year
                        print(f"   Updating existing assignment")
                        existing_assignment.academic_year = subject.academic_year
                        existing_assignment.semester = subject.semester
                        existing_assignment.status = 'ENROLLED'
                        existing_assignment.enrolled_by = current_user.id
                    else:
                        # Create new assignment
                        print(f"   Creating new assignment")
                        student_subject = StudentSubject(
                            student_id=student_id,
                            subject_id=int(subject_id),
                            academic_year=subject.academic_year,
                            semester=subject.semester,
                            enrolled_by=current_user.id,
                            status='ENROLLED'
                        )
                        db.session.add(student_subject)
            
            db.session.commit()
            
            flash(f'Successfully assigned {len(subject_ids)} subjects to {student.first_name} {student.last_name}', 'success')
            return redirect(url_for('enrollment_dashboard'))
            
        except Exception as e:
            db.session.rollback()
            print(f"Error assigning subjects: {str(e)}")
            print(f"Error type: {type(e).__name__}")
            flash(f'An error occurred while assigning subjects: {str(e)}', 'error')
            return redirect(url_for('assign_subjects', student_id=student_id))
    
    # Get filter parameters
    department_filter = request.args.get('department', '') or None
    section_filter = request.args.get('section', '') or None
    year_level_filter = request.args.get('year_level', '') or None
    semester_filter = request.args.get('semester', '') or None
    
    # Only show subjects if at least one filter is applied
    filtered_subjects = []
    if department_filter or section_filter or year_level_filter or semester_filter:
        # Build query with filters
        query = Subject.query
        
        if department_filter:
            query = query.filter(Subject.department == department_filter)
        if year_level_filter:
            query = query.filter(Subject.year_level == int(year_level_filter))
        if semester_filter:
            query = query.filter(Subject.semester == int(semester_filter))
        
        # Get filtered subjects
        filtered_subjects = query.all()
    
    # Get all subjects for filter options
    all_subjects = Subject.query.all()
    
    # Get unique departments for filter dropdown
    departments = db.session.query(Subject.department).distinct().order_by(Subject.department).all()
    departments = [dept[0] for dept in departments]
    
    # Get unique sections from students table
    sections_query = db.session.query(Student.section).filter(
        Student.section.isnot(None),
        Student.section != ''
    ).distinct().order_by(Student.section).all()
    sections = [section[0] for section in sections_query]
    
    # Get already assigned subjects - use the subject's academic_year and semester
    assigned_subjects = StudentSubject.query.filter_by(
        student_id=student_id,
        status='ENROLLED'
    ).all()
    
    # Filter by matching academic_year and semester with the subjects
    assigned_subject_ids = []
    assigned_subjects_list = []  # For the "Currently Assigned Subjects" section
    for ss in assigned_subjects:
        subject = Subject.query.get(ss.subject_id)
        if subject and subject.academic_year == ss.academic_year and subject.semester == ss.semester:
            assigned_subject_ids.append(ss.subject_id)
            assigned_subjects_list.append(subject)  # Add the full subject object
    
    return render_template('registrar/assign_subjects.html',
                         student=student,
                         current_enrollment=current_enrollment,
                         all_subjects=filtered_subjects if filtered_subjects else all_subjects,
                         assigned_subject_ids=assigned_subject_ids,
                         assigned_subjects_list=assigned_subjects_list,
                         departments=departments,
                         sections=sections,
                         department_filter=department_filter,
                         section_filter=section_filter,
                         year_level_filter=year_level_filter,
                         semester_filter=semester_filter)



# =====================================
# API ROUTES
# =====================================

@app.route('/api/create-section', methods=['POST'])
@login_required
def create_section():
    """Create a new section"""
    if current_user.role != 'registrar':
        return jsonify({'error': 'Access denied. Only registrars can create sections.'}), 403
    
    try:
        data = request.get_json()
        section_name = data.get('section_name', '').strip()
        
        if not section_name:
            return jsonify({'error': 'Section name is required'}), 400
        
        # Check if section already exists
        existing_section = Section.query.filter_by(section_name=section_name).first()
        if existing_section:
            return jsonify({'error': 'This section already exists'}), 400
        
        # Create new section
        section = Section(
            section_name=section_name,
            created_by=current_user.id
        )
        
        db.session.add(section)
        db.session.commit()
        
        return jsonify({
            'status': 'success',
            'message': f'Section "{section_name}" created successfully',
            'section': {
                'id': section.id,
                'name': section.section_name
            }
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to create section: {str(e)}'}), 500

@app.route('/api/create-curriculum', methods=['POST'])
@login_required
def create_curriculum():
    """Create a new curriculum"""
    if current_user.role != 'registrar':
        return jsonify({'error': 'Access denied. Only registrars can create curricula.'}), 403
    
    try:
        data = request.get_json()
        curriculum_name = data.get('curriculum_name', '').strip()
        curriculum_description = data.get('curriculum_description', '').strip()
        
        if not curriculum_name:
            return jsonify({'error': 'Curriculum name is required'}), 400
        
        # Check if curriculum already exists
        existing_curriculum = Curriculum.query.filter_by(curriculum_name=curriculum_name).first()
        if existing_curriculum:
            return jsonify({'error': 'This curriculum already exists'}), 400
        
        # Create new curriculum
        curriculum = Curriculum(
            curriculum_name=curriculum_name,
            curriculum_description=curriculum_description,
            created_by=current_user.id
        )
        
        db.session.add(curriculum)
        db.session.commit()
        
        return jsonify({
            'status': 'success',
            'message': f'Curriculum "{curriculum_name}" created successfully',
            'curriculum': {
                'id': curriculum.id,
                'name': curriculum.curriculum_name,
                'description': curriculum.curriculum_description
            }
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to create curriculum: {str(e)}'}), 500

@app.route('/api/create-academic-year', methods=['POST'])
@login_required
def create_academic_year():
    """Create a new academic year"""
    if current_user.role != 'registrar':
        return jsonify({'error': 'Access denied. Only registrars can create academic years.'}), 403
    
    try:
        data = request.get_json()
        academic_year_name = data.get('academic_year_name', '').strip()
        
        if not academic_year_name:
            return jsonify({'error': 'Academic year name is required'}), 400
        
        # Check if academic year already exists
        existing_academic_year = AcademicYear.query.filter_by(academic_year_name=academic_year_name).first()
        if existing_academic_year:
            return jsonify({'error': 'This academic year already exists'}), 400
        
        # Create new academic year
        academic_year = AcademicYear(
            academic_year_name=academic_year_name,
            created_by=current_user.id
        )
        
        db.session.add(academic_year)
        db.session.commit()
        
        return jsonify({
            'status': 'success',
            'message': f'Academic year "{academic_year_name}" created successfully',
            'academic_year': {
                'id': academic_year.id,
                'name': academic_year.academic_year_name
            }
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to create academic year: {str(e)}'}), 500

@app.route('/api/toggle-theme', methods=['POST'])
@login_required
def toggle_theme():
    """Toggle between dark and light theme"""
    data = request.get_json() or {}
    theme = data.get('theme', 'light')
    session['theme'] = theme
    return jsonify({'status': 'success', 'theme': theme})

# =====================================
# DATABASE INITIALIZATION
# =====================================

def update_database_schema():
    """Database schema update - DISABLED FOR CLEAN DATABASE STRUCTURE"""
    print("Database schema update disabled - using clean structure")
    print("User table contains ONLY authentication and basic info")
    print("Student data is stored in dedicated 'students' table")
    return
    
    # ORIGINAL FUNCTION COMPLETELY DISABLED
    # This function was adding student-specific fields back to the user table
    # which conflicts with our clean database structure where:
    # - User table = Authentication + Basic info only
    # - Students table = All student-specific data

def init_database():
    """Initialize database with tables and demo data"""
    try:
        # Create all tables
        db.create_all()
        
        # Update existing tables with new columns - DISABLED (database already cleaned)
        # update_database_schema()
        
        # Create demo accounts
        create_demo_accounts()
        print("Database initialized successfully!")
        return True
    except Exception as e:
        print(f"Database initialization error: {e}")
        return False

# =====================================
# APPLICATION STARTUP
# =====================================

if __name__ == '__main__':
    # Initialize database
    with app.app_context():
        init_database()
    
    # Run the application
    app.run(host='0.0.0.0', port=5000, debug=True)
